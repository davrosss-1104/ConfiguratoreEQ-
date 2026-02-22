from fastapi import FastAPI, Depends, HTTPException, Query, Header, UploadFile, File as FastAPIFile, Form, Request
from fastapi import FastAPI, Depends, HTTPException, Query, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, or_
from typing import List, Optional
from datetime import datetime, timedelta
import json
import os
import io
import math
import re

import base64
from template_engine import (
    DocumentTemplate, STATIC_SECTIONS,
    get_available_fields_from_db, get_default_template_config_from_db,
    genera_docx_da_template, load_valori_dinamici, load_defaults_info
)

from database import engine, SessionLocal
from models import (
    Base, Preventivo, DatiCommessa, DatiPrincipali, Normative,
    DisposizioneVano, Porte, Materiale, ProductTemplate,
    Utente, GruppoUtenti, PermessoGruppo, Ruolo, PermessoRuolo,
    Articolo, CategoriaArticoli, Cliente,
    RigaRicambio, Argano, ParametriSistema, BomStruttura
)
from schemas import (
    PreventivoCreate, PreventivoUpdate, Preventivo as PreventivoSchema,
    DatiCommessaCreate, DatiCommessaUpdate, DatiCommessa as DatiCommessaSchema,
    DatiPrincipaliCreate, DatiPrincipaliUpdate, DatiPrincipali as DatiPrincipaliSchema,
    NormativeCreate, NormativeUpdate, Normative as NormativeSchema,
    DisposizioneVanoCreate, DisposizioneVanoUpdate, DisposizioneVano as DisposizioneVanoSchema,
    PorteCreate, PorteUpdate, Porte as PorteSchema,
    MaterialeCreate, MaterialeUpdate, Materiale as MaterialeSchema,
    ProductTemplateCreate, ProductTemplateUpdate, ProductTemplate as ProductTemplateSchema
)
from auth import (
    authenticate_user, create_access_token, get_password_hash,
    get_user_from_token, create_default_admin,
    get_user_permissions,
    ACCESS_TOKEN_EXPIRE_MINUTES
)
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
import json as json_module
from excel_data_loader import ExcelDataLoader
    

# Custom JSON encoder per SQLAlchemy objects
class SQLAlchemyEncoder(json_module.JSONEncoder):
    def default(self, obj):
        if hasattr(obj, '__table__'):
            row = {}
            for col in obj.__table__.columns:
                val = getattr(obj, col.name, None)
                if hasattr(val, 'isoformat'):
                    val = val.isoformat()
                row[col.name] = val
            return row
        if hasattr(obj, 'isoformat'):
            return obj.isoformat()
        return super().default(obj)

def _orm_to_dict(obj):
    """Converte un oggetto ORM in dict serializzabile"""
    if obj is None:
        return None
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, list):
        return [_orm_to_dict(item) for item in obj]
    if hasattr(obj, '__table__'):
        row = {}
        for col in obj.__table__.columns:
            val = getattr(obj, col.name, None)
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            row[col.name] = val
        return row
    return obj

# Crea tabelle
Base.metadata.create_all(bind=engine)

# FastAPI app
app = FastAPI(title="Configuratore Elettroquadri API", version="0.11.0")

# ==========================================
# CORS CONFIGURATION
# ==========================================
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "http://127.0.0.1:5173",
        "http://127.0.0.1:3000"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# DEPENDENCY
# ==========================================
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ==========================================
# HELPER: Accesso sicuro ai campi Preventivo
# Nomi reali da models.py: stato, numero, tipo, totale_materiali, totale_netto
# ==========================================
def _prev_get(preventivo, *names, default=None):
    """Cerca il primo attributo esistente tra i nomi forniti"""
    for name in names:
        val = getattr(preventivo, name, None)
        if val is not None:
            return val
    return default

def _prev_set(preventivo, value, *names):
    """Setta il primo attributo esistente tra i nomi forniti"""
    for name in names:
        if hasattr(preventivo, name):
            setattr(preventivo, name, value)
            return True
    setattr(preventivo, names[0], value)
    return False

# Shortcuts - nomi REALI prima, fallback dopo
def _prev_stato(p):      return _prev_get(p, 'stato', 'status', default='bozza')
def _prev_numero(p):     return _prev_get(p, 'numero', 'numero_preventivo', default='')
def _prev_tipo(p):       return _prev_get(p, 'tipo', 'tipo_preventivo', default='')
def _prev_totale(p):     return _prev_get(p, 'totale_materiali', 'total_price', default=0)
def _prev_netto(p):      return _prev_get(p, 'totale_netto', 'total_price_finale', default=0)
def _prev_categoria(p):  return _prev_get(p, 'categoria', default='')

def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


# ==========================================
# RULE ENGINE - VERSIONE MIGLIORATA
# ==========================================
def load_rules():
    """Carica le regole dai file JSON nella directory rules/"""
    rules = []
    rules_dir = "./rules"
    
    if not os.path.exists(rules_dir):
        print("Ã¢Å¡Â Ã¯Â¸Â Directory rules/ non trovata")
        return rules
    
    for filename in os.listdir(rules_dir):
        if filename.endswith(".json"):
            try:
                with open(os.path.join(rules_dir, filename), "r", encoding="utf-8") as f:
                    rule = json.load(f)
                    rules.append(rule)
                    print(f"Ã¢Å“â€¦ Regola caricata: {filename}")
            except Exception as e:
                print(f"Ã¢ÂÅ’ Errore caricamento {filename}: {e}")
    
    print(f"Ã°Å¸â€œâ€¹ Caricate {len(rules)} regole")
    return rules

def evaluate_condition(condition: dict, config_data: dict) -> bool:
    """Valuta una singola condizione con gestione robusta di NULL/None e tipi misti"""
    if isinstance(condition, str):
        return False
    
    field = condition.get("field")
    operator = condition.get("operator")
    expected_value = condition.get("value")
    
    if field not in config_data:
        return False
    
    config_value = config_data.get(field)
    
    if config_value is None:
        return False
    
    if isinstance(config_value, str) and config_value.strip() == "":
        return False
    
    result = False
    
    if operator == "equals":
        # Confronto robusto: prova prima diretto, poi stringa vs stringa
        result = config_value == expected_value
        if not result:
            result = str(config_value).strip().lower() == str(expected_value).strip().lower()
    elif operator == "not_equals":
        result = str(config_value).strip().lower() != str(expected_value).strip().lower()
    elif operator == "contains":
        result = str(expected_value).lower() in str(config_value).lower() if config_value else False
    elif operator in ("greater_than", "gt"):
        try:
            result = float(config_value) > float(expected_value)
        except (ValueError, TypeError):
            result = False
    elif operator in ("greater_equal", "gte", "greater_than_or_equal"):
        try:
            result = float(config_value) >= float(expected_value)
        except (ValueError, TypeError):
            result = False
    elif operator in ("less_than", "lt"):
        try:
            result = float(config_value) < float(expected_value)
        except (ValueError, TypeError):
            result = False
    elif operator in ("less_equal", "lte", "less_than_or_equal"):
        try:
            result = float(config_value) <= float(expected_value)
        except (ValueError, TypeError):
            result = False
    elif operator == "in":
        if isinstance(expected_value, list):
            cv_lower = str(config_value).strip().lower()
            result = cv_lower in [str(v).strip().lower() for v in expected_value]
        else:
            result = False
    elif operator == "not_in":
        if isinstance(expected_value, list):
            cv_lower = str(config_value).strip().lower()
            result = cv_lower not in [str(v).strip().lower() for v in expected_value]
        else:
            result = True
    elif operator == "is_true":
        result = str(config_value).strip().lower() in ("true", "1", "si", "yes")
    elif operator == "is_false":
        result = str(config_value).strip().lower() in ("false", "0", "no", "")
    
    return result

def build_config_context(preventivo_id: int, db: Session) -> dict:
    """
    Costruisce il contesto COMPLETO per la valutazione regole,
    leggendo da TUTTE le fonti dati del preventivo.
    
    Fonti (in ordine di priorita, le ultime sovrascrivono):
      0. defaults da campi_configuratore (base, priorità minima)
      1. metadati preventivo (tipo, categoria)
      2. dati_principali (tabella ORM)
      3. normative (tabella ORM)
      4. argano (query diretta)
      5. porte (query diretta)
      6. disposizione_vano (query diretta)
      7. valori_configurazione (tabella chiave/valore per campi dinamici)
    """
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        return {}
    
    config_data = {}
    
    # 0. DEFAULTS da campi_configuratore (priorità più bassa, base per tutte le sezioni)
    #    Garantisce che i valori default siano disponibili anche per sezioni non ancora visitate
    try:
        defaults_rows = db.execute(
            text("SELECT codice, valore_default FROM campi_configuratore WHERE attivo=1 AND valore_default IS NOT NULL AND valore_default != '-'")
        ).fetchall()
        for codice, valore_default in defaults_rows:
            if codice and valore_default:
                config_data[codice] = valore_default
    except Exception:
        pass
    
    # 1. Metadati preventivo
    config_data["preventivo_tipo"] = _prev_tipo(preventivo)
    config_data["preventivo_categoria"] = _prev_categoria(preventivo)
    
    # 1b. Info template prodotto (per regole TEMPLATE_BASE)
    try:
        tmpl_id = getattr(preventivo, 'template_id', None)
        if tmpl_id:
            tmpl = db.query(ProductTemplate).filter(ProductTemplate.id == tmpl_id).first()
            if tmpl:
                config_data["template_id"] = tmpl.id
                config_data["template_categoria"] = tmpl.categoria
                config_data["template_sottocategoria"] = tmpl.sottocategoria
                config_data["template_nome"] = tmpl.nome_display
    except Exception:
        pass
    
    # 2. Dati principali (query diretta, NO relationship)
    try:
        dp = db.query(DatiPrincipali).filter(DatiPrincipali.preventivo_id == preventivo_id).first()
        if dp:
            config_data.update({
                "tipo_impianto": dp.tipo_impianto,
                "nuovo_impianto": dp.nuovo_impianto,
                "numero_fermate": dp.numero_fermate,
                "numero_servizi": dp.numero_servizi,
                "velocita": dp.velocita,
                "corsa": dp.corsa,
                "con_locale_macchina": dp.con_locale_macchina,
                "posizione_locale_macchina": dp.posizione_locale_macchina,
                "tipo_trazione": dp.tipo_trazione,
                "forza_motrice": dp.forza_motrice,
                "luce": dp.luce,
                "tensione_manovra": dp.tensione_manovra,
                "tensione_freno": dp.tensione_freno,
            })
    except Exception:
        pass
    
    # 3. Normative (query diretta, NO relationship)
    try:
        norm = db.query(Normative).filter(Normative.preventivo_id == preventivo_id).first()
        if norm:
            config_data.update({
                "en_81_1": norm.en_81_1,
                "en_81_20": norm.en_81_20,
                "en_81_21": norm.en_81_21,
                "en_81_28": norm.en_81_28,
                "en_81_70": norm.en_81_70,
                "en_81_72": norm.en_81_72,
                "en_81_73": norm.en_81_73,
                "a3_95_16": norm.a3_95_16,
                "dm236_legge13": norm.dm236_legge13,
                "emendamento_a3": norm.emendamento_a3,
                "uni_10411_1": norm.uni_10411_1,
            })
    except Exception:
        pass
    
    # 4. Argano (non ha relazione ORM, query diretta)
    try:
        argano = db.query(Argano).filter(Argano.preventivo_id == preventivo_id).first()
        if argano:
            config_data.update({
                "trazione": argano.trazione,
                "potenza_motore_kw": argano.potenza_motore_kw,
                "corrente_nom_motore_amp": argano.corrente_nom_motore_amp,
                "tipo_vvvf": argano.tipo_vvvf,
                "vvvf_nel_vano": argano.vvvf_nel_vano,
                "freno_tensione": argano.freno_tensione,
                "ventilazione_forzata": argano.ventilazione_forzata,
                "tipo_teleruttore": argano.tipo_teleruttore,
            })
    except Exception:
        pass
    
    # 5. Porte (query diretta, leggi tutti gli attributi)
    try:
        porte = db.query(Porte).filter(Porte.preventivo_id == preventivo_id).first()
        if porte:
            skip = {'id', 'preventivo_id', 'preventivo', 'metadata', 'registry',
                    '_sa_instance_state', '_sa_class_manager'}
            for col in porte.__table__.columns:
                if col.name not in ('id', 'preventivo_id'):
                    val = getattr(porte, col.name, None)
                    if val is not None:
                        config_data[col.name] = val
    except Exception:
        pass
    
    # 6. Disposizione vano (query diretta)
    try:
        dv = db.query(DisposizioneVano).filter(
            DisposizioneVano.preventivo_id == preventivo_id
        ).first()
        if dv:
            for col in dv.__table__.columns:
                if col.name not in ('id', 'preventivo_id'):
                    val = getattr(dv, col.name, None)
                    if val is not None:
                        config_data[col.name] = val
    except Exception:
        pass
    
    # 7. Valori configurazione (tabella chiave/valore, massima priorita)
    try:
        result = db.execute(
            text("SELECT codice_campo, valore FROM valori_configurazione WHERE preventivo_id = :pid"),
            {"pid": preventivo_id}
        )
        for row in result.fetchall():
            campo, valore = row[0], row[1]
            if campo and valore is not None:
                config_data[campo] = valore
    except Exception:
        pass  # Tabella potrebbe non esistere
    
    # Pulizia: rimuovi None
    config_data = {k: v for k, v in config_data.items() if v is not None}
    
    return config_data


def evaluate_rules(preventivo_id: int, db: Session):
    """
    Valuta le regole JSON per un preventivo.
    
    1. Costruisce contesto da TUTTE le sezioni (ORM + valori_configurazione + JSON)
    2. Carica regole da file JSON (./rules/)
    3. Per ogni regola: valuta condizioni â†’ se tutte OK â†’ aggiunge materiali
    4. Rimuove materiali orfani (regole non piu attive)
    
    Returns: dict con risultato valutazione
    """
    db.expire_all()
    
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        return {"error": "Preventivo non trovato"}
    
    # 1. Costruisci contesto completo
    config_data = build_config_context(preventivo_id, db)
    
    # 2. Carica regole da file JSON
    rules = load_rules()
    
    if not rules:
        return {
            "status": "warning",
            "message": "Nessuna regola trovata in ./rules/",
            "materiali_aggiunti": 0,
            "materiali_rimossi": 0,
            "regole_attive": [],
            "context_keys": list(config_data.keys())
        }
    
    # 3. Valuta regole
    active_rules = set()
    materials_to_add = []
    
    for rule in rules:
        if not rule.get("enabled", True):
            continue
        
        rule_id = rule.get("id", "unknown")
        conditions = rule.get("conditions", [])
        
        # Valuta tutte le condizioni (AND implicito)
        all_met = True
        for condition in conditions:
            if not evaluate_condition(condition, config_data):
                all_met = False
                break
        
        if all_met and conditions:
            active_rules.add(rule_id)
            for material in rule.get("materials", []):
                materials_to_add.append({
                    "rule_id": rule_id,
                    "codice": material.get("codice"),
                    "descrizione": material.get("descrizione"),
                    "quantita": material.get("quantita", 1),
                    "prezzo_unitario": material.get("prezzo_unitario", 0.0),
                    "categoria": material.get("categoria", "Materiale Automatico"),
                    "note": material.get("note", ""),
                })
    
    # 4. Lookup prezzi da tabella articoli per materiali senza prezzo
    codici_senza_prezzo = set(
        m["codice"] for m in materials_to_add
        if not m.get("prezzo_unitario") and m.get("codice")
    )
    prezzi_articoli = {}
    if codici_senza_prezzo:
        articoli_trovati = db.query(Articolo).filter(
            Articolo.codice.in_(codici_senza_prezzo)
        ).all()
        for art in articoli_trovati:
            costo_base = art.costo_fisso or 0
            ricarico = art.ricarico_percentuale or 0
            prezzo = costo_base * (1 + ricarico / 100) if ricarico else costo_base
            prezzi_articoli[art.codice] = round(prezzo, 4)
            
        # Fallback: cerca anche in articoli_bom (tabella legacy)
        codici_mancanti = codici_senza_prezzo - set(prezzi_articoli.keys())
        if codici_mancanti:
            try:
                placeholders = ",".join(f":c{i}" for i in range(len(codici_mancanti)))
                params = {f"c{i}": c for i, c in enumerate(codici_mancanti)}
                result = db.execute(text(
                    f"SELECT codice, costo_fisso, costo_variabile FROM articoli_bom WHERE codice IN ({placeholders})"
                ), params)
                for row in result.fetchall():
                    costo = (row[1] or 0) + (row[2] or 0)
                    if costo > 0:
                        prezzi_articoli[row[0]] = round(costo, 4)
            except Exception:
                pass  # tabella articoli_bom potrebbe non esistere

    # Aggiorna prezzi nei materiali da aggiungere
    for mat_data in materials_to_add:
        if not mat_data.get("prezzo_unitario") and mat_data["codice"] in prezzi_articoli:
            mat_data["prezzo_unitario"] = prezzi_articoli[mat_data["codice"]]

    # 5. Rimozione orfani: materiali da regole non piu attive
    materiali_rimossi = 0
    existing_auto = db.query(Materiale).filter(
        Materiale.preventivo_id == preventivo_id,
        Materiale.aggiunto_da_regola == True
    ).all()
    
    for mat in existing_auto:
        if mat.regola_id and mat.regola_id not in active_rules:
            db.delete(mat)
            materiali_rimossi += 1
    
    # 6. Aggiungi nuovi materiali (evita duplicati) + aggiorna prezzi mancanti
    materiali_aggiunti = 0
    materiali_prezzo_aggiornato = 0
    for mat_data in materials_to_add:
        existing = db.query(Materiale).filter(
            Materiale.preventivo_id == preventivo_id,
            Materiale.codice == mat_data["codice"],
            Materiale.regola_id == mat_data["rule_id"]
        ).first()
        
        if existing:
            # Aggiorna prezzo se era 0 e ora abbiamo un prezzo valido
            if (not existing.prezzo_unitario or existing.prezzo_unitario == 0) and mat_data["prezzo_unitario"]:
                existing.prezzo_unitario = mat_data["prezzo_unitario"]
                existing.prezzo_totale = existing.quantita * mat_data["prezzo_unitario"]
                materiali_prezzo_aggiornato += 1
        else:
            qta = mat_data["quantita"]
            prezzo = mat_data["prezzo_unitario"]
            nuovo = Materiale(
                preventivo_id=preventivo_id,
                codice=mat_data["codice"],
                descrizione=mat_data["descrizione"],
                quantita=qta,
                prezzo_unitario=prezzo,
                prezzo_totale=qta * prezzo,
                categoria=mat_data["categoria"],
                aggiunto_da_regola=True,
                regola_id=mat_data["rule_id"],
                note=mat_data.get("note", ""),
            )
            db.add(nuovo)
            materiali_aggiunti += 1
    
    # 7. Ricalcola totale preventivo
    if materiali_aggiunti > 0 or materiali_rimossi > 0 or materiali_prezzo_aggiornato > 0:
        db.commit()
        tutti_materiali = db.query(Materiale).filter(
            Materiale.preventivo_id == preventivo_id
        ).all()
        totale_calc = sum(m.prezzo_totale or 0 for m in tutti_materiali)
        _prev_set(preventivo, totale_calc, "totale_materiali", "total_price")
        db.commit()
    
    return {
        "status": "ok",
        "materiali_aggiunti": materiali_aggiunti,
        "materiali_rimossi": materiali_rimossi,
        "materiali_prezzo_aggiornato": materiali_prezzo_aggiornato,
        "regole_attive": list(active_rules),
        "regole_totali": len(rules),
        "prezzi_da_articoli": len(prezzi_articoli),
        "codici_senza_prezzo": list(codici_senza_prezzo - set(prezzi_articoli.keys())),
        "context_keys": list(config_data.keys()),
    }


def safe_evaluate_rules(preventivo_id: int, db: Session):
    """Wrapper che non crasha mai - logga errori ma non blocca il chiamante"""
    try:
        return evaluate_rules(preventivo_id, db)
    except Exception as e:
        import traceback
        print(f"[RULE ENGINE ERROR] preventivo {preventivo_id}: {e}")
        traceback.print_exc()
        return {"error": str(e)}



# ============================================================
# ORDINI E BOM - FUNZIONI HELPER
# ============================================================

def calcola_lead_time(preventivo_id: int, db: Session) -> int:
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    materiali = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
    if not materiali:
        return 15
    max_lt = 0
    for mat in materiali:
        cursor.execute("SELECT lead_time_produzione, lead_time_acquisto, tipo FROM articoli_bom WHERE codice = ?", (mat.codice,))
        row = cursor.fetchone()
        if row:
            lt = row[0] if row[2] in ("MASTER", "SEMILAVORATO") else row[1]
            max_lt = max(max_lt, lt or 0)
    return max_lt + 5 + 3


def esplodi_bom_ricorsiva(cursor, codice_padre, qta_padre, contesto, ordine_id,
                           livello=0, percorso="", risultati=None):
    if risultati is None:
        risultati = []
    if livello > 10:
        return risultati

    cursor.execute(
        "SELECT bs.figlio_codice, bs.quantita, bs.formula_quantita, bs.condizione_esistenza, "
        "ab.descrizione, ab.tipo, ab.categoria, ab.costo_fisso, ab.costo_variabile, "
        "ab.unita_misura, ab.lead_time_produzione, ab.lead_time_acquisto, "
        "ab.parametro1_nome, ab.parametro2_nome, ab.parametro3_nome "
        "FROM bom_struttura bs "
        "JOIN articoli_bom ab ON ab.codice = bs.figlio_codice "
        "WHERE bs.padre_codice = ? ORDER BY bs.posizione",
        (codice_padre,)
    )
    figli = cursor.fetchall()

    if not figli:
        # LEAF
        cursor.execute(
            "SELECT descrizione, tipo, categoria, costo_fisso, costo_variabile, "
            "unita_misura, lead_time_produzione, lead_time_acquisto, "
            "parametro1_nome, parametro2_nome, parametro3_nome "
            "FROM articoli_bom WHERE codice = ?",
            (codice_padre,)
        )
        art = cursor.fetchone()
        if art:
            costo_unit = (art[3] or 0) + (art[4] or 0)
            param_vals = {}
            for i, pname in enumerate(art[8:11]):
                if pname:
                    val = contesto.get(pname, contesto.get(pname.upper(), ""))
                    param_vals[f"parametro{i+1}_nome"] = pname
                    param_vals[f"parametro{i+1}_valore"] = str(val) if val else ""
            risultati.append({
                "ordine_id": ordine_id, "codice": codice_padre,
                "descrizione": art[0], "tipo": art[1], "categoria": art[2],
                "quantita": qta_padre, "unita_misura": art[5] or "PZ",
                "costo_unitario": costo_unit, "costo_totale": costo_unit * qta_padre,
                "padre_codice": percorso.split(" > ")[-2] if " > " in percorso else None,
                "livello_esplosione": livello, "percorso": percorso,
                "lead_time_giorni": art[6] if art[1] != "ACQUISTO" else art[7],
                **param_vals
            })
        return risultati

    for figlio in figli:
        (f_cod, f_qta, f_formula, f_cond, f_desc, f_tipo, f_cat,
         f_cf, f_cv, f_um, f_lt_p, f_lt_a, f_p1, f_p2, f_p3) = figlio

        if f_cond and not _valuta_condizione_bom(f_cond, contesto):
            continue

        qta = f_qta
        if f_formula:
            qta_calc = _calcola_formula(f_formula, contesto)
            if qta_calc is not None:
                qta = qta_calc

        qta_tot = qta * qta_padre
        nuovo_percorso = f"{percorso} > {f_cod}" if percorso else f_cod

        if f_tipo == "ACQUISTO":
            costo_unit = (f_cf or 0) + (f_cv or 0)
            param_vals = {}
            for i, pname in enumerate([f_p1, f_p2, f_p3]):
                if pname:
                    val = contesto.get(pname, contesto.get(pname.upper(), ""))
                    param_vals[f"parametro{i+1}_nome"] = pname
                    param_vals[f"parametro{i+1}_valore"] = str(val) if val else ""
            risultati.append({
                "ordine_id": ordine_id, "codice": f_cod,
                "descrizione": f_desc, "tipo": f_tipo, "categoria": f_cat,
                "quantita": qta_tot, "unita_misura": f_um or "PZ",
                "costo_unitario": costo_unit, "costo_totale": costo_unit * qta_tot,
                "padre_codice": codice_padre,
                "livello_esplosione": livello + 1, "percorso": nuovo_percorso,
                "lead_time_giorni": f_lt_a or 0, **param_vals
            })
        else:
            esplodi_bom_ricorsiva(cursor, f_cod, qta_tot, contesto, ordine_id,
                                  livello + 1, nuovo_percorso, risultati)
    return risultati


def _valuta_condizione_bom(condizione, contesto):
    try:
        expr = condizione
        for match in re.finditer(r'\[(\w+)\]', condizione):
            campo = match.group(1)
            valore = contesto.get(campo, contesto.get(campo.upper(), ""))
            if isinstance(valore, (int, float)):
                expr = expr.replace(f"[{campo}]", str(valore))
            else:
                expr = expr.replace(f"[{campo}]", f"'{valore}'" if valore else "''")
        return bool(eval(expr))
    except:
        return True


def _calcola_formula(formula, contesto):
    try:
        expr = formula
        for match in re.finditer(r'\[(\w+)\]', formula):
            campo = match.group(1)
            valore = contesto.get(campo, contesto.get(campo.upper(), 0))
            try:
                valore = float(valore)
            except (ValueError, TypeError):
                valore = 0
            expr = expr.replace(f"[{campo}]", str(valore))
        return max(1, math.ceil(eval(expr)))
    except:
        return None



# ==========================================
# AUTH
# ==========================================
@app.post("/auth/login")
def login(data: dict, db: Session = Depends(get_db)):
    """Login utente, restituisce JWT token"""
    username = data.get("username", "")
    password = data.get("password", "")
    
    user = authenticate_user(db, username, password)
    if not user:
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    
    # Aggiorna last_login
    user.last_login = datetime.now()
    db.commit()
    
    # Carica permessi e info ruolo/gruppo
    permessi = get_user_permissions(user, db)
    
    gruppo_nome = None
    if user.gruppo_id:
        gruppo = db.query(GruppoUtenti).filter(GruppoUtenti.id == user.gruppo_id).first()
        gruppo_nome = gruppo.nome if gruppo else None
    
    ruolo_nome = None
    ruolo_codice = None
    if user.ruolo_id:
        ruolo = db.query(Ruolo).filter(Ruolo.id == user.ruolo_id).first()
        if ruolo:
            ruolo_nome = ruolo.nome
            ruolo_codice = ruolo.codice
    
    access_token = create_access_token(data={"sub": user.username})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "username": user.username,
            "nome": user.nome,
            "cognome": user.cognome,
            "email": user.email,
            "is_admin": user.is_admin,
            "gruppo_id": user.gruppo_id,
            "gruppo_nome": gruppo_nome,
            "ruolo_id": user.ruolo_id,
            "ruolo_nome": ruolo_nome,
            "ruolo_codice": ruolo_codice,
            "permessi": permessi,
        }
    }

@app.get("/auth/me")
def get_current_user_info(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db)
):
    """Ottieni info utente corrente dal token"""
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
    
    if not token:
        raise HTTPException(status_code=401, detail="Token non fornito")
    
    user = get_user_from_token(token, db)
    if not user:
        raise HTTPException(status_code=401, detail="Token non valido")
    
    permessi = get_user_permissions(user, db)
    
    gruppo_nome = None
    if user.gruppo_id:
        gruppo = db.query(GruppoUtenti).filter(GruppoUtenti.id == user.gruppo_id).first()
        gruppo_nome = gruppo.nome if gruppo else None
    
    ruolo_nome = None
    ruolo_codice = None
    if user.ruolo_id:
        ruolo = db.query(Ruolo).filter(Ruolo.id == user.ruolo_id).first()
        if ruolo:
            ruolo_nome = ruolo.nome
            ruolo_codice = ruolo.codice
    
    return {
        "id": user.id,
        "username": user.username,
        "nome": user.nome,
        "cognome": user.cognome,
        "email": user.email,
        "is_admin": user.is_admin,
        "is_active": user.is_active,
        "gruppo_id": user.gruppo_id,
        "gruppo_nome": gruppo_nome,
        "ruolo_id": user.ruolo_id,
        "ruolo_nome": ruolo_nome,
        "ruolo_codice": ruolo_codice,
        "permessi": permessi,
    }

# ==========================================
# PREVENTIVI
# ==========================================
@app.get("/preventivi")
def get_preventivi(db: Session = Depends(get_db)):
    """Ottieni tutti i preventivi ordinati per data"""
    preventivi = db.query(Preventivo).order_by(Preventivo.created_at.desc()).all()
    results = []
    for p in preventivi:
        row = {}
        for col in p.__table__.columns:
            val = getattr(p, col.name, None)
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            row[col.name] = val
        results.append(row)
    return results


@app.get("/preventivi/search")
def search_preventivi(
    q: str = None,
    status: str = None,
    categoria: str = None,
    data_da: str = None,
    data_a: str = None,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Ricerca preventivi con filtri multipli.
    q = testo libero (cerca in numero_preventivo, customer_name, ragione_sociale cliente)
    status = draft|confermato|inviato
    categoria = RISE|HOME
    data_da/data_a = YYYY-MM-DD
    """
    query = db.query(Preventivo).outerjoin(
        Cliente, Preventivo.cliente_id == Cliente.id
    )

    if q and q.strip():
        pattern = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Preventivo.numero_preventivo.ilike(pattern),
                Preventivo.customer_name.ilike(pattern),
                Cliente.ragione_sociale.ilike(pattern),
                Cliente.codice.ilike(pattern),
            )
        )

    if status:
        query = query.filter(Preventivo.status == status)

    if categoria:
        query = query.filter(Preventivo.categoria == categoria.upper())

    if data_da:
        try:
            from datetime import datetime as dt
            d = dt.strptime(data_da, "%Y-%m-%d")
            query = query.filter(Preventivo.created_at >= d)
        except ValueError:
            pass

    if data_a:
        try:
            from datetime import datetime as dt
            d = dt.strptime(data_a, "%Y-%m-%d")
            # Include tutto il giorno
            d = d.replace(hour=23, minute=59, second=59)
            query = query.filter(Preventivo.created_at <= d)
        except ValueError:
            pass

    preventivi = query.order_by(Preventivo.created_at.desc()).limit(limit).all()

    results = []
    for p in preventivi:
        row = {}
        for col in p.__table__.columns:
            row[col.name] = getattr(p, col.name, None)
        # Aggiungi ragione_sociale cliente se disponibile
        if p.cliente_id:
            cl = db.query(Cliente).filter(Cliente.id == p.cliente_id).first()
            if cl:
                row['cliente_ragione_sociale'] = cl.ragione_sociale
                row['cliente_codice'] = cl.codice
        results.append(row)

    return results

@app.get("/preventivi/{preventivo_id}")
def get_preventivo(preventivo_id: int, db: Session = Depends(get_db)):
    """Ottieni un preventivo specifico"""
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    # Ritorna dict con tutte le colonne (incluso cliente_id)
    result = {}
    for col in preventivo.__table__.columns:
        result[col.name] = getattr(preventivo, col.name, None)
    return result

@app.post("/preventivi", status_code=201)
def create_preventivo(data: PreventivoCreate, db: Session = Depends(get_db)):
    """Crea nuovo preventivo con generazione automatica numero preventivo."""
    # Genera numero preventivo automatico se non fornito
    if not data.numero_preventivo:
        current_year = datetime.now().year
        last_preventivo = (
            db.query(Preventivo)
            .filter(Preventivo.numero_preventivo.like(f"{current_year}/%"))
            .order_by(Preventivo.id.desc())
            .first()
        )
        
        if last_preventivo and "/" in (_prev_numero(last_preventivo)):
            try:
                last_number = int((_prev_numero(last_preventivo)).split("/")[1])
                new_number = last_number + 1
            except (ValueError, IndexError):
                new_number = 1
        else:
            new_number = 1
        
        numero_preventivo = f"{current_year}/{new_number:04d}"
    else:
        numero_preventivo = data.numero_preventivo
    
    # Carica template se specificato
    template_data_parsed = None
    template = None
    if data.template_id:
        template = db.query(ProductTemplate).filter(ProductTemplate.id == data.template_id).first()
        if template and template.template_data:
            try:
                template_data_parsed = json.loads(template.template_data)
            except json.JSONDecodeError:
                pass
    
    # Crea preventivo
    preventivo_dict = data.dict(exclude_unset=True)
    preventivo_dict['numero_preventivo'] = numero_preventivo
    template_id = preventivo_dict.pop('template_id', None)
    
    if template_id:
        preventivo_dict['template_id'] = template_id
        if template:
            preventivo_dict['categoria'] = template.categoria
    
    if 'created_at' not in preventivo_dict:
        preventivo_dict['created_at'] = datetime.now()
    if 'updated_at' not in preventivo_dict:
        preventivo_dict['updated_at'] = datetime.now()
    
    preventivo = Preventivo(**preventivo_dict)
    db.add(preventivo)
    db.commit()
    db.refresh(preventivo)
    
    # Crea record ORM vuoti per le sezioni dedicate (backward compat)
    db.add_all([
        DatiCommessa(preventivo_id=preventivo.id),
        DatiPrincipali(preventivo_id=preventivo.id),
        Normative(preventivo_id=preventivo.id),
        DisposizioneVano(preventivo_id=preventivo.id),
        Porte(preventivo_id=preventivo.id),
        Argano(preventivo_id=preventivo.id),
    ])
    db.commit()
    
    # Scrivi valori template in valori_configurazione (fonte primaria)
    if template_data_parsed:
        skip_keys = {"materiali", "field_config"}
        for sez_codice, campi in template_data_parsed.items():
            if sez_codice in skip_keys or not isinstance(campi, dict):
                continue
            for codice_campo, valore in campi.items():
                # Salta valori che sono dict/list (sotto-config, non valori reali)
                if isinstance(valore, (dict, list)):
                    continue
                if valore is not None and str(valore).strip() != "":
                    db.execute(text("""
                        INSERT INTO valori_configurazione
                            (preventivo_id, sezione, codice_campo, valore, is_default)
                        VALUES (:pid, :sez, :campo, :val, 1)
                    """), {
                        "pid": preventivo.id,
                        "sez": sez_codice,
                        "campo": codice_campo,
                        "val": str(valore),
                    })
        db.commit()
    
    # Aggiungi materiali da template se presenti
    if template_data_parsed and "materiali" in template_data_parsed:
        for mat in template_data_parsed["materiali"]:
            quantita = mat.get("quantita", 1)
            prezzo_unitario = mat.get("prezzo_unitario", 0.0)
            materiale = Materiale(
                preventivo_id=preventivo.id,
                codice=mat.get("codice", ""),
                descrizione=mat.get("descrizione", ""),
                quantita=quantita,
                prezzo_unitario=prezzo_unitario,
                prezzo_totale=quantita * prezzo_unitario,
                categoria=mat.get("categoria", "Template"),
                aggiunto_da_regola=False,
                note=f"Da template: {template.nome_display}" if template else None
            )
            db.add(materiale)
        
        db.commit()
        all_materials = db.query(Materiale).filter(Materiale.preventivo_id == preventivo.id).all()
        totale_calc = sum(m.prezzo_totale for m in all_materials)
        _prev_set(preventivo, totale_calc, "totale_materiali", "total_price")
        db.commit()
        db.refresh(preventivo)
    
    # Valuta regole (incluse TEMPLATE_BASE) per aggiungere articoli base automatici
    safe_evaluate_rules(preventivo.id, db)
    db.refresh(preventivo)

    # Inizializza defaults e applica field_config dal template
    try:
        campi_default = db.execute(text("""
            SELECT codice, sezione, valore_default
            FROM campi_configuratore
            WHERE attivo = 1 AND valore_default IS NOT NULL AND valore_default != ''
        """)).fetchall()

        existing_campos = db.execute(text("""
            SELECT codice_campo FROM valori_configurazione WHERE preventivo_id = :pid
        """), {"pid": preventivo.id}).fetchall()
        existing_set = {r[0] for r in existing_campos}

        for codice, sezione, valore_default in campi_default:
            if codice not in existing_set:
                db.execute(text("""
                    INSERT INTO valori_configurazione
                        (preventivo_id, sezione, codice_campo, valore, is_default)
                    VALUES (:pid, :sez, :campo, :val, 1)
                """), {"pid": preventivo.id, "sez": sezione,
                       "campo": codice, "val": str(valore_default)})
        db.commit()

        if data.template_id:
            _apply_template_field_config(db, preventivo.id, data.template_id)
            db.commit()
    except Exception as e:
        logger.warning(f"Errore inizializzazione defaults/field_config: {e}")
    
    return _orm_to_dict(preventivo)

@app.put("/preventivi/{preventivo_id}")
def update_preventivo(preventivo_id: int, data: PreventivoUpdate, db: Session = Depends(get_db)):
    """Aggiorna preventivo"""
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    
    update_dict = data.dict(exclude_unset=True)
    update_dict['updated_at'] = datetime.now()
    
    for key, value in update_dict.items():
        setattr(preventivo, key, value)
    
    db.commit()
    db.refresh(preventivo)
    return _orm_to_dict(preventivo)

@app.delete("/preventivi/{preventivo_id}")
def delete_preventivo(preventivo_id: int, db: Session = Depends(get_db)):
    """Elimina preventivo"""
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    db.delete(preventivo)
    db.commit()
    return {"message": "Preventivo eliminato"}

# ==========================================
# DATI COMMESSA
# ==========================================
@app.get("/preventivi/{preventivo_id}/dati-commessa")
def get_dati_commessa(preventivo_id: int, db: Session = Depends(get_db)):
    dati = db.query(DatiCommessa).filter(DatiCommessa.preventivo_id == preventivo_id).first()
    if not dati:
        dati = DatiCommessa(preventivo_id=preventivo_id)
        db.add(dati)
        db.commit()
        db.refresh(dati)
    # Ritorna dict manuale per evitare errori schema
    result = {}
    for col in dati.__table__.columns:
        result[col.name] = getattr(dati, col.name, None)
    # Alias per frontend
    if "consegna_richiesta" in result:
        result["data_consegna_richiesta"] = result["consegna_richiesta"]
    return result

@app.put("/preventivi/{preventivo_id}/dati-commessa")
def update_dati_commessa(preventivo_id: int, data: DatiCommessaUpdate, db: Session = Depends(get_db)):
    try:
        dati = db.query(DatiCommessa).filter(DatiCommessa.preventivo_id == preventivo_id).first()
        if not dati:
            dati = DatiCommessa(preventivo_id=preventivo_id)
            db.add(dati)
        update_data = data.dict(exclude_unset=True)
        print(f"[DATI COMMESSA] PUT preventivo_id={preventivo_id}, fields={list(update_data.keys())}")
        # Salva campi dati_commessa
        cliente_id_value = update_data.pop('cliente_id', None)
        # Mapping nomi frontend -> nomi ORM
        field_aliases = {
            "data_consegna_richiesta": "consegna_richiesta",
        }
        for key, value in update_data.items():
            orm_key = field_aliases.get(key, key)
            if hasattr(dati, orm_key):
                setattr(dati, orm_key, value)
            else:
                print(f"[DATI COMMESSA] WARNING: campo '{key}' (orm: '{orm_key}') non esiste nel modello DatiCommessa")
        
        # Se presente cliente_id, aggiorna direttamente il preventivo
        if cliente_id_value is not None:
            preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
            if preventivo:
                preventivo.cliente_id = cliente_id_value
                print(f"[DATI COMMESSA] Aggiornato cliente_id={cliente_id_value} sul preventivo {preventivo_id}")
        
        db.commit()
        db.refresh(dati)
        safe_evaluate_rules(preventivo_id, db)
        # Ritorna dict manuale per evitare errori schema
        result = {}
        for col in dati.__table__.columns:
            result[col.name] = getattr(dati, col.name, None)
        return result
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Errore salvataggio dati commessa: {str(e)}")
    
# ============================================================
# FILE: main.py â€” DELETE CLIENTE (aggiungere dopo PUT /clienti/{id})
# ============================================================

@app.delete("/clienti/{cliente_id}")
def delete_cliente(cliente_id: int, db: Session = Depends(get_db)):
    """Disattiva un cliente (soft delete)"""
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(404, "Cliente non trovato")
    cliente.is_active = False
    db.commit()
    return {"status": "ok"}

# ==========================================
# DATI PRINCIPALI
# ==========================================
@app.get("/preventivi/{preventivo_id}/dati-principali")
def get_dati_principali(preventivo_id: int, db: Session = Depends(get_db)):
    dati = db.query(DatiPrincipali).filter(DatiPrincipali.preventivo_id == preventivo_id).first()
    if not dati:
        dati = DatiPrincipali(preventivo_id=preventivo_id)
        db.add(dati)
        db.commit()
        db.refresh(dati)
    return _orm_to_dict(dati)

@app.put("/preventivi/{preventivo_id}/dati-principali")
def update_dati_principali(preventivo_id: int, data: DatiPrincipaliUpdate, db: Session = Depends(get_db)):
    dati = db.query(DatiPrincipali).filter(DatiPrincipali.preventivo_id == preventivo_id).first()
    if not dati:
        dati = DatiPrincipali(preventivo_id=preventivo_id)
        db.add(dati)
    update_data = data.dict(exclude_unset=True)
    print(f"[DATI_PRINCIPALI] PUT preventivo_id={preventivo_id}, fields={list(update_data.keys())}")
    for key, value in update_data.items():
        if hasattr(dati, key):
            setattr(dati, key, value)
        else:
            print(f"[DATI_PRINCIPALI] WARNING: campo '{key}' non esiste nel modello")
    db.commit()
    db.refresh(dati)
    rule_result = safe_evaluate_rules(preventivo_id, db)
    result = _orm_to_dict(dati)
    if rule_result and isinstance(rule_result, dict):
        result["materials_added"] = rule_result.get("materiali_aggiunti", 0)
        result["materials_removed"] = rule_result.get("materiali_rimossi", 0)
        result["active_rules"] = rule_result.get("regole_attive", [])
    return result

# ==========================================
# NORMATIVE
# ==========================================
@app.get("/preventivi/{preventivo_id}/normative")
def get_normative(preventivo_id: int, db: Session = Depends(get_db)):
    normative = db.query(Normative).filter(Normative.preventivo_id == preventivo_id).first()
    if not normative:
        normative = Normative(preventivo_id=preventivo_id)
        db.add(normative)
        db.commit()
        db.refresh(normative)
    return _orm_to_dict(normative)

@app.put("/preventivi/{preventivo_id}/normative")
def update_normative(preventivo_id: int, data: NormativeUpdate, db: Session = Depends(get_db)):
    normative = db.query(Normative).filter(Normative.preventivo_id == preventivo_id).first()
    if not normative:
        normative = Normative(preventivo_id=preventivo_id)
        db.add(normative)
    update_data = data.dict(exclude_unset=True)
    print(f"[NORMATIVE] PUT preventivo_id={preventivo_id}, fields={update_data}")
    for key, value in update_data.items():
        if hasattr(normative, key):
            setattr(normative, key, value)
        else:
            print(f"[NORMATIVE] WARNING: campo '{key}' non esiste nel modello")
    db.commit()
    db.refresh(normative)
    rule_result = safe_evaluate_rules(preventivo_id, db)
    result = _orm_to_dict(normative)
    # Aggiungi info regole alla risposta
    if rule_result and isinstance(rule_result, dict):
        result["materials_added"] = rule_result.get("materiali_aggiunti", 0)
        result["materials_removed"] = rule_result.get("materiali_rimossi", 0)
        result["active_rules"] = rule_result.get("regole_attive", [])
    return result

# ==========================================
# DISPOSIZIONE VANO
# ==========================================
@app.get("/preventivi/{preventivo_id}/disposizione-vano")
def get_disposizione_vano(preventivo_id: int, db: Session = Depends(get_db)):
    disposizione = db.query(DisposizioneVano).filter(DisposizioneVano.preventivo_id == preventivo_id).first()
    if not disposizione:
        disposizione = DisposizioneVano(preventivo_id=preventivo_id)
        db.add(disposizione)
        db.commit()
        db.refresh(disposizione)
    return _orm_to_dict(disposizione)

@app.put("/preventivi/{preventivo_id}/disposizione-vano")
def update_disposizione_vano(preventivo_id: int, data: DisposizioneVanoUpdate, db: Session = Depends(get_db)):
    disposizione = db.query(DisposizioneVano).filter(DisposizioneVano.preventivo_id == preventivo_id).first()
    if not disposizione:
        disposizione = DisposizioneVano(preventivo_id=preventivo_id)
        db.add(disposizione)
    
    data_dict = data.dict(exclude_unset=True)
    cleaned_data = {}
    for key, value in data_dict.items():
        if value == '' or value == '{}' or value == '[]':
            cleaned_data[key] = None
        else:
            cleaned_data[key] = value
    
    for key, value in cleaned_data.items():
        if hasattr(disposizione, key):
            setattr(disposizione, key, value)
        else:
            print(f"[DISPOSIZIONE_VANO] WARNING: campo '{key}' non esiste nel modello")
    
    db.commit()
    db.refresh(disposizione)
    return _orm_to_dict(disposizione)

@app.delete("/preventivi/{preventivo_id}/disposizione-vano")
def delete_disposizione_vano(preventivo_id: int, db: Session = Depends(get_db)):
    disposizione = db.query(DisposizioneVano).filter(DisposizioneVano.preventivo_id == preventivo_id).first()
    if disposizione:
        db.delete(disposizione)
        db.commit()
    return {"message": "Disposizione vano eliminata"}

# ==========================================
# PORTE
# ==========================================
@app.get("/preventivi/{preventivo_id}/porte")
def get_porte(preventivo_id: int, db: Session = Depends(get_db)):
    porte = db.query(Porte).filter(Porte.preventivo_id == preventivo_id).first()
    if not porte:
        porte = Porte(preventivo_id=preventivo_id)
        db.add(porte)
        db.commit()
        db.refresh(porte)
    return _orm_to_dict(porte)

@app.put("/preventivi/{preventivo_id}/porte")
def update_porte(preventivo_id: int, data: PorteUpdate, db: Session = Depends(get_db)):
    porte = db.query(Porte).filter(Porte.preventivo_id == preventivo_id).first()
    if not porte:
        porte = Porte(preventivo_id=preventivo_id)
        db.add(porte)
    for key, value in data.dict(exclude_unset=True).items():
        if hasattr(porte, key):
            setattr(porte, key, value)
        else:
            print(f"[PORTE] WARNING: campo '{key}' non esiste nel modello")
    db.commit()
    db.refresh(porte)
    safe_evaluate_rules(preventivo_id, db)
    return _orm_to_dict(porte)

# ==========================================
# ARGANO
# ==========================================
@app.get("/preventivi/{preventivo_id}/argano")
def get_argano(preventivo_id: int, db: Session = Depends(get_db)):
    """Ottieni dati argano, crea se non esiste"""
    argano = db.query(Argano).filter(Argano.preventivo_id == preventivo_id).first()
    if not argano:
        argano = Argano(preventivo_id=preventivo_id)
        db.add(argano)
        db.commit()
        db.refresh(argano)
    return {
        "id": argano.id,
        "preventivo_id": argano.preventivo_id,
        "trazione": argano.trazione,
        "potenza_motore_kw": argano.potenza_motore_kw,
        "corrente_nom_motore_amp": argano.corrente_nom_motore_amp,
        "tipo_vvvf": argano.tipo_vvvf,
        "vvvf_nel_vano": argano.vvvf_nel_vano,
        "freno_tensione": argano.freno_tensione,
        "ventilazione_forzata": argano.ventilazione_forzata,
        "tipo_teleruttore": argano.tipo_teleruttore
    }

@app.put("/preventivi/{preventivo_id}/argano")
def update_argano(preventivo_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna dati argano"""
    argano = db.query(Argano).filter(Argano.preventivo_id == preventivo_id).first()
    if not argano:
        argano = Argano(preventivo_id=preventivo_id)
        db.add(argano)
    
    for key in ["trazione", "potenza_motore_kw", "corrente_nom_motore_amp",
                "tipo_vvvf", "vvvf_nel_vano", "freno_tensione",
                "ventilazione_forzata", "tipo_teleruttore"]:
        if key in data:
            setattr(argano, key, data[key])
    
    db.commit()
    db.refresh(argano)

    safe_evaluate_rules(preventivo_id, db)    # <-- AGGIUNGI QUESTA RIGA

    return {
        "id": argano.id,
        "preventivo_id": argano.preventivo_id,
        "trazione": argano.trazione,
        "potenza_motore_kw": argano.potenza_motore_kw,
        "corrente_nom_motore_amp": argano.corrente_nom_motore_amp,
        "tipo_vvvf": argano.tipo_vvvf,
        "vvvf_nel_vano": argano.vvvf_nel_vano,
        "freno_tensione": argano.freno_tensione,
        "ventilazione_forzata": argano.ventilazione_forzata,
        "tipo_teleruttore": argano.tipo_teleruttore
    }


# ==========================================
# MATERIALI
# ==========================================
@app.get("/preventivi/{preventivo_id}/materiali")
def get_materiali(preventivo_id: int, db: Session = Depends(get_db)):
    return _orm_to_dict(db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all())

@app.post("/preventivi/{preventivo_id}/materiali")
def create_materiale(preventivo_id: int, data: MaterialeCreate, db: Session = Depends(get_db)):
    prezzo_totale = data.quantita * data.prezzo_unitario
    materiale = Materiale(
        preventivo_id=preventivo_id,
        codice=data.codice, descrizione=data.descrizione,
        quantita=data.quantita, prezzo_unitario=data.prezzo_unitario,
        prezzo_totale=prezzo_totale, categoria=data.categoria,
        aggiunto_da_regola=data.aggiunto_da_regola,
        regola_id=data.regola_id, note=data.note
    )
    db.add(materiale)
    db.commit()
    db.refresh(materiale)
    
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if preventivo:
        all_materials = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
        totale_calc = sum(m.prezzo_totale for m in all_materials)
        _prev_set(preventivo, totale_calc, "totale_materiali", "total_price")
        db.commit()
    return _orm_to_dict(materiale)

@app.put("/preventivi/{preventivo_id}/materiali/{materiale_id}")
def update_materiale(preventivo_id: int, materiale_id: int, data: MaterialeUpdate, db: Session = Depends(get_db)):
    materiale = db.query(Materiale).filter(Materiale.id == materiale_id, Materiale.preventivo_id == preventivo_id).first()
    if not materiale:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    
    update_data = data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(materiale, key, value)
    
    if 'quantita' in update_data or 'prezzo_unitario' in update_data:
        materiale.prezzo_totale = materiale.quantita * materiale.prezzo_unitario
    
    db.commit()
    db.refresh(materiale)
    
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if preventivo:
        all_materials = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
        totale_calc = sum(m.prezzo_totale for m in all_materials)
        _prev_set(preventivo, totale_calc, "totale_materiali", "total_price")
        db.commit()
    return _orm_to_dict(materiale)

@app.delete("/preventivi/{preventivo_id}/materiali/{materiale_id}")
def delete_materiale(preventivo_id: int, materiale_id: int, db: Session = Depends(get_db)):
    materiale = db.query(Materiale).filter(Materiale.id == materiale_id, Materiale.preventivo_id == preventivo_id).first()
    if not materiale:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    db.delete(materiale)
    db.commit()
    
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if preventivo:
        all_materials = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
        totale_calc = sum(m.prezzo_totale for m in all_materials)
        _prev_set(preventivo, totale_calc, "totale_materiali", "total_price")
        db.commit()
    return {"message": "Materiale eliminato"}

# ==========================================
# RIGHE RICAMBIO
# ==========================================
@app.get("/preventivi/{preventivo_id}/righe-ricambio")
def get_righe_ricambio(preventivo_id: int, db: Session = Depends(get_db)):
    righe = db.query(RigaRicambio).filter(
        RigaRicambio.preventivo_id == preventivo_id
    ).order_by(RigaRicambio.ordine).all()
    return [{
        "id": r.id, "preventivo_id": r.preventivo_id,
        "articolo_id": r.articolo_id, "codice": r.codice,
        "descrizione": r.descrizione, "tipo_articolo": r.tipo_articolo,
        "quantita": r.quantita,
        "parametro_1": r.parametro_1, "unita_param_1": r.unita_param_1,
        "desc_param_1": r.desc_param_1, "costo_var_1": r.costo_var_1,
        "parametro_2": r.parametro_2, "unita_param_2": r.unita_param_2,
        "desc_param_2": r.desc_param_2, "costo_var_2": r.costo_var_2,
        "parametro_3": r.parametro_3, "unita_param_3": r.unita_param_3,
        "desc_param_3": r.desc_param_3, "costo_var_3": r.costo_var_3,
        "parametro_4": r.parametro_4, "unita_param_4": r.unita_param_4,
        "desc_param_4": r.desc_param_4, "costo_var_4": r.costo_var_4,
        "costo_fisso": r.costo_fisso, "costo_base_unitario": r.costo_base_unitario,
        "ricarico_percentuale": r.ricarico_percentuale,
        "prezzo_listino_unitario": r.prezzo_listino_unitario,
        "sconto_cliente": r.sconto_cliente,
        "prezzo_cliente_unitario": r.prezzo_cliente_unitario,
        "prezzo_totale_listino": r.prezzo_totale_listino,
        "prezzo_totale_cliente": r.prezzo_totale_cliente,
        "ordine": r.ordine, "note": r.note
    } for r in righe]

@app.post("/preventivi/{preventivo_id}/righe-ricambio")
def create_riga_ricambio(preventivo_id: int, data: dict, db: Session = Depends(get_db)):
    riga = RigaRicambio(preventivo_id=preventivo_id)
    for key in [
        "articolo_id", "codice", "descrizione", "tipo_articolo", "quantita",
        "parametro_1", "unita_param_1", "desc_param_1", "costo_var_1",
        "parametro_2", "unita_param_2", "desc_param_2", "costo_var_2",
        "parametro_3", "unita_param_3", "desc_param_3", "costo_var_3",
        "parametro_4", "unita_param_4", "desc_param_4", "costo_var_4",
        "costo_fisso", "costo_base_unitario", "ricarico_percentuale",
        "prezzo_listino_unitario", "sconto_cliente", "prezzo_cliente_unitario",
        "prezzo_totale_listino", "prezzo_totale_cliente", "ordine", "note"
    ]:
        if key in data:
            setattr(riga, key, data[key])
    
    db.add(riga)
    db.commit()
    db.refresh(riga)
    return {"id": riga.id, "status": "ok"}

@app.put("/preventivi/{preventivo_id}/righe-ricambio/{riga_id}")
def update_riga_ricambio(preventivo_id: int, riga_id: int, data: dict, db: Session = Depends(get_db)):
    riga = db.query(RigaRicambio).filter(
        RigaRicambio.id == riga_id,
        RigaRicambio.preventivo_id == preventivo_id
    ).first()
    if not riga:
        raise HTTPException(status_code=404, detail="Riga ricambio non trovata")
    
    for key in [
        "articolo_id", "codice", "descrizione", "tipo_articolo", "quantita",
        "parametro_1", "unita_param_1", "desc_param_1", "costo_var_1",
        "parametro_2", "unita_param_2", "desc_param_2", "costo_var_2",
        "parametro_3", "unita_param_3", "desc_param_3", "costo_var_3",
        "parametro_4", "unita_param_4", "desc_param_4", "costo_var_4",
        "costo_fisso", "costo_base_unitario", "ricarico_percentuale",
        "prezzo_listino_unitario", "sconto_cliente", "prezzo_cliente_unitario",
        "prezzo_totale_listino", "prezzo_totale_cliente", "ordine", "note"
    ]:
        if key in data:
            setattr(riga, key, data[key])
    
    db.commit()
    db.refresh(riga)
    return {"id": riga.id, "status": "ok"}

@app.delete("/preventivi/{preventivo_id}/righe-ricambio/{riga_id}")
def delete_riga_ricambio(preventivo_id: int, riga_id: int, db: Session = Depends(get_db)):
    riga = db.query(RigaRicambio).filter(
        RigaRicambio.id == riga_id,
        RigaRicambio.preventivo_id == preventivo_id
    ).first()
    if not riga:
        raise HTTPException(status_code=404, detail="Riga ricambio non trovata")
    db.delete(riga)
    db.commit()
    return {"status": "ok"}

# ==========================================
# SCONTO ADMIN
# ==========================================
@app.put("/preventivi/{preventivo_id}/sconto-admin")
def update_sconto_admin(preventivo_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna sconto extra admin per un preventivo"""
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    
    if "sconto_extra_admin" in data:
        preventivo.sconto_extra_admin = data["sconto_extra_admin"]
    
    # Ricalcola prezzo finale
    sconto_totale = (preventivo.sconto_cliente or 0) + (preventivo.sconto_extra_admin or 0)
    if _prev_totale(preventivo):
        totale_netto_calc = _prev_totale(preventivo) * (1 - sconto_totale / 100)
        _prev_set(preventivo, totale_netto_calc, 'totale_netto', 'total_price_finale')
    
    preventivo.updated_at = datetime.now()
    db.commit()
    db.refresh(preventivo)
    return {
        "sconto_extra_admin": preventivo.sconto_extra_admin,
        "total_price_finale": _prev_netto(preventivo)
    }

# ==========================================
# EXPORT PREVENTIVO
# ==========================================
@app.get("/preventivi/{preventivo_id}/export/{formato}")
def export_preventivo(preventivo_id: int, formato: str, request: Request, db: Session = Depends(get_db)):    
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    
    dati_commessa = db.query(DatiCommessa).filter(DatiCommessa.preventivo_id == preventivo_id).first()
    dati_principali = db.query(DatiPrincipali).filter(DatiPrincipali.preventivo_id == preventivo_id).first()
    normative_data = db.query(Normative).filter(Normative.preventivo_id == preventivo_id).first()
    materiali = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
    argano_data = None
    try:
        argano_data = db.query(Argano).filter(Argano.preventivo_id == preventivo_id).first()
    except Exception:
        pass
    cliente = None
    try:
        cid = getattr(preventivo, 'cliente_id', None)
        if cid:
            cliente = db.query(Cliente).filter(Cliente.id == cid).first()
    except Exception:
        pass
    
    fmt = formato.lower()
    filename_base = f"preventivo_{_prev_numero(preventivo).replace('/', '_')}"
    
    if fmt == "json":
        return {
            "preventivo": {
                "id": preventivo.id,
                "numero_preventivo": _prev_numero(preventivo),
                "status": _prev_stato(preventivo),
                "total_price": _prev_totale(preventivo),
                "customer_name": getattr(preventivo, 'customer_name', ''),
                "created_at": str(preventivo.created_at) if preventivo.created_at else None,
            },
            "dati_commessa": {k: v for k, v in (dati_commessa.__dict__.items() if dati_commessa else {}) if not k.startswith('_')},
            "dati_principali": {k: v for k, v in (dati_principali.__dict__.items() if dati_principali else {}) if not k.startswith('_')},
            "materiali": [{
                "codice": m.codice, "descrizione": m.descrizione,
                "quantita": m.quantita, "prezzo_unitario": m.prezzo_unitario,
                "prezzo_totale": m.prezzo_totale, "categoria": m.categoria
            } for m in materiali]
        }
    
    elif fmt in ("docx", "pdf"):
        # --- Template personalizzato ---
        doc_template_id = request.query_params.get("template_id")
        doc_template = None
        if doc_template_id:
            doc_template = db.query(DocumentTemplate).filter(
                DocumentTemplate.id == int(doc_template_id)
            ).first()
        else:
            doc_template = db.query(DocumentTemplate).filter(
                DocumentTemplate.tipo == "preventivo",
                DocumentTemplate.is_default == True,
                DocumentTemplate.attivo == True
            ).first()

        if doc_template and doc_template.config:
            try:
                available = get_available_fields_from_db(db)
                valori_din = load_valori_dinamici(db, preventivo_id)
                def_info = load_defaults_info(db, preventivo_id)
                buf = genera_docx_da_template(
                    template_config=doc_template.config,
                    preventivo=preventivo,
                    dati_commessa=dati_commessa,
                    dati_principali=dati_principali,
                    normative=normative_data,
                    argano=argano_data,
                    materiali=materiali,
                    cliente=cliente,
                    logo_data=doc_template.logo_data,
                    logo_mime=doc_template.logo_mime,
                    valori_dinamici=valori_din,
                    available_fields=available,
                    defaults_info=def_info,
                )
                return StreamingResponse(buf,
                    media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    headers={"Content-Disposition": f'attachment; filename="{filename_base}.docx"'})
            except Exception as e:
                print(f"Errore template personalizzato, fallback: {e}")

        # --- Fallback: generazione standard ---
        try:
            dati_doc = get_dati_documento_preventivo(preventivo_id, db)
            preventivo_info = {
                "numero": _prev_numero(preventivo),
                "customer": getattr(preventivo, 'customer_name', '') or '',
                "status": _prev_stato(preventivo),
                "totale": safe_float(_prev_totale(preventivo)),
                "sconto": safe_float(getattr(preventivo, 'sconto_cliente', 0)) + safe_float(getattr(preventivo, 'sconto_extra_admin', 0)),
                "netto": safe_float(_prev_netto(preventivo)),
                "note": getattr(preventivo, 'note', '') or '',
                "revisione": getattr(preventivo, 'revisione_corrente', 0) or 0,
            }
            if cliente:
                preventivo_info["customer"] = getattr(cliente, 'ragione_sociale', preventivo_info["customer"]) or preventivo_info["customer"]
            from export_utils import genera_docx_preventivo_v2
            buf = genera_docx_preventivo_v2(preventivo_info, dati_doc, materiali)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}.docx"'})
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"Dipendenza mancante: {e}. Installa: pip install python-docx")
    
    elif fmt in ("xlsx", "excel", "xls"):
        try:
            from export_utils import genera_xlsx_preventivo
            buf = genera_xlsx_preventivo(preventivo, dati_commessa, dati_principali, normative_data, argano_data, materiali, cliente)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'})
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"Dipendenza mancante: {e}. Installa: pip install openpyxl")
    
    raise HTTPException(status_code=400, detail=f"Formato '{formato}' non supportato. Usa: json, docx, xlsx")


@app.get("/ordini/{ordine_id}/export/{formato}")
def export_ordine(ordine_id: int, formato: str, db: Session = Depends(get_db)):
    """Esporta ordine in formato JSON, DOCX o XLSX (include BOM esplosa e lista acquisti)"""
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM ordini WHERE id = ?", (ordine_id,))
    r = cursor.fetchone()
    if not r:
        raise HTTPException(status_code=404, detail="Ordine non trovato")
    cols = [desc[0] for desc in cursor.description]
    ordine_data = dict(zip(cols, r))
    
    if ordine_data.get('cliente_id'):
        try:
            cl = db.query(Cliente).filter(Cliente.id == ordine_data['cliente_id']).first()
            ordine_data['cliente'] = cl.ragione_sociale if cl else ''
        except Exception:
            ordine_data['cliente'] = ''
    
    materiali = []
    if ordine_data.get('preventivo_id'):
        materiali = db.query(Materiale).filter(
            Materiale.preventivo_id == ordine_data['preventivo_id']
        ).order_by(Materiale.ordine).all()
    
    esplosi = []
    try:
        cursor.execute("SELECT * FROM esplosi WHERE ordine_id = ? ORDER BY tipo, categoria, codice", (ordine_id,))
        ecols = [d[0] for d in cursor.description]
        esplosi = [dict(zip(ecols, row)) for row in cursor.fetchall()]
    except Exception:
        pass
    
    lista_acquisti = None
    try:
        cursor.execute(
            "SELECT e.codice, e.descrizione, e.quantita, e.unita_misura, e.costo_unitario, e.costo_totale, "
            "ab.fornitore, ab.codice_fornitore, ab.lead_time_acquisto "
            "FROM esplosi e LEFT JOIN articoli_bom ab ON ab.codice = e.codice "
            "WHERE e.ordine_id = ? AND e.tipo = 'ACQUISTO' ORDER BY ab.fornitore, e.codice",
            (ordine_id,)
        )
        fornitori = {}
        for row in cursor.fetchall():
            f = row[6] or "Non assegnato"
            if f not in fornitori:
                fornitori[f] = {"articoli": [], "totale": 0}
            fornitori[f]["articoli"].append({
                "codice": row[0], "descrizione": row[1], "quantita": row[2],
                "unita": row[3], "costo_unitario": row[4], "costo_totale": row[5],
                "codice_fornitore": row[7], "lead_time_giorni": row[8]
            })
            fornitori[f]["totale"] += row[5] or 0
        if fornitori:
            lista_acquisti = {"fornitori": {n: {"totale": round(d["totale"], 2), "articoli": d["articoli"]} for n, d in fornitori.items()}}
    except Exception:
        pass
    
    fmt = formato.lower()
    num_ordine = ordine_data.get('numero_ordine', str(ordine_id)).replace('/', '_')
    
    if fmt == "json":
        return {
            "ordine": ordine_data,
            "materiali": [{"codice": m.codice, "descrizione": m.descrizione,
                "quantita": m.quantita, "prezzo_unitario": m.prezzo_unitario,
                "prezzo_totale": m.prezzo_totale} for m in materiali],
            "esplosi": esplosi,
            "lista_acquisti": lista_acquisti
        }
    
    elif fmt in ("docx", "pdf"):
        try:
            from export_utils import genera_docx_ordine
            buf = genera_docx_ordine(ordine_data, materiali, esplosi, lista_acquisti)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={"Content-Disposition": f'attachment; filename="ordine_{num_ordine}.docx"'})
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"Dipendenza mancante: {e}. Installa: pip install python-docx")
    
    elif fmt in ("xlsx", "excel", "xls"):
        try:
            from export_utils import genera_xlsx_ordine
            buf = genera_xlsx_ordine(ordine_data, materiali, esplosi, lista_acquisti)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="ordine_{num_ordine}.xlsx"'})
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"Dipendenza mancante: {e}. Installa: pip install openpyxl")
    
    raise HTTPException(status_code=400, detail=f"Formato '{formato}' non supportato. Usa: json, docx, xlsx")


# ==========================================
# HELPER: Applica field_config da template a valori_configurazione
# ==========================================

def _apply_template_field_config(db, preventivo_id: int, template_id: int):
    """
    Legge field_config dal template e aggiorna i flag su valori_configurazione.
    Chiamare DOPO che i valori sono stati inseriti in valori_configurazione.
    """
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template or not template.template_data:
        return 0

    try:
        td = json.loads(template.template_data)
    except json.JSONDecodeError:
        return 0

    field_config = td.get("field_config", {})
    if not field_config:
        return 0

    count = 0
    for codice_campo, cfg in field_config.items():
        # Costruisci UPDATE dinamico: solo flag esplicitamente presenti nel template
        updates = []
        params = {"pid": preventivo_id, "campo": codice_campo}

        if "readonly" in cfg:
            updates.append("is_readonly = :ro")
            params["ro"] = 1 if cfg["readonly"] else 0

        if "includi_preventivo" in cfg:
            updates.append("includi_preventivo = :inc")
            params["inc"] = 1 if cfg["includi_preventivo"] else 0

        if "mostra_default" in cfg:
            updates.append("mostra_default_preventivo = :md")
            params["md"] = 1 if cfg["mostra_default"] else 0

        if not updates:
            continue

        result = db.execute(text(f"""
            UPDATE valori_configurazione
            SET {', '.join(updates)}
            WHERE preventivo_id = :pid AND codice_campo = :campo
        """), params)

        if result.rowcount > 0:
            count += 1

    return count


# ==========================================
# PRODUCT TEMPLATES
# ==========================================
@app.get("/templates")
def get_templates(categoria: str = None, db: Session = Depends(get_db)):
    query = db.query(ProductTemplate).filter(ProductTemplate.attivo == True)
    if categoria:
        query = query.filter(ProductTemplate.categoria == categoria.upper())
    return _orm_to_dict(query.order_by(ProductTemplate.categoria, ProductTemplate.ordine).all())

@app.get("/templates/all")
def get_all_templates(db: Session = Depends(get_db)):
    return _orm_to_dict(db.query(ProductTemplate).order_by(ProductTemplate.categoria, ProductTemplate.ordine).all())

@app.get("/templates/categories/summary")
def get_categories_summary(db: Session = Depends(get_db)):
    templates = db.query(ProductTemplate).filter(ProductTemplate.attivo == True).order_by(
        ProductTemplate.categoria, ProductTemplate.ordine
    ).all()
    categories = {}
    for t in templates:
        cat = t.categoria
        if cat not in categories:
            categories[cat] = {"categoria": cat, "sottocategorie": []}
        categories[cat]["sottocategorie"].append({
            "id": t.id, "sottocategoria": t.sottocategoria,
            "nome_display": t.nome_display, "descrizione": t.descrizione,
            "icona": t.icona, "ordine": t.ordine
        })
    return list(categories.values())

@app.get("/templates/{template_id}")
def get_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")
    return _orm_to_dict(template)

@app.post("/templates", status_code=201)
def create_template(data: ProductTemplateCreate, db: Session = Depends(get_db)):
    if data.template_data:
        try:
            json.loads(data.template_data)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="template_data non ÃƒÂ¨ un JSON valido")
    
    template_dict = data.dict()
    template_dict['categoria'] = template_dict['categoria'].upper()
    template_dict['created_at'] = datetime.now()
    template_dict['updated_at'] = datetime.now()
    template = ProductTemplate(**template_dict)
    db.add(template)
    db.commit()
    db.refresh(template)
    return _orm_to_dict(template)

@app.put("/templates/{template_id}")
def update_template(template_id: int, data: ProductTemplateUpdate, db: Session = Depends(get_db)):
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")
    if data.template_data:
        try:
            json.loads(data.template_data)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="template_data non ÃƒÂ¨ un JSON valido")
    
    update_dict = data.dict(exclude_unset=True)
    if 'categoria' in update_dict and update_dict['categoria']:
        update_dict['categoria'] = update_dict['categoria'].upper()
    update_dict['updated_at'] = datetime.now()
    for key, value in update_dict.items():
        setattr(template, key, value)
    db.commit()
    db.refresh(template)
    return _orm_to_dict(template)

@app.delete("/templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")
    db.delete(template)
    db.commit()
    return {"message": "Template eliminato"}


@app.get("/templates/{template_id}/field-config")
def get_template_field_config(template_id: int, db: Session = Depends(get_db)):
    """
    Restituisce la configurazione campi per un template.
    Per ogni campo: readonly, includi_preventivo, mostra_default.
    """
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")

    field_config = {}
    if template.template_data:
        try:
            td = json.loads(template.template_data)
            field_config = td.get("field_config", {})
        except json.JSONDecodeError:
            pass

    try:
        campi = db.execute(text("""
            SELECT codice, etichetta, sezione, tipo, valore_default
            FROM campi_configuratore WHERE attivo = 1
            ORDER BY sezione, ordine
        """)).fetchall()
    except Exception:
        campi = []

    result = []
    for codice, etichetta, sezione, tipo, valore_default in campi:
        cfg = field_config.get(codice, {})
        result.append({
            "codice": codice,
            "etichetta": etichetta,
            "sezione": sezione,
            "tipo": tipo,
            "valore_default": valore_default,
            "readonly": cfg.get("readonly", False),
            "includi_preventivo": cfg.get("includi_preventivo", True),
            "mostra_default": cfg.get("mostra_default", False),
        })

    return {"template_id": template_id, "fields": result, "raw_config": field_config}


@app.put("/templates/{template_id}/field-config")
def update_template_field_config(template_id: int, data: dict, db: Session = Depends(get_db)):
    """
    Aggiorna la configurazione campi per un template.
    Body: {"field_config": {"campo": {"readonly": true, "includi_preventivo": true, "mostra_default": false}, ...}}
    """
    template = db.query(ProductTemplate).filter(ProductTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")

    new_field_config = data.get("field_config", {})

    existing_td = {}
    if template.template_data:
        try:
            existing_td = json.loads(template.template_data)
        except json.JSONDecodeError:
            existing_td = {}

    existing_td["field_config"] = new_field_config
    template.template_data = json.dumps(existing_td, ensure_ascii=False)
    db.commit()

    return {"status": "ok", "fields_configured": len(new_field_config)}


# ==========================================
# RULE ENGINE ENDPOINT
# ==========================================
@app.post("/preventivi/{preventivo_id}/evaluate-rules")
def evaluate_rules_endpoint(preventivo_id: int, db: Session = Depends(get_db)):
    return evaluate_rules(preventivo_id, db)

# ==========================================
# REGOLE JSON - CRUD su file in ./rules/
# ==========================================
RULES_DIR = "./rules"

@app.get("/regole")
def get_regole():
    """Lista tutte le regole JSON"""
    rules = load_rules()
    return sorted(rules, key=lambda r: r.get("priority", 99))

@app.get("/regole/{rule_id}")
def get_regola(rule_id: str):
    """Ottieni una regola per ID"""
    rules = load_rules()
    for r in rules:
        if r.get("id") == rule_id:
            return r
    raise HTTPException(status_code=404, detail=f"Regola {rule_id} non trovata")

@app.post("/regole/check-value-usage")
def check_value_usage(data: dict, db: Session = Depends(get_db)):
    """
    Controlla se un valore di un gruppo dropdown e' usato in condizioni di regole.
    Body: { "gruppo": "en_81_20_anno", "valore": "2020" }
    """
    gruppo = data.get("gruppo", "")
    valore = str(data.get("valore", ""))
    
    if not gruppo or not valore:
        return {"affected_rules": [], "count": 0}
    
    # Trova i codici campo che usano questo gruppo dropdown
    try:
        result = db.execute(text(
            "SELECT codice FROM campi_configuratore WHERE gruppo_dropdown = :g"
        ), {"g": gruppo})
        codici = [r[0] for r in result.fetchall()]
    except:
        codici = []
    
    # Genera varianti nome campo (regole possono usare con o senza prefisso sezione)
    field_variants = set(codici)
    for codice in codici:
        if "." in codice:
            field_variants.add(codice.split(".", 1)[1])
        field_variants.add(gruppo.replace("_anno", "").replace("_tipo", ""))
    
    # Cerca nelle regole
    rules = load_rules()
    affected = []
    for rule in rules:
        for cond in rule.get("conditions", []):
            if cond.get("field") in field_variants and str(cond.get("value")) == valore:
                affected.append({
                    "rule_id": rule.get("id"),
                    "rule_name": rule.get("name", rule.get("id")),
                    "enabled": rule.get("enabled", True),
                    "field": cond.get("field"),
                    "value": str(cond.get("value")),
                })
                break
    
    return {"affected_rules": affected, "count": len(affected), "field_variants": list(field_variants)}

@app.put("/regole/cascade-update-value")
def cascade_update_value(data: dict, db: Session = Depends(get_db)):
    """
    Aggiorna un valore nelle condizioni di tutte le regole impattate.
    Body: { "gruppo": "en_81_20_anno", "old_value": "2020", "new_value": "2020_rev" }
    """
    gruppo = data.get("gruppo", "")
    old_value = str(data.get("old_value", ""))
    new_value = str(data.get("new_value", ""))
    
    if not old_value or not new_value:
        raise HTTPException(status_code=400, detail="old_value e new_value obbligatori")
    
    # Stessa logica di field_variants
    try:
        result = db.execute(text(
            "SELECT codice FROM campi_configuratore WHERE gruppo_dropdown = :g"
        ), {"g": gruppo})
        codici = [r[0] for r in result.fetchall()]
    except:
        codici = []
    
    field_variants = set(codici)
    for codice in codici:
        if "." in codice:
            field_variants.add(codice.split(".", 1)[1])
        field_variants.add(gruppo.replace("_anno", "").replace("_tipo", ""))
    
    # Aggiorna
    rules = load_rules()
    updated_rules = []
    for rule in rules:
        modified = False
        for cond in rule.get("conditions", []):
            if cond.get("field") in field_variants and str(cond.get("value")) == old_value:
                cond["value"] = new_value
                modified = True
        if modified:
            rule_id = rule.get("id")
            filepath = os.path.join(RULES_DIR, f"{rule_id}.json")
            try:
                with open(filepath, "w", encoding="utf-8") as f:
                    json.dump(rule, f, indent=2, ensure_ascii=False)
                updated_rules.append(rule_id)
            except Exception as e:
                print(f"Errore aggiornamento regola {rule_id}: {e}")
    
    return {"status": "ok", "updated_rules": updated_rules, "count": len(updated_rules)}

@app.post("/regole")
def create_regola(data: dict):
    """Crea una nuova regola JSON"""
    rule_id = data.get("id")
    if not rule_id:
        raise HTTPException(status_code=400, detail="Campo 'id' obbligatorio")
    
    os.makedirs(RULES_DIR, exist_ok=True)
    filepath = os.path.join(RULES_DIR, f"{rule_id}.json")
    if os.path.exists(filepath):
        raise HTTPException(status_code=409, detail=f"Regola {rule_id} esiste gia")
    
    # Assicura campi obbligatori
    data.setdefault("enabled", True)
    data.setdefault("priority", 50)
    data.setdefault("conditions", [])
    data.setdefault("materials", [])
    data.setdefault("version", "1.0")
    
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    return {"status": "created", "id": rule_id}

@app.put("/regole/{rule_id}")
def update_regola(rule_id: str, data: dict):
    """Aggiorna una regola JSON esistente"""
    os.makedirs(RULES_DIR, exist_ok=True)
    
    # Cerca il file: prima match diretto, poi per ID interno
    filepath = os.path.join(RULES_DIR, f"{rule_id}.json")
    if not os.path.exists(filepath):
        for filename in os.listdir(RULES_DIR):
            if filename.endswith(".json"):
                fpath = os.path.join(RULES_DIR, filename)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        rule = json.load(f)
                    if rule.get("id") == rule_id:
                        filepath = fpath
                        break
                except Exception:
                    continue
    
    data["id"] = rule_id  # Forza ID consistente
    data.setdefault("enabled", True)
    data.setdefault("priority", 50)
    data.setdefault("conditions", [])
    data.setdefault("materials", [])
    
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    return {"status": "updated", "id": rule_id}

@app.delete("/regole/{rule_id}")
def delete_regola(rule_id: str):
    """Elimina una regola JSON cercando per ID interno"""
    print(f">>> DELETE chiamato per rule_id={rule_id}")
    filepath = os.path.join(RULES_DIR, f"{rule_id}.json")
    if os.path.exists(filepath):
        os.remove(filepath)
        return {"status": "deleted", "id": rule_id}
    if os.path.exists(RULES_DIR):
        for filename in os.listdir(RULES_DIR):
            if filename.endswith(".json"):
                fpath = os.path.join(RULES_DIR, filename)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        rule = json.load(f)
                    if rule.get("id") == rule_id:
                        os.remove(fpath)
                        return {"status": "deleted", "id": rule_id, "file": filename}
                except Exception:
                    continue
    raise HTTPException(status_code=404, detail=f"Regola {rule_id} non trovata")
    
    # Altrimenti cerca in tutti i file JSON per campo "id"
    if os.path.exists(RULES_DIR):
        for filename in os.listdir(RULES_DIR):
            if filename.endswith(".json"):
                fpath = os.path.join(RULES_DIR, filename)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        rule = json.load(f)
                    if rule.get("id") == rule_id:
                        os.remove(fpath)
                        return {"status": "deleted", "id": rule_id, "file": filename}
                except Exception:
                    continue
    
    raise HTTPException(status_code=404, detail=f"Regola {rule_id} non trovata")

@app.get("/regole-campi-disponibili")
def get_campi_disponibili(db: Session = Depends(get_db)):
    """
    Lista tutti i campi utilizzabili nelle condizioni delle regole, con tipo e opzioni.
    
    Approccio DYNAMIC-FIRST:
    1. Legge TUTTO da campi_configuratore (fonte di verità per campi sezione)
    2. Aggiunge campi template/preventivo (meta-campi, non in campi_configuratore)
    3. Aggiunge fallback per campi ORM legacy NON presenti in campi_configuratore
    """
    campi = []
    codici_gia_aggiunti = set()

    # Pre-carica mappa sezione codice → etichetta per nomi leggibili
    sez_labels = {}
    try:
        sez_rows = db.execute(text("SELECT codice, etichetta FROM sezioni_configuratore")).fetchall()
        sez_labels = {r[0]: r[1] for r in sez_rows}
    except Exception:
        pass

    def sez_display(codice_sezione):
        return sez_labels.get(codice_sezione, codice_sezione)

    def add_campo(field, source, label, tipo="testo", options=None):
        if field in codici_gia_aggiunti:
            return  # Evita duplicati
        codici_gia_aggiunti.add(field)
        entry = {"field": field, "source": source, "label": label, "type": tipo}
        if options:
            entry["options"] = options
        campi.append(entry)

    # =========================================================
    # STEP 1: Campi da campi_configuratore (FONTE DI VERITÀ)
    # =========================================================
    try:
        result = db.execute(text(
            "SELECT codice, etichetta, sezione, tipo, gruppo_dropdown "
            "FROM campi_configuratore WHERE attivo=1 ORDER BY sezione, ordine"
        ))
        for row in result.fetchall():
            codice, etichetta, sezione, tipo, gruppo = row[0], row[1], row[2], row[3], row[4]
            options = None
            if tipo == "dropdown" and gruppo:
                options = _load_dropdown_options(db, gruppo)
            add_campo(codice, sez_display(sezione), etichetta, tipo or "testo", options)
    except Exception:
        pass

    # =========================================================
    # STEP 2: Meta-campi TEMPLATE (non sono in campi_configuratore)
    # =========================================================
    try:
        cat_result = db.execute(text("SELECT DISTINCT categoria FROM product_templates WHERE attivo=1 ORDER BY categoria"))
        template_categorie = [r[0] for r in cat_result.fetchall() if r[0]]
    except Exception:
        template_categorie = ["RISE", "HOME"]
    try:
        sub_result = db.execute(text("SELECT DISTINCT sotto_categoria FROM product_templates WHERE attivo=1 ORDER BY sotto_categoria"))
        template_sottocategorie = [r[0] for r in sub_result.fetchall() if r[0]]
    except Exception:
        template_sottocategorie = []
    try:
        nome_result = db.execute(text("SELECT DISTINCT nome_display FROM product_templates WHERE attivo=1 ORDER BY nome_display"))
        template_nomi = [r[0] for r in nome_result.fetchall() if r[0]]
    except Exception:
        template_nomi = []
    try:
        id_result = db.execute(text("SELECT id, nome_display FROM product_templates WHERE attivo=1 ORDER BY id"))
        template_ids = [{"value": str(r[0]), "label": f"{r[0]} - {r[1]}"} for r in id_result.fetchall()]
    except Exception:
        template_ids = []

    add_campo("template_categoria", "Prodotto / Template", "Template Categoria", "dropdown", template_categorie)
    add_campo("template_sottocategoria", "Prodotto / Template", "Template Sottocategoria", "dropdown", template_sottocategorie)
    add_campo("template_nome", "Prodotto / Template", "Template Nome Display", "dropdown", template_nomi)
    add_campo("template_id", "Prodotto / Template", "Template ID", "dropdown", template_ids)
    # =========================================================
    # STEP 3: Meta-campi PREVENTIVO (non sono in campi_configuratore)
    # =========================================================
    add_campo("preventivo_tipo", "Preventivo", "Tipo Preventivo", "dropdown", ["COMPLETO", "RICAMBI"])
    add_campo("preventivo_categoria", "Preventivo", "Categoria Preventivo", "dropdown", template_categorie)
    # =========================================================
    # STEP 4: Fallback campi ORM legacy
    #   Solo per campi NON già definiti in campi_configuratore.
    #   Questi servono se il sistema ha tabelle ORM legacy
    #   non ancora migrate a campi_configuratore.
    # =========================================================
    
    # Dati principali
    legacy_dp = {
        "tipo_impianto": ("dropdown", "tipo_impianto"),
        "nuovo_impianto": ("booleano", None),
        "numero_fermate": ("numero", None),
        "numero_servizi": ("numero", None),
        "velocita": ("numero", None),
        "corsa": ("numero", None),
        "con_locale_macchina": ("booleano", None),
        "posizione_locale_macchina": ("dropdown", "posizione_locale_macchina"),
        "tipo_trazione": ("dropdown", "tipo_trazione"),
        "forza_motrice": ("dropdown", "forza_motrice"),
        "luce": ("numero", None),
        "tensione_manovra": ("dropdown", "tensione_manovra"),
        "tensione_freno": ("dropdown", "tensione_freno"),
    }
    for campo, (tipo, gruppo) in legacy_dp.items():
        if campo not in codici_gia_aggiunti:
            opts = _load_dropdown_options(db, gruppo) if gruppo else None
            add_campo(campo, sez_display("dati_principali"), campo.replace("_", " ").title(), tipo, opts)

    # Normative
    legacy_norm = {
        "en_81_1": ("dropdown", "en_81_1_anno", [{"value": "1998", "label": "1998"}, {"value": "2010", "label": "2010"}]),
        "en_81_20": ("dropdown", "en_81_20_anno", [{"value": "2014", "label": "2014"}, {"value": "2020", "label": "2020"}]),
        "en_81_21": ("dropdown", "en_81_21_anno", [{"value": "2009", "label": "2009"}, {"value": "2018", "label": "2018"}]),
        "en_81_28": ("booleano", None, None),
        "en_81_70": ("booleano", None, None),
        "en_81_72": ("booleano", None, None),
        "en_81_73": ("booleano", None, None),
        "a3_95_16": ("booleano", None, None),
        "dm236_legge13": ("booleano", None, None),
        "emendamento_a3": ("booleano", None, None),
        "uni_10411_1": ("booleano", None, None),
    }
    for campo, (tipo, gruppo, fallback_opts) in legacy_norm.items():
        if campo not in codici_gia_aggiunti:
            opts = None
            if gruppo:
                opts = _load_dropdown_options(db, gruppo) or fallback_opts
            add_campo(campo, sez_display("normative"), campo.replace("_", " ").upper(), tipo, opts)

    # Argano
    legacy_argano = {
        "trazione": ("dropdown", "trazione"),
        "potenza_motore_kw": ("numero", None),
        "corrente_nom_motore_amp": ("numero", None),
        "tipo_vvvf": ("dropdown", "tipo_vvvf"),
        "vvvf_nel_vano": ("booleano", None),
        "freno_tensione": ("numero", None),
        "ventilazione_forzata": ("booleano", None),
        "tipo_teleruttore": ("dropdown", "tipo_teleruttore"),
    }
    for campo, (tipo, gruppo) in legacy_argano.items():
        if campo not in codici_gia_aggiunti:
            opts = _load_dropdown_options(db, gruppo) if gruppo else None
            add_campo(campo, sez_display("argano"), campo.replace("_", " ").title(), tipo, opts)

    return campi


def _load_dropdown_options(db, gruppo: str):
    """Carica opzioni dropdown dal DB per un dato gruppo"""
    try:
        result = db.execute(text(
            "SELECT valore, etichetta FROM opzioni_dropdown "
            "WHERE gruppo = :g AND attivo = 1 ORDER BY ordine"
        ), {"g": gruppo})
        rows = result.fetchall()
        if rows:
            return [{"value": r[0], "label": r[1] or r[0]} for r in rows]
    except Exception:
        pass
    return None

# ==========================================
# CAMPI <-> REGOLE: USAGE MAP + RINOMINA
# ==========================================

def _scan_rules_fields() -> dict:
    """
    Scansiona tutti i file JSON in rules/ e restituisce:
      { "nome_campo": [ { rule_id, rule_name, file, enabled, usage, detail }, ... ] }
    """
    field_map = {}
    rules_dir = "./rules"
    if not os.path.exists(rules_dir):
        return field_map

    for filename in os.listdir(rules_dir):
        if not filename.endswith(".json"):
            continue
        try:
            filepath = os.path.join(rules_dir, filename)
            with open(filepath, "r", encoding="utf-8") as f:
                rule = json.load(f)

            rule_id = rule.get("id", filename)
            rule_name = rule.get("name", rule_id)
            enabled = rule.get("enabled", True)

            def _extract_conditions(conds):
                fields = []
                if isinstance(conds, list):
                    for c in conds:
                        if "field" in c:
                            fields.append((c["field"], c.get("operator", "?"), c.get("value", "?")))
                elif isinstance(conds, dict):
                    if "field" in conds:
                        fields.append((conds["field"], conds.get("operator", "?"), conds.get("value", "?")))
                    if "conditions" in conds:
                        fields.extend(_extract_conditions(conds["conditions"]))
                return fields

            for field, op, val in _extract_conditions(rule.get("conditions", [])):
                entry = {
                    "rule_id": rule_id, "rule_name": rule_name,
                    "file": filename, "enabled": enabled,
                    "usage": "condition",
                    "detail": f"{op} {val}",
                }
                field_map.setdefault(field, []).append(entry)

            for mat in rule.get("materials", []):
                for val in [mat.get("codice", ""), mat.get("descrizione", "")]:
                    for m in re.findall(r"\{\{(\w[\w.]*)\}\}", str(val)):
                        entry = {
                            "rule_id": rule_id, "rule_name": rule_name,
                            "file": filename, "enabled": enabled,
                            "usage": "placeholder",
                            "detail": f"in {mat.get('codice', '?')}",
                        }
                        field_map.setdefault(m, []).append(entry)

        except Exception:
            continue

    return field_map


@app.get("/campi-regole-usage")
def get_campi_regole_usage(db: Session = Depends(get_db)):
    """
    Per ogni campo restituisce info + regole che lo usano + preventivi con valori.
    Include anche campi che appaiono solo nelle regole (non nel DB).
    """
    # 1. Campi dal DB
    try:
        result = db.execute(text("""
            SELECT id, codice, etichetta, sezione, tipo, attivo, usabile_regole
            FROM campi_configuratore ORDER BY sezione, ordine
        """))
        campi = []
        for r in result.fetchall():
            campi.append({
                "id": r[0], "codice": r[1], "etichetta": r[2],
                "sezione": r[3], "tipo": r[4], "attivo": bool(r[5]),
                "usabile_regole": bool(r[6]) if r[6] is not None else True,
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    # 2. Mappa regole
    field_rules = _scan_rules_fields()

    # 3. Conteggio valori
    try:
        result = db.execute(text("""
            SELECT codice_campo, COUNT(DISTINCT preventivo_id) as n
            FROM valori_configurazione
            WHERE valore IS NOT NULL AND valore != ''
            GROUP BY codice_campo
        """))
        usage_count = {r[0]: r[1] for r in result.fetchall()}
    except Exception:
        usage_count = {}

    # 4. Assembla
    output = []
    for campo in campi:
        codice = campo["codice"]
        regole = list(field_rules.get(codice, []))
        if "." in codice:
            senza = codice.split(".", 1)[1]
            regole += field_rules.get(senza, [])
        # Deduplica
        seen = set()
        uniche = []
        for r in regole:
            key = (r["rule_id"], r["usage"])
            if key not in seen:
                seen.add(key)
                uniche.append(r)

        output.append({
            **campo,
            "regole": uniche,
            "num_regole": len(uniche),
            "num_preventivi": usage_count.get(codice, 0),
            "solo_regole": False,
        })

    # 5. Campi che esistono solo nelle regole (non in campi_configuratore)
    codici_db = {c["codice"] for c in campi}
    codici_senza = set()
    for c in campi:
        if "." in c["codice"]:
            codici_senza.add(c["codice"].split(".", 1)[1])

    for field, regole in field_rules.items():
        if field not in codici_db and field not in codici_senza:
            output.append({
                "id": None, "codice": field,
                "etichetta": field.replace("_", " ").title(),
                "sezione": field.split(".")[0] if "." in field else "???",
                "tipo": "???", "attivo": True, "usabile_regole": True,
                "regole": regole,
                "num_regole": len(regole),
                "num_preventivi": usage_count.get(field, 0),
                "solo_regole": True,
            })

    return output


@app.post("/campi-rinomina/preview")
def preview_rinomina(data: dict, db: Session = Depends(get_db)):
    """Anteprima impatto rinomina senza applicare."""
    vecchio = data.get("vecchio_codice", "").strip()
    nuovo = data.get("nuovo_codice", "").strip()
    if not vecchio or not nuovo:
        raise HTTPException(status_code=400, detail="vecchio_codice e nuovo_codice obbligatori")

    preview = {"vecchio": vecchio, "nuovo": nuovo, "impatto": []}

    # DB campi
    try:
        r = db.execute(text("SELECT COUNT(*) FROM campi_configuratore WHERE codice=:v"), {"v": vecchio}).fetchone()
        if r[0]:
            preview["impatto"].append({"area": "campi_configuratore", "desc": f"{r[0]} record", "tipo": "db"})
    except Exception:
        pass

    # Valori preventivi
    try:
        r = db.execute(text("""
            SELECT COUNT(*), COUNT(DISTINCT preventivo_id)
            FROM valori_configurazione WHERE codice_campo=:v
        """), {"v": vecchio}).fetchone()
        if r[0]:
            preview["impatto"].append({"area": "valori_configurazione", "desc": f"{r[0]} valori in {r[1]} preventivi", "tipo": "db"})
    except Exception:
        pass

    # Opzioni dropdown
    try:
        r = db.execute(text("SELECT COUNT(*) FROM opzioni_dropdown WHERE gruppo=:v"), {"v": vecchio}).fetchone()
        if r[0]:
            preview["impatto"].append({"area": "opzioni_dropdown", "desc": f"{r[0]} opzioni nel gruppo", "tipo": "db"})
    except Exception:
        pass

    # Regole JSON
    field_rules = _scan_rules_fields()
    vecchio_senza = vecchio.split(".", 1)[1] if "." in vecchio else vecchio
    files = set()
    for f in [vecchio, vecchio_senza]:
        for entry in field_rules.get(f, []):
            files.add(entry["file"])
    if files:
        preview["impatto"].append({"area": "regole_json", "desc": f"{len(files)} file regola", "tipo": "file", "files": sorted(files)})

    return preview


@app.post("/campi-rinomina")
def rinomina_campo(data: dict, db: Session = Depends(get_db)):
    """
    Rinomina campo su tutti i punti:
      1. campi_configuratore.codice
      2. valori_configurazione.codice_campo + sezione
      3. opzioni_dropdown.gruppo
      4. File JSON regole (conditions + placeholders)
    """
    vecchio = data.get("vecchio_codice", "").strip()
    nuovo = data.get("nuovo_codice", "").strip()
    if not vecchio or not nuovo:
        raise HTTPException(status_code=400, detail="vecchio_codice e nuovo_codice obbligatori")
    if vecchio == nuovo:
        raise HTTPException(status_code=400, detail="Codici identici")

    report = {"vecchio": vecchio, "nuovo": nuovo,
              "campi_configuratore": 0, "valori_configurazione": 0,
              "opzioni_dropdown": 0, "regole_aggiornate": [], "errori": []}

    # 1. campi_configuratore
    try:
        r = db.execute(text("UPDATE campi_configuratore SET codice=:n WHERE codice=:v"), {"v": vecchio, "n": nuovo})
        report["campi_configuratore"] = r.rowcount
        if "." in nuovo:
            nuova_sez = nuovo.split(".", 1)[0]
            db.execute(text("UPDATE campi_configuratore SET sezione=:s WHERE codice=:c"), {"s": nuova_sez, "c": nuovo})
    except Exception as e:
        report["errori"].append(f"campi_configuratore: {e}")

    # 2. valori_configurazione
    try:
        r = db.execute(text("UPDATE valori_configurazione SET codice_campo=:n WHERE codice_campo=:v"), {"v": vecchio, "n": nuovo})
        report["valori_configurazione"] = r.rowcount
        if "." in nuovo:
            nuova_sez = nuovo.split(".", 1)[0]
            db.execute(text("UPDATE valori_configurazione SET sezione=:s WHERE codice_campo=:c"), {"s": nuova_sez, "c": nuovo})
    except Exception as e:
        report["errori"].append(f"valori_configurazione: {e}")

    # 3. opzioni_dropdown
    try:
        r = db.execute(text("UPDATE opzioni_dropdown SET gruppo=:n WHERE gruppo=:v"), {"v": vecchio, "n": nuovo})
        report["opzioni_dropdown"] = r.rowcount
    except Exception as e:
        report["errori"].append(f"opzioni_dropdown: {e}")

    # 4. Regole JSON
    rules_dir = "./rules"
    if os.path.exists(rules_dir):
        vecchio_senza = vecchio.split(".", 1)[1] if "." in vecchio else vecchio
        nuovo_senza = nuovo.split(".", 1)[1] if "." in nuovo else nuovo

        for filename in os.listdir(rules_dir):
            if not filename.endswith(".json"):
                continue
            filepath = os.path.join(rules_dir, filename)
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    content = f.read()
                if vecchio not in content and vecchio_senza not in content:
                    continue

                rule = json.loads(content)
                mod = False

                def _upd(conds):
                    nonlocal mod
                    if isinstance(conds, list):
                        for c in conds:
                            if c.get("field") == vecchio:
                                c["field"] = nuovo; mod = True
                            elif c.get("field") == vecchio_senza:
                                c["field"] = nuovo_senza; mod = True
                    elif isinstance(conds, dict):
                        if conds.get("field") == vecchio:
                            conds["field"] = nuovo; mod = True
                        elif conds.get("field") == vecchio_senza:
                            conds["field"] = nuovo_senza; mod = True
                        if "conditions" in conds:
                            _upd(conds["conditions"])

                _upd(rule.get("conditions", []))

                for mat in rule.get("materials", []):
                    for key in ["codice", "descrizione"]:
                        if key in mat and isinstance(mat[key], str):
                            for old_f, new_f in [(vecchio, nuovo), (vecchio_senza, nuovo_senza)]:
                                old_ph = "{{" + old_f + "}}"
                                new_ph = "{{" + new_f + "}}"
                                if old_ph in mat[key]:
                                    mat[key] = mat[key].replace(old_ph, new_ph)
                                    mod = True

                if mod:
                    with open(filepath, "w", encoding="utf-8") as f:
                        json.dump(rule, f, indent=2, ensure_ascii=False)
                    report["regole_aggiornate"].append(filename)

            except Exception as e:
                report["errori"].append(f"{filename}: {e}")

    # 5. Template JSON — template_data keys e field_config keys
    try:
        templates = db.execute(text("SELECT id, template_data FROM product_templates WHERE template_data IS NOT NULL")).fetchall()
        report["templates_aggiornati"] = 0
        vecchio_sez = vecchio.split(".", 1)[0] if "." in vecchio else ""
        vecchio_campo_only = vecchio.split(".", 1)[1] if "." in vecchio else vecchio
        nuovo_sez = nuovo.split(".", 1)[0] if "." in nuovo else ""
        nuovo_campo_only = nuovo.split(".", 1)[1] if "." in nuovo else nuovo

        for tmpl_id, tmpl_json in templates:
            if not tmpl_json:
                continue
            tmpl_data = json.loads(tmpl_json)
            changed = False

            # Rinomina campo nelle sezioni dati (es: templateData.argano.trazione → templateData.argano.nuovo_nome)
            if vecchio_sez and vecchio_sez in tmpl_data:
                sez_data = tmpl_data[vecchio_sez]
                if isinstance(sez_data, dict) and vecchio_campo_only in sez_data:
                    sez_data[nuovo_campo_only] = sez_data.pop(vecchio_campo_only)
                    # Se la sezione è cambiata, sposta
                    if nuovo_sez != vecchio_sez:
                        if nuovo_sez not in tmpl_data:
                            tmpl_data[nuovo_sez] = {}
                        tmpl_data[nuovo_sez][nuovo_campo_only] = sez_data.pop(nuovo_campo_only, tmpl_data[nuovo_sez].get(nuovo_campo_only))
                    changed = True

            # Rinomina chiave in field_config
            if "field_config" in tmpl_data:
                fc = tmpl_data["field_config"]
                if vecchio in fc:
                    fc[nuovo] = fc.pop(vecchio)
                    changed = True

            if changed:
                db.execute(text("UPDATE product_templates SET template_data=:data WHERE id=:id"),
                    {"data": json.dumps(tmpl_data), "id": tmpl_id})
                report["templates_aggiornati"] += 1
    except Exception as e:
        report["errori"].append(f"templates: {e}")

    db.commit()
    return report

# ==========================================
# ARTICOLI
# ==========================================
@app.get("/articoli")
def get_articoli(
    search: str = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Lista articoli con ricerca opzionale"""
    query = db.query(Articolo)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Articolo.codice.ilike(search_term),
                Articolo.descrizione.ilike(search_term)
            )
        )
    articoli = query.order_by(Articolo.codice).offset(offset).limit(limit).all()
    return [{
        "id": a.id, "codice": a.codice, "descrizione": a.descrizione,
        "descrizione_estesa": a.descrizione_estesa,
        "tipo_articolo": a.tipo_articolo, "categoria_id": a.categoria_id,
        "costo_fisso": a.costo_fisso,
        "costo_variabile_1": a.costo_variabile_1, "unita_misura_var_1": a.unita_misura_var_1,
        "descrizione_var_1": a.descrizione_var_1,
        "costo_variabile_2": a.costo_variabile_2, "unita_misura_var_2": a.unita_misura_var_2,
        "descrizione_var_2": a.descrizione_var_2,
        "costo_variabile_3": a.costo_variabile_3, "unita_misura_var_3": a.unita_misura_var_3,
        "descrizione_var_3": a.descrizione_var_3,
        "costo_variabile_4": a.costo_variabile_4, "unita_misura_var_4": a.unita_misura_var_4,
        "descrizione_var_4": a.descrizione_var_4,
        "ricarico_percentuale": a.ricarico_percentuale,
        "unita_misura": a.unita_misura,
        "fornitore": a.fornitore, "codice_fornitore": a.codice_fornitore,
        "is_active": a.is_active
    } for a in articoli]

@app.get("/articoli/search")
def search_articoli(q: str = "", limit: int = 10, db: Session = Depends(get_db)):
    """Ricerca articoli per codice o descrizione"""
    if not q:
        return []
    search_term = f"%{q}%"
    articoli = db.query(Articolo).filter(
        or_(
            Articolo.codice.ilike(search_term),
            Articolo.descrizione.ilike(search_term)
        )
    ).limit(limit).all()
    return [{
        "id": a.id, "codice": a.codice, "descrizione": a.descrizione,
        "tipo_articolo": a.tipo_articolo, "costo_fisso": a.costo_fisso,
        "costo_variabile_1": a.costo_variabile_1, "unita_misura_var_1": a.unita_misura_var_1,
        "descrizione_var_1": a.descrizione_var_1,
        "costo_variabile_2": a.costo_variabile_2, "unita_misura_var_2": a.unita_misura_var_2,
        "descrizione_var_2": a.descrizione_var_2,
        "costo_variabile_3": a.costo_variabile_3, "unita_misura_var_3": a.unita_misura_var_3,
        "descrizione_var_3": a.descrizione_var_3,
        "costo_variabile_4": a.costo_variabile_4, "unita_misura_var_4": a.unita_misura_var_4,
        "descrizione_var_4": a.descrizione_var_4,
        "ricarico_percentuale": a.ricarico_percentuale,
        "unita_misura": a.unita_misura
    } for a in articoli]

@app.post("/articoli")
def create_articolo(data: dict, db: Session = Depends(get_db)):
    articolo = Articolo()
    for key in [
        "codice", "descrizione", "descrizione_estesa", "tipo_articolo", "categoria_id",
        "costo_fisso", "costo_variabile_1", "unita_misura_var_1", "descrizione_var_1",
        "costo_variabile_2", "unita_misura_var_2", "descrizione_var_2",
        "costo_variabile_3", "unita_misura_var_3", "descrizione_var_3",
        "costo_variabile_4", "unita_misura_var_4", "descrizione_var_4",
        "ricarico_percentuale", "unita_misura", "fornitore", "codice_fornitore",
        "giacenza", "scorta_minima", "lead_time_giorni", "manodopera_giorni", "is_active"
    ]:
        if key in data:
            setattr(articolo, key, data[key])
    db.add(articolo)
    db.commit()
    db.refresh(articolo)
    return {"id": articolo.id, "status": "ok"}

@app.put("/articoli/{articolo_id}")
def update_articolo(articolo_id: int, data: dict, db: Session = Depends(get_db)):
    articolo = db.query(Articolo).filter(Articolo.id == articolo_id).first()
    if not articolo:
        raise HTTPException(status_code=404, detail="Articolo non trovato")
    for key in [
        "codice", "descrizione", "descrizione_estesa", "tipo_articolo", "categoria_id",
        "costo_fisso", "costo_variabile_1", "unita_misura_var_1", "descrizione_var_1",
        "costo_variabile_2", "unita_misura_var_2", "descrizione_var_2",
        "costo_variabile_3", "unita_misura_var_3", "descrizione_var_3",
        "costo_variabile_4", "unita_misura_var_4", "descrizione_var_4",
        "ricarico_percentuale", "unita_misura", "fornitore", "codice_fornitore",
        "giacenza", "scorta_minima", "lead_time_giorni", "manodopera_giorni", "is_active"
    ]:
        if key in data:
            setattr(articolo, key, data[key])
    articolo.updated_at = datetime.now()
    db.commit()
    return {"id": articolo.id, "status": "ok"}

@app.delete("/articoli/{articolo_id}")
def delete_articolo(articolo_id: int, db: Session = Depends(get_db)):
    articolo = db.query(Articolo).filter(Articolo.id == articolo_id).first()
    if not articolo:
        raise HTTPException(status_code=404, detail="Articolo non trovato")
    db.delete(articolo)
    db.commit()
    return {"status": "ok"}

@app.post("/articoli/calcola-prezzo")
def calcola_prezzo_articolo(data: dict, db: Session = Depends(get_db)):
    """Calcola prezzo di un articolo in base ai parametri variabili"""
    costo_fisso = data.get("costo_fisso", 0) or 0
    costo_totale = costo_fisso
    
    for i in range(1, 5):
        costo_var = data.get(f"costo_variabile_{i}", 0) or 0
        parametro = data.get(f"parametro_{i}", 0) or 0
        costo_totale += costo_var * parametro
    
    ricarico = data.get("ricarico_percentuale", 0) or 0
    prezzo_listino = costo_totale * (1 + ricarico / 100) if ricarico else costo_totale
    
    sconto = data.get("sconto_cliente", 0) or 0
    prezzo_cliente = prezzo_listino * (1 - sconto / 100) if sconto else prezzo_listino
    
    return {
        "costo_base": costo_totale,
        "prezzo_listino": round(prezzo_listino, 4),
        "prezzo_cliente": round(prezzo_cliente, 4)
    }

# --- BOM: Distinta Base ---

def _detect_bom_schema(db: Session) -> str:
    """
    Rileva se bom_struttura usa lo schema vecchio (padre_codice/figlio_codice TEXT)
    o quello nuovo (articolo_padre_id/articolo_figlio_id INT).
    """
    try:
        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(bom_struttura)")
        columns = {row[1] for row in cursor.fetchall()}
        if 'articolo_padre_id' in columns and 'articolo_figlio_id' in columns:
            return 'id'
        return 'codice'
    except Exception:
        return 'id'


@app.get("/bom/counts")
def get_bom_counts(db: Session = Depends(get_db)):
    """Conta i figli per ogni articolo padre"""
    schema = _detect_bom_schema(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()

    if schema == 'id':
        rows = db.execute(text(
            "SELECT articolo_padre_id, COUNT(*) as cnt FROM bom_struttura GROUP BY articolo_padre_id"
        )).fetchall()
        return {row[0]: row[1] for row in rows}
    else:
        # Schema vecchio: padre_codice â†’ mappa a ID articolo dalla tabella articoli
        cursor.execute(
            "SELECT bs.padre_codice, COUNT(*) as cnt "
            "FROM bom_struttura bs GROUP BY bs.padre_codice"
        )
        counts_by_codice = {row[0]: row[1] for row in cursor.fetchall()}
        if not counts_by_codice:
            return {}

        # Mappa codice â†’ articoli.id
        result = {}
        placeholders = ','.join(['?'] * len(counts_by_codice))
        cursor.execute(
            f"SELECT id, codice FROM articoli WHERE codice IN ({placeholders})",
            list(counts_by_codice.keys())
        )
        for row in cursor.fetchall():
            art_id, codice = row[0], row[1]
            if codice in counts_by_codice:
                result[art_id] = counts_by_codice[codice]
        return result


@app.get("/bom/{articolo_padre_id}")
def get_bom(articolo_padre_id: int, db: Session = Depends(get_db)):
    """Lista componenti di un articolo padre"""
    schema = _detect_bom_schema(db)

    if schema == 'id':
        # â”€â”€ Schema nuovo: SQLAlchemy con articolo_padre_id/articolo_figlio_id â”€â”€
        righe = db.query(BomStruttura).filter(
            BomStruttura.articolo_padre_id == articolo_padre_id
        ).order_by(BomStruttura.ordine, BomStruttura.id).all()

        result = []
        for r in righe:
            figlio = db.query(Articolo).filter(Articolo.id == r.articolo_figlio_id).first()
            result.append({
                "id": r.id,
                "articolo_padre_id": r.articolo_padre_id,
                "articolo_figlio_id": r.articolo_figlio_id,
                "articolo_figlio": {
                    "id": figlio.id, "codice": figlio.codice,
                    "descrizione": figlio.descrizione,
                    "tipo_articolo": figlio.tipo_articolo,
                    "costo_fisso": figlio.costo_fisso,
                    "unita_misura": figlio.unita_misura,
                    "fornitore": getattr(figlio, 'fornitore', None),
                    "is_active": figlio.is_active,
                } if figlio else None,
                "quantita": r.quantita,
                "formula_quantita": r.formula_quantita,
                "unita_misura": r.unita_misura,
                "condizione_esistenza": r.condizione_esistenza,
                "note": r.note,
                "ordine": r.ordine,
            })
        return result

    else:
        # â”€â”€ Schema vecchio: padre_codice/figlio_codice (TEXT) â”€â”€
        # 1. Trova il codice dell'articolo padre dalla tabella articoli
        articolo_padre = db.query(Articolo).filter(Articolo.id == articolo_padre_id).first()
        if not articolo_padre:
            return []

        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()

        # 2. Query bom_struttura con padre_codice
        cursor.execute(
            "SELECT bs.id, bs.padre_codice, bs.figlio_codice, bs.quantita, "
            "bs.formula_quantita, bs.condizione_esistenza, "
            "COALESCE(bs.note, '') as note, "
            "COALESCE(bs.posizione, 0) as ordine "
            "FROM bom_struttura bs "
            "WHERE bs.padre_codice = ? "
            "ORDER BY ordine, bs.id",
            (articolo_padre.codice,)
        )
        rows = cursor.fetchall()

        result = []
        for r in rows:
            bs_id, padre_cod, figlio_cod, qta, formula, condizione, note, ordine = r

            # 3. Cerca figlio prima in 'articoli' (SQLAlchemy), poi in 'articoli_bom' (raw)
            figlio = db.query(Articolo).filter(Articolo.codice == figlio_cod).first()
            figlio_data = None

            if figlio:
                figlio_data = {
                    "id": figlio.id,
                    "codice": figlio.codice,
                    "descrizione": figlio.descrizione,
                    "tipo_articolo": figlio.tipo_articolo,
                    "costo_fisso": figlio.costo_fisso,
                    "unita_misura": figlio.unita_misura,
                    "fornitore": getattr(figlio, 'fornitore', None),
                    "is_active": figlio.is_active,
                }
            else:
                # Fallback: cerca in articoli_bom (tabella creata da migrate_prototipo)
                cursor.execute(
                    "SELECT codice, descrizione, tipo, categoria, costo_fisso, "
                    "unita_misura FROM articoli_bom WHERE codice = ?",
                    (figlio_cod,)
                )
                ab = cursor.fetchone()
                if ab:
                    figlio_data = {
                        "id": 0,
                        "codice": ab[0],
                        "descrizione": ab[1],
                        "tipo_articolo": ab[2] or "ACQUISTO",
                        "costo_fisso": ab[4] or 0,
                        "unita_misura": ab[5] or "PZ",
                        "fornitore": None,
                        "is_active": True,
                    }

            result.append({
                "id": bs_id,
                "articolo_padre_id": articolo_padre_id,
                "articolo_figlio_id": figlio.id if figlio else 0,
                "articolo_figlio": figlio_data,
                "quantita": qta or 1,
                "formula_quantita": formula,
                "unita_misura": "PZ",
                "condizione_esistenza": condizione,
                "note": note,
                "ordine": ordine or 0,
            })

        return result


@app.post("/bom")
def create_bom_riga(data: dict, db: Session = Depends(get_db)):
    """Aggiunge un componente alla distinta"""
    padre_id = data.get("articolo_padre_id")
    figlio_id = data.get("articolo_figlio_id")

    if not padre_id or not figlio_id:
        raise HTTPException(400, "articolo_padre_id e articolo_figlio_id obbligatori")
    if padre_id == figlio_id:
        raise HTTPException(400, "Un articolo non puÃ² contenere sÃ© stesso")

    schema = _detect_bom_schema(db)

    if schema == 'id':
        exists = db.query(BomStruttura).filter(
            BomStruttura.articolo_padre_id == padre_id,
            BomStruttura.articolo_figlio_id == figlio_id,
        ).first()
        if exists:
            raise HTTPException(400, "Questo componente Ã¨ giÃ  nella distinta")

        def is_ancestor(potential_ancestor_id, target_id, visited=None):
            if visited is None:
                visited = set()
            if potential_ancestor_id in visited:
                return False
            visited.add(potential_ancestor_id)
            parents = db.query(BomStruttura.articolo_padre_id).filter(
                BomStruttura.articolo_figlio_id == target_id
            ).all()
            for (pid,) in parents:
                if pid == potential_ancestor_id:
                    return True
                if is_ancestor(potential_ancestor_id, pid, visited):
                    return True
            return False

        if is_ancestor(figlio_id, padre_id):
            raise HTTPException(400, "Riferimento circolare")

        max_ord = db.query(BomStruttura).filter(
            BomStruttura.articolo_padre_id == padre_id
        ).count()

        riga = BomStruttura(
            articolo_padre_id=padre_id,
            articolo_figlio_id=figlio_id,
            quantita=data.get("quantita", 1),
            formula_quantita=data.get("formula_quantita"),
            unita_misura=data.get("unita_misura", "PZ"),
            condizione_esistenza=data.get("condizione_esistenza"),
            note=data.get("note"),
            ordine=max_ord + 1,
        )
        db.add(riga)
        db.commit()
        db.refresh(riga)
        return {"id": riga.id, "status": "ok"}

    else:
        # Schema vecchio: padre_codice/figlio_codice
        padre = db.query(Articolo).filter(Articolo.id == padre_id).first()
        figlio = db.query(Articolo).filter(Articolo.id == figlio_id).first()
        if not padre or not figlio:
            raise HTTPException(400, "Articolo padre o figlio non trovato")

        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT COUNT(*) FROM bom_struttura WHERE padre_codice=? AND figlio_codice=?",
            (padre.codice, figlio.codice)
        )
        if cursor.fetchone()[0] > 0:
            raise HTTPException(400, "Questo componente Ã¨ giÃ  nella distinta")

        cursor.execute(
            "SELECT COALESCE(MAX(posizione), 0) FROM bom_struttura WHERE padre_codice=?",
            (padre.codice,)
        )
        next_pos = (cursor.fetchone()[0] or 0) + 1

        cursor.execute(
            "INSERT INTO bom_struttura (padre_codice, figlio_codice, quantita, "
            "formula_quantita, condizione_esistenza, note, posizione, obbligatorio) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 1)",
            (padre.codice, figlio.codice, data.get("quantita", 1),
             data.get("formula_quantita"), data.get("condizione_esistenza"),
             data.get("note"), next_pos)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "status": "ok"}


@app.put("/bom/{riga_id}")
def update_bom_riga(riga_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna una riga della distinta"""
    schema = _detect_bom_schema(db)

    if schema == 'id':
        riga = db.query(BomStruttura).filter(BomStruttura.id == riga_id).first()
        if not riga:
            raise HTTPException(404, "Riga BOM non trovata")
        for key in ["quantita", "formula_quantita", "unita_misura", "condizione_esistenza", "note", "ordine"]:
            if key in data:
                setattr(riga, key, data[key])
        db.commit()
    else:
        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()
        updates = []
        values = []
        for key in ["quantita", "formula_quantita", "condizione_esistenza", "note"]:
            if key in data:
                updates.append(f"{key} = ?")
                values.append(data[key])
        if "ordine" in data:
            updates.append("posizione = ?")
            values.append(data["ordine"])
        if not updates:
            return {"id": riga_id, "status": "ok"}
        values.append(riga_id)
        cursor.execute(f"UPDATE bom_struttura SET {', '.join(updates)} WHERE id = ?", values)
        if cursor.rowcount == 0:
            raise HTTPException(404, "Riga BOM non trovata")
        conn.commit()

    return {"id": riga_id, "status": "ok"}


@app.delete("/bom/{riga_id}")
def delete_bom_riga(riga_id: int, db: Session = Depends(get_db)):
    """Rimuove una riga dalla distinta"""
    schema = _detect_bom_schema(db)

    if schema == 'id':
        riga = db.query(BomStruttura).filter(BomStruttura.id == riga_id).first()
        if not riga:
            raise HTTPException(404, "Riga BOM non trovata")
        db.delete(riga)
        db.commit()
    else:
        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM bom_struttura WHERE id = ?", (riga_id,))
        if cursor.rowcount == 0:
            raise HTTPException(404, "Riga BOM non trovata")
        conn.commit()

    return {"status": "ok"}


@app.post("/bom/duplica")
def duplica_bom(data: dict, db: Session = Depends(get_db)):
    """Duplica l'intera BOM da un articolo sorgente a un articolo destinazione"""
    source_id = data.get("source_articolo_id")
    target_id = data.get("target_articolo_id")

    if not source_id or not target_id:
        raise HTTPException(400, "source_articolo_id e target_articolo_id obbligatori")
    if source_id == target_id:
        raise HTTPException(400, "Sorgente e destinazione devono essere diversi")

    schema = _detect_bom_schema(db)

    if schema == 'id':
        db.query(BomStruttura).filter(BomStruttura.articolo_padre_id == target_id).delete()
        righe_source = db.query(BomStruttura).filter(
            BomStruttura.articolo_padre_id == source_id
        ).order_by(BomStruttura.ordine).all()
        for r in righe_source:
            nuova = BomStruttura(
                articolo_padre_id=target_id,
                articolo_figlio_id=r.articolo_figlio_id,
                quantita=r.quantita,
                formula_quantita=r.formula_quantita,
                unita_misura=r.unita_misura,
                condizione_esistenza=r.condizione_esistenza,
                note=r.note,
                ordine=r.ordine,
            )
            db.add(nuova)
        db.commit()
        return {"status": "ok", "righe_duplicate": len(righe_source)}

    else:
        source = db.query(Articolo).filter(Articolo.id == source_id).first()
        target = db.query(Articolo).filter(Articolo.id == target_id).first()
        if not source or not target:
            raise HTTPException(400, "Articoli non trovati")

        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM bom_struttura WHERE padre_codice = ?", (target.codice,))
        cursor.execute(
            "SELECT figlio_codice, quantita, formula_quantita, condizione_esistenza, "
            "note, posizione, obbligatorio FROM bom_struttura WHERE padre_codice = ? ORDER BY posizione",
            (source.codice,)
        )
        righe = cursor.fetchall()
        for r in righe:
            cursor.execute(
                "INSERT INTO bom_struttura (padre_codice, figlio_codice, quantita, "
                "formula_quantita, condizione_esistenza, note, posizione, obbligatorio) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (target.codice, r[0], r[1], r[2], r[3], r[4], r[5], r[6])
            )
        conn.commit()
        return {"status": "ok", "righe_duplicate": len(righe)}

# ==========================================
# CATEGORIE ARTICOLI
# ==========================================
@app.get("/categorie-articoli")
def get_categorie_articoli(db: Session = Depends(get_db)):
    categorie = db.query(CategoriaArticoli).order_by(CategoriaArticoli.ordine).all()
    return [{
        "id": c.id, "codice": c.codice, "nome": c.nome,
        "descrizione": c.descrizione, "ordine": c.ordine, "is_active": c.is_active
    } for c in categorie]

# ==========================================
# CLIENTI
# ==========================================
@app.get("/clienti")
def get_clienti(db: Session = Depends(get_db)):
    clienti = db.query(Cliente).filter(Cliente.is_active == True).order_by(Cliente.ragione_sociale).all()
    return [{
        "id": c.id, "codice": c.codice, "ragione_sociale": c.ragione_sociale,
        "partita_iva": c.partita_iva, "codice_fiscale": c.codice_fiscale,
        "indirizzo": c.indirizzo, "cap": c.cap, "citta": c.citta,
        "provincia": c.provincia, "nazione": c.nazione,
        "telefono": c.telefono, "email": c.email, "pec": c.pec,
        "sconto_globale": c.sconto_globale, "sconto_produzione": c.sconto_produzione,
        "sconto_acquisto": c.sconto_acquisto, "aliquota_iva": c.aliquota_iva,
        "pagamento_default": c.pagamento_default, "imballo_default": c.imballo_default,
        "reso_fco_default": c.reso_fco_default, "trasporto_default": c.trasporto_default,
        "destinazione_default": c.destinazione_default,
        "riferimento_cliente_default": c.riferimento_cliente_default,
        "listino": c.listino, "note": c.note, "is_active": c.is_active
    } for c in clienti]

@app.get("/clienti/search")
def search_clienti(q: str = "", db: Session = Depends(get_db)):
    if not q:
        return []
    search_term = f"%{q}%"
    clienti = db.query(Cliente).filter(
        or_(
            Cliente.codice.ilike(search_term),
            Cliente.ragione_sociale.ilike(search_term),
            Cliente.partita_iva.ilike(search_term)
        )
    ).limit(20).all()
    return [{
        "id": c.id, "codice": c.codice, "ragione_sociale": c.ragione_sociale,
        "partita_iva": c.partita_iva, "citta": c.citta, "provincia": c.provincia,
        "sconto_globale": c.sconto_globale, "sconto_produzione": c.sconto_produzione,
        "sconto_acquisto": c.sconto_acquisto,
        "pagamento_default": c.pagamento_default, "imballo_default": c.imballo_default,
        "reso_fco_default": c.reso_fco_default, "trasporto_default": c.trasporto_default,
        "destinazione_default": c.destinazione_default,
        "riferimento_cliente_default": c.riferimento_cliente_default
    } for c in clienti]

@app.get("/clienti/{cliente_id}")
def get_cliente(cliente_id: int, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    return {
        "id": cliente.id, "codice": cliente.codice, "ragione_sociale": cliente.ragione_sociale,
        "partita_iva": cliente.partita_iva, "codice_fiscale": cliente.codice_fiscale,
        "indirizzo": cliente.indirizzo, "cap": cliente.cap, "citta": cliente.citta,
        "provincia": cliente.provincia, "nazione": cliente.nazione,
        "telefono": cliente.telefono, "email": cliente.email, "pec": cliente.pec,
        "sconto_globale": cliente.sconto_globale, "sconto_produzione": cliente.sconto_produzione,
        "sconto_acquisto": cliente.sconto_acquisto, "aliquota_iva": cliente.aliquota_iva,
        "pagamento_default": cliente.pagamento_default, "imballo_default": cliente.imballo_default,
        "reso_fco_default": cliente.reso_fco_default, "trasporto_default": cliente.trasporto_default,
        "destinazione_default": cliente.destinazione_default,
        "riferimento_cliente_default": cliente.riferimento_cliente_default,
        "listino": cliente.listino, "note": cliente.note, "is_active": cliente.is_active
    }

@app.post("/clienti")
def create_cliente(data: dict, db: Session = Depends(get_db)):
    cliente = Cliente()
    for key in [
        "codice", "ragione_sociale", "partita_iva", "codice_fiscale",
        "indirizzo", "cap", "citta", "provincia", "nazione",
        "telefono", "email", "pec",
        "sconto_globale", "sconto_produzione", "sconto_acquisto", "aliquota_iva",
        "pagamento_default", "imballo_default", "reso_fco_default",
        "trasporto_default", "destinazione_default", "riferimento_cliente_default",
        "listino", "note", "is_active"
    ]:
        if key in data:
            setattr(cliente, key, data[key])
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return {"id": cliente.id, "status": "ok"}

@app.put("/clienti/{cliente_id}")
def update_cliente(cliente_id: int, data: dict, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    for key in [
        "codice", "ragione_sociale", "partita_iva", "codice_fiscale",
        "indirizzo", "cap", "citta", "provincia", "nazione",
        "telefono", "email", "pec",
        "sconto_globale", "sconto_produzione", "sconto_acquisto", "aliquota_iva",
        "pagamento_default", "imballo_default", "reso_fco_default",
        "trasporto_default", "destinazione_default", "riferimento_cliente_default",
        "listino", "note", "is_active"
    ]:
        if key in data:
            setattr(cliente, key, data[key])
    cliente.updated_at = datetime.now()
    db.commit()
    return {"id": cliente.id, "status": "ok"}

# ==========================================
# UTENTI
# ==========================================
@app.get("/utenti")
def get_utenti(db: Session = Depends(get_db)):
    utenti = db.query(Utente).order_by(Utente.username).all()
    result = []
    for u in utenti:
        gruppo_nome = None
        if u.gruppo_id:
            g = db.query(GruppoUtenti).filter(GruppoUtenti.id == u.gruppo_id).first()
            gruppo_nome = g.nome if g else None
        ruolo_nome = None
        ruolo_codice = None
        if u.ruolo_id:
            r = db.query(Ruolo).filter(Ruolo.id == u.ruolo_id).first()
            if r:
                ruolo_nome = r.nome
                ruolo_codice = r.codice
        result.append({
            "id": u.id,
            "username": u.username,
            "nome": u.nome,
            "cognome": u.cognome,
            "email": u.email,
            "gruppo_id": u.gruppo_id,
            "gruppo_nome": gruppo_nome,
            "ruolo_id": u.ruolo_id,
            "ruolo_nome": ruolo_nome,
            "ruolo_codice": ruolo_codice,
            "is_admin": u.is_admin,
            "is_active": u.is_active,
            "created_at": str(u.created_at) if u.created_at else None,
            "last_login": str(u.last_login) if u.last_login else None,
        })
    return result

@app.post("/utenti")
def create_utente(
    username: str = Query(...),
    password: str = Query(None),
    nome: str = Query(None),
    cognome: str = Query(None),
    email: str = Query(None),
    gruppo_id: int = Query(None),
    ruolo_id: int = Query(None),
    is_admin: bool = Query(False),
    is_active: bool = Query(True),
    db: Session = Depends(get_db)
):
    existing = db.query(Utente).filter(Utente.username == username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username già esistente")
    
    utente = Utente(
        username=username,
        password_hash=get_password_hash(password or "changeme"),
        nome=nome, cognome=cognome, email=email,
        gruppo_id=gruppo_id, ruolo_id=ruolo_id,
        is_admin=is_admin, is_active=is_active
    )
    db.add(utente)
    db.commit()
    db.refresh(utente)
    return {"id": utente.id, "status": "ok"}

@app.put("/utenti/{utente_id}")
def update_utente(
    utente_id: int,
    username: str = Query(None),
    password: str = Query(None),
    nome: str = Query(None),
    cognome: str = Query(None),
    email: str = Query(None),
    gruppo_id: int = Query(None),
    ruolo_id: int = Query(None),
    is_admin: bool = Query(None),
    is_active: bool = Query(None),
    db: Session = Depends(get_db)
):
    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if not utente:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    if username is not None:
        utente.username = username
    if password is not None:
        utente.password_hash = get_password_hash(password)
    if nome is not None:
        utente.nome = nome
    if cognome is not None:
        utente.cognome = cognome
    if email is not None:
        utente.email = email
    if gruppo_id is not None:
        utente.gruppo_id = gruppo_id
    if ruolo_id is not None:
        utente.ruolo_id = ruolo_id
    if is_admin is not None:
        utente.is_admin = is_admin
    if is_active is not None:
        utente.is_active = is_active
    
    db.commit()
    db.refresh(utente)
    return {"id": utente.id, "status": "ok"}

@app.delete("/utenti/{utente_id}")
def delete_utente(utente_id: int, db: Session = Depends(get_db)):
    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if not utente:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    db.delete(utente)
    db.commit()
    return {"status": "ok"}

# ==========================================
# PARAMETRI SISTEMA
# ==========================================
@app.get("/parametri-sistema")
def get_parametri_sistema(db: Session = Depends(get_db)):
    parametri = db.query(ParametriSistema).order_by(ParametriSistema.gruppo, ParametriSistema.chiave).all()
    return [{
        "id": p.id, "chiave": p.chiave, "valore": p.valore,
        "descrizione": p.descrizione, "tipo_dato": p.tipo_dato,
        "gruppo": p.gruppo
    } for p in parametri]

@app.put("/parametri-sistema/{chiave}")
def update_parametro_sistema(chiave: str, data: dict, db: Session = Depends(get_db)):
    parametro = db.query(ParametriSistema).filter(ParametriSistema.chiave == chiave).first()
    if not parametro:
        raise HTTPException(status_code=404, detail="Parametro non trovato")
    if "valore" in data:
        parametro.valore = data["valore"]
    parametro.updated_at = datetime.now()
    db.commit()
    return {"status": "ok"}

# ==========================================
# OPZIONI DROPDOWN
# ==========================================
@app.get("/opzioni-dropdown/gruppi")
def get_gruppi_dropdown(db: Session = Depends(get_db)):
    try:
        result = db.execute(text("""
            SELECT gruppo,
                   COUNT(*) as totale,
                   SUM(CASE WHEN attivo = 1 THEN 1 ELSE 0 END) as attive
            FROM opzioni_dropdown
            GROUP BY gruppo
            ORDER BY gruppo
        """))
        return [{"gruppo": r[0], "totale": r[1], "attive": r[2]} for r in result.fetchall()]
    except:
        return []
    
@app.get("/opzioni-dropdown/gruppi-sezioni-map")
def get_gruppi_sezioni_map(db: Session = Depends(get_db)):
    """
    Restituisce { gruppo_dropdown: { sezione_codice, sezione_etichetta } }
    costruita dai campi_configuratore + sezioni_configuratore.
    """
    try:
        result = db.execute(text("""
            SELECT DISTINCT c.gruppo_dropdown, c.sezione, 
                   COALESCE(s.etichetta, c.sezione) as sezione_etichetta
            FROM campi_configuratore c
            LEFT JOIN sezioni_configuratore s ON s.codice = c.sezione
            WHERE c.gruppo_dropdown IS NOT NULL 
              AND c.gruppo_dropdown != ''
        """))
        mapping = {}
        for row in result.fetchall():
            mapping[row[0]] = {
                "sezione_codice": row[1],
                "sezione_etichetta": row[2]
            }
        return mapping
    except Exception as e:
        print(f"Errore gruppi-sezioni-map: {e}")
        return {}

@app.get("/opzioni-dropdown/{gruppo}")
def get_opzioni_dropdown(gruppo: str, solo_attive: bool = True, db: Session = Depends(get_db)):
    try:
        q = "SELECT id, gruppo, valore, etichetta, ordine, attivo FROM opzioni_dropdown WHERE gruppo = :g"
        if solo_attive:
            q += " AND attivo = 1"
        q += " ORDER BY ordine"
        result = db.execute(text(q), {"g": gruppo})
        return [{
            "id": r[0], "gruppo": r[1], "valore": r[2],
            "etichetta": r[3],   # <-- era "label"
            "label": r[3],       # backward compat
            "ordine": r[4], "attivo": bool(r[5])
        } for r in result.fetchall()]
    except:
        return []


# SOSTITUIRE @app.post("/opzioni-dropdown") (riga ~2371-2379)
@app.post("/opzioni-dropdown")
def create_opzione_dropdown(data: dict, db: Session = Depends(get_db)):
    try:
        # Accetta sia "etichetta" che "label"
        etichetta = data.get("etichetta") or data.get("label") or data.get("valore", "")
        db.execute(text(
            "INSERT INTO opzioni_dropdown (gruppo, valore, etichetta, ordine, attivo) "
            "VALUES (:gruppo, :valore, :etichetta, :ordine, :attivo)"
        ), {
            "gruppo": data["gruppo"],
            "valore": data["valore"],
            "etichetta": etichetta,
            "ordine": data.get("ordine", 0),
            "attivo": 1 if data.get("attivo", True) else 0,
        })
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# SOSTITUIRE @app.put("/opzioni-dropdown/{opzione_id}") (riga ~2381-2389)
@app.put("/opzioni-dropdown/{opzione_id}")
def update_opzione_dropdown(opzione_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna SOLO i campi presenti nel payload."""
    try:
        updates = []
        params = {"id": opzione_id}

        if "valore" in data and data["valore"] is not None:
            updates.append("valore = :valore")
            params["valore"] = data["valore"]

        # Accetta sia "etichetta" che "label"
        etichetta = data.get("etichetta", data.get("label"))
        if etichetta is not None:
            updates.append("etichetta = :etichetta")
            params["etichetta"] = etichetta

        if "ordine" in data:
            updates.append("ordine = :ordine")
            params["ordine"] = data["ordine"]

        if "attivo" in data:
            updates.append("attivo = :attivo")
            params["attivo"] = 1 if data["attivo"] else 0

        if not updates:
            return {"status": "noop"}

        sql = f"UPDATE opzioni_dropdown SET {', '.join(updates)} WHERE id = :id"
        db.execute(text(sql), params)
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/opzioni-dropdown/{opzione_id}")
def delete_opzione_dropdown(opzione_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM opzioni_dropdown WHERE id=:id"), {"id": opzione_id})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/opzioni-dropdown/{gruppo}/riordina")
def riordina_opzioni(gruppo: str, data: dict, db: Session = Depends(get_db)):
    try:
        for i, oid in enumerate(data.get("ids", [])):
            db.execute(text("UPDATE opzioni_dropdown SET ordine=:o WHERE id=:id"), {"o": i, "id": oid})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# DOCUMENT TEMPLATES — CRUD + Export
# ==========================================

@app.get("/document-templates/available-fields")
def get_available_fields(db: Session = Depends(get_db)):
    """Restituisce tutti i campi disponibili (statici + dinamici dal DB)"""
    return get_available_fields_from_db(db)


@app.get("/document-templates/default-config/{tipo}")
def get_default_config(tipo: str, db: Session = Depends(get_db)):
    """Restituisce una configurazione template di default per tipo (preventivo/ordine)"""
    if tipo not in ("preventivo", "ordine"):
        raise HTTPException(status_code=400, detail="Tipo deve essere 'preventivo' o 'ordine'")
    return get_default_template_config_from_db(db, tipo)


@app.get("/document-templates")
def list_document_templates(tipo: str = None, db: Session = Depends(get_db)):
    """Lista tutti i document templates, opzionalmente filtrati per tipo"""
    query = db.query(DocumentTemplate)
    if tipo:
        query = query.filter(DocumentTemplate.tipo == tipo)
    templates = query.order_by(DocumentTemplate.tipo, DocumentTemplate.nome).all()
    return [
        {
            "id": t.id,
            "nome": t.nome,
            "tipo": t.tipo,
            "descrizione": t.descrizione,
            "attivo": t.attivo,
            "is_default": t.is_default,
            "has_logo": t.logo_data is not None,
            "logo_filename": t.logo_filename,
            "config": t.config,
            "created_at": str(t.created_at) if t.created_at else None,
            "updated_at": str(t.updated_at) if t.updated_at else None,
        }
        for t in templates
    ]


@app.get("/document-templates/{template_id}")
def get_document_template(template_id: int, db: Session = Depends(get_db)):
    """Recupera un singolo document template con config completa"""
    t = db.query(DocumentTemplate).filter(DocumentTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template non trovato")
    result = {
        "id": t.id,
        "nome": t.nome,
        "tipo": t.tipo,
        "descrizione": t.descrizione,
        "attivo": t.attivo,
        "is_default": t.is_default,
        "has_logo": t.logo_data is not None,
        "logo_filename": t.logo_filename,
        "logo_mime": t.logo_mime,
        "config": t.config,
        "created_at": str(t.created_at) if t.created_at else None,
        "updated_at": str(t.updated_at) if t.updated_at else None,
    }
    if t.logo_data:
        result["logo_base64"] = base64.b64encode(t.logo_data).decode("utf-8")
    return result


@app.post("/document-templates")
def create_document_template(
    nome: str = Form(...),
    tipo: str = Form(...),
    descrizione: str = Form(""),
    config_json: str = Form(...),
    is_default: bool = Form(False),
    logo: UploadFile = FastAPIFile(None),
    db: Session = Depends(get_db)
):
    """Crea un nuovo document template"""
    if tipo not in ("preventivo", "ordine"):
        raise HTTPException(status_code=400, detail="Tipo deve essere 'preventivo' o 'ordine'")

    try:
        config = json.loads(config_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="config_json non e' un JSON valido")

    if is_default:
        db.query(DocumentTemplate).filter(
            DocumentTemplate.tipo == tipo,
            DocumentTemplate.is_default == True
        ).update({"is_default": False})

    template = DocumentTemplate(
        nome=nome, tipo=tipo, descrizione=descrizione,
        config=config, is_default=is_default, attivo=True,
    )

    if logo:
        logo_bytes = logo.file.read()
        template.logo_data = logo_bytes
        template.logo_filename = logo.filename
        template.logo_mime = logo.content_type

    db.add(template)
    db.commit()
    db.refresh(template)
    return {"id": template.id, "nome": template.nome, "message": "Template creato"}


@app.put("/document-templates/{template_id}")
def update_document_template(
    template_id: int,
    nome: str = Form(None),
    tipo: str = Form(None),
    descrizione: str = Form(None),
    config_json: str = Form(None),
    attivo: bool = Form(None),
    is_default: bool = Form(None),
    remove_logo: bool = Form(False),
    logo: UploadFile = FastAPIFile(None),
    db: Session = Depends(get_db)
):
    """Aggiorna un document template esistente"""
    template = db.query(DocumentTemplate).filter(DocumentTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")

    if nome is not None: template.nome = nome
    if tipo is not None: template.tipo = tipo
    if descrizione is not None: template.descrizione = descrizione
    if attivo is not None: template.attivo = attivo
    if config_json is not None:
        try:
            template.config = json.loads(config_json)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="config_json non e' un JSON valido")

    if is_default is not None:
        if is_default:
            db.query(DocumentTemplate).filter(
                DocumentTemplate.tipo == (tipo or template.tipo),
                DocumentTemplate.is_default == True,
                DocumentTemplate.id != template_id
            ).update({"is_default": False})
        template.is_default = is_default

    if remove_logo:
        template.logo_data = None
        template.logo_filename = None
        template.logo_mime = None
    elif logo:
        logo_bytes = logo.file.read()
        template.logo_data = logo_bytes
        template.logo_filename = logo.filename
        template.logo_mime = logo.content_type

    db.commit()
    return {"id": template.id, "message": "Template aggiornato"}


@app.delete("/document-templates/{template_id}")
def delete_document_template(template_id: int, db: Session = Depends(get_db)):
    """Elimina un document template"""
    template = db.query(DocumentTemplate).filter(DocumentTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template non trovato")
    db.delete(template)
    db.commit()
    return {"message": "Template eliminato"}

# ============================================================================
# SEZIONE A: CRUD sezioni_preventivo (NUOVI endpoint)
# ============================================================================

@app.get("/sezioni-preventivo")
def get_sezioni_preventivo(db: Session = Depends(get_db)):
    """Lista sezioni del documento preventivo, ordinate."""
    try:
        result = db.execute(text("""
            SELECT id, codice, titolo, ordine, tipo, visibile, mostra_titolo, nota_default, stile
            FROM sezioni_preventivo
            ORDER BY ordine
        """))
        return [
            {"id": r[0], "codice": r[1], "titolo": r[2], "ordine": r[3],
             "tipo": r[4], "visibile": bool(r[5]), "mostra_titolo": bool(r[6]),
             "nota_default": r[7], "stile": r[8]}
            for r in result.fetchall()
        ]
    except Exception as e:
        return []


@app.post("/sezioni-preventivo")
def create_sezione_preventivo(data: dict, db: Session = Depends(get_db)):
    """Crea nuova sezione documento."""
    try:
        db.execute(text("""
            INSERT INTO sezioni_preventivo (codice, titolo, ordine, tipo, visibile, mostra_titolo, nota_default, stile)
            VALUES (:codice, :titolo, :ordine, :tipo, :vis, :mt, :nota, :stile)
        """), {
            "codice": data["codice"], "titolo": data["titolo"],
            "ordine": data.get("ordine", 0), "tipo": data.get("tipo", "tabella"),
            "vis": 1 if data.get("visibile", True) else 0,
            "mt": 1 if data.get("mostra_titolo", True) else 0,
            "nota": data.get("nota_default"), "stile": data.get("stile"),
        })
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@app.put("/sezioni-preventivo/riordina")
def riordina_sezioni_preventivo(data: dict, db: Session = Depends(get_db)):
    """Riordina sezioni. Body: {"ordine": [{"id": 1, "ordine": 10}, ...]}"""
    try:
        for item in data.get("ordine", []):
            db.execute(text("UPDATE sezioni_preventivo SET ordine=:o WHERE id=:id"),
                       {"o": item["ordine"], "id": item["id"]})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

@app.put("/sezioni-preventivo/{sezione_id}")
def update_sezione_preventivo(sezione_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna sezione documento."""
    try:
        field_map = {
            "codice": "codice", "titolo": "titolo", "ordine": "ordine",
            "tipo": "tipo", "visibile": "visibile", "mostra_titolo": "mostra_titolo",
            "nota_default": "nota_default", "stile": "stile",
        }
        fields, params = [], {"id": sezione_id}
        for key, col in field_map.items():
            if key in data:
                val = data[key]
                if key in ("visibile", "mostra_titolo"):
                    val = 1 if val else 0
                fields.append(f"{col}=:{col}")
                params[col] = val
        if not fields:
            return {"status": "noop"}
        fields.append("updated_at=CURRENT_TIMESTAMP")
        db.execute(text(f"UPDATE sezioni_preventivo SET {', '.join(fields)} WHERE id=:id"), params)
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/sezioni-preventivo/{sezione_id}")
def delete_sezione_preventivo(sezione_id: int, db: Session = Depends(get_db)):
    """Elimina sezione documento."""
    try:
        db.execute(text("DELETE FROM sezioni_preventivo WHERE id=:id"), {"id": sezione_id})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




# ==========================================
# SEZIONI CONFIGURATORE
# ==========================================

def _ensure_sezioni_table(db):
    """Crea tabella sezioni_configuratore se non esiste"""
    try:
        db.execute(text('''
            CREATE TABLE IF NOT EXISTS sezioni_configuratore (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codice TEXT NOT NULL UNIQUE,
                etichetta TEXT NOT NULL,
                descrizione TEXT,
                icona TEXT DEFAULT 'Settings',
                ordine INTEGER DEFAULT 0,
                attivo INTEGER DEFAULT 1,
                product_template_id INTEGER,
                product_template_ids TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP
            )
        '''))
        db.commit()
    except:
        pass
    # Migrazione: aggiungi colonna se manca (tabella gia' esistente)
    try:
        db.execute(text("SELECT product_template_ids FROM sezioni_configuratore LIMIT 1"))
    except:
        try:
            db.execute(text("ALTER TABLE sezioni_configuratore ADD COLUMN product_template_ids TEXT"))
            db.commit()
        except:
            pass


@app.get("/sezioni-configuratore")
def get_sezioni_configuratore_list(db: Session = Depends(get_db)):
    """Lista tutte le sezioni con info prodotti associati e conteggio campi"""
    _ensure_sezioni_table(db)
    try:
        import json as _json
        result = db.execute(text("""
            SELECT s.id, s.codice, s.etichetta, s.descrizione, s.icona,
                   s.ordine, s.attivo, s.product_template_id, s.product_template_ids,
                   (SELECT COUNT(*) FROM campi_configuratore c WHERE c.sezione = s.codice) as num_campi,
                   (SELECT COUNT(*) FROM campi_configuratore c WHERE c.sezione = s.codice AND c.attivo = 1) as num_campi_attivi
            FROM sezioni_configuratore s
            ORDER BY s.ordine, s.etichetta
        """))
        rows = result.fetchall()
        
        # Carica tutti i product_templates per risolvere i nomi
        pt_result = db.execute(text("SELECT id, categoria, sottocategoria, nome_display FROM product_templates"))
        pt_map = {r[0]: {"categoria": r[1], "sottocategoria": r[2], "nome_display": r[3]} for r in pt_result.fetchall()}
        
        sezioni = []
        for r in rows:
            # Risolvi prodotti associati
            pt_ids = []
            if r[8]:  # product_template_ids (JSON array)
                try:
                    pt_ids = _json.loads(r[8])
                except:
                    pt_ids = []
            elif r[7]:  # fallback: vecchio product_template_id singolo
                pt_ids = [r[7]]
            
            prodotti = [pt_map[pid] for pid in pt_ids if pid in pt_map]
            
            sezioni.append({
                "id": r[0], "codice": r[1], "etichetta": r[2],
                "descrizione": r[3], "icona": r[4], "ordine": r[5],
                "attivo": bool(r[6]),
                "product_template_ids": pt_ids,
                "prodotti": prodotti,
                # Backward compat
                "product_template_id": pt_ids[0] if pt_ids else None,
                "prodotto": prodotti[0] if prodotti else None,
                "num_campi": r[9], "num_campi_attivi": r[10]
            })
        return sezioni
    except Exception as e:
        print(f"Errore get_sezioni: {e}")
        return []


@app.get("/sezioni-configuratore/campi-non-assegnati")
def get_campi_non_assegnati(db: Session = Depends(get_db)):
    """Campi la cui sezione non corrisponde a nessuna sezione registrata"""
    _ensure_sezioni_table(db)
    try:
        result = db.execute(text("""
            SELECT c.id, c.codice, c.etichetta, c.tipo, c.sezione, c.attivo
            FROM campi_configuratore c
            WHERE c.sezione NOT IN (SELECT codice FROM sezioni_configuratore)
               OR c.sezione IS NULL OR c.sezione = ''
            ORDER BY c.sezione, c.ordine
        """))
        return [{"id": r[0], "codice": r[1], "etichetta": r[2], "tipo": r[3], "sezione": r[4], "attivo": bool(r[5])} for r in result.fetchall()]
    except:
        return []


@app.get("/sezioni-configuratore/{sezione_id}")
def get_sezione_detail(sezione_id: int, db: Session = Depends(get_db)):
    """Dettaglio sezione con campi associati"""
    _ensure_sezioni_table(db)
    try:
        import json as _json
        result = db.execute(text("""
            SELECT s.id, s.codice, s.etichetta, s.descrizione, s.icona,
                   s.ordine, s.attivo, s.product_template_id, s.product_template_ids
            FROM sezioni_configuratore s
            WHERE s.id = :id
        """), {"id": sezione_id})
        r = result.fetchone()
        if not r:
            raise HTTPException(status_code=404, detail="Sezione non trovata")
        
        # Risolvi prodotti
        pt_ids = []
        if r[8]:
            try: pt_ids = _json.loads(r[8])
            except: pt_ids = []
        elif r[7]:
            pt_ids = [r[7]]
        
        prodotti = []
        if pt_ids:
            pt_result = db.execute(text("SELECT id, categoria, sottocategoria, nome_display FROM product_templates"))
            pt_map = {pr[0]: {"categoria": pr[1], "sottocategoria": pr[2], "nome_display": pr[3]} for pr in pt_result.fetchall()}
            prodotti = [pt_map[pid] for pid in pt_ids if pid in pt_map]
        
        campi_result = db.execute(text("""
            SELECT id, codice, etichetta, tipo, sezione, gruppo_dropdown, ordine, attivo,
                   obbligatorio, unita_misura, valore_default, visibile_form, usabile_regole
            FROM campi_configuratore WHERE sezione = :s ORDER BY ordine
        """), {"s": r[1]})
        campi = [{
            "id": c[0], "codice": c[1], "etichetta": c[2], "tipo": c[3],
            "sezione": c[4], "gruppo_dropdown": c[5], "ordine": c[6],
            "attivo": bool(c[7]), "obbligatorio": bool(c[8]),
            "unita_misura": c[9], "valore_default": c[10],
            "visibile_form": bool(c[11]) if c[11] is not None else True,
            "usabile_regole": bool(c[12]) if c[12] is not None else False
        } for c in campi_result.fetchall()]
        
        return {
            "id": r[0], "codice": r[1], "etichetta": r[2],
            "descrizione": r[3], "icona": r[4], "ordine": r[5],
            "attivo": bool(r[6]),
            "product_template_ids": pt_ids,
            "prodotti": prodotti,
            "product_template_id": pt_ids[0] if pt_ids else None,
            "prodotto": prodotti[0] if prodotti else None,
            "campi": campi
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sezioni-configuratore")
def create_sezione_configuratore(data: dict, db: Session = Depends(get_db)):
    """Crea una nuova sezione"""
    _ensure_sezioni_table(db)
    try:
        import json as _json
        existing = db.execute(text("SELECT id FROM sezioni_configuratore WHERE codice = :c"), {"c": data["codice"]}).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail=f"Codice '{data['codice']}' gia' esistente")
        
        # Gestisci product_template_ids (array) o fallback a product_template_id (singolo)
        pt_ids_json = None
        if "product_template_ids" in data and data["product_template_ids"]:
            pt_ids_json = _json.dumps(data["product_template_ids"])
        elif "product_template_id" in data and data["product_template_id"]:
            pt_ids_json = _json.dumps([data["product_template_id"]])
        
        max_ord = db.execute(text("SELECT COALESCE(MAX(ordine), -1) FROM sezioni_configuratore")).fetchone()[0]
        db.execute(text("""
            INSERT INTO sezioni_configuratore (codice, etichetta, descrizione, icona, ordine, attivo, product_template_ids)
            VALUES (:codice, :etichetta, :descrizione, :icona, :ordine, :attivo, :pt_ids)
        """), {
            "codice": data["codice"],
            "etichetta": data.get("etichetta", data["codice"]),
            "descrizione": data.get("descrizione"),
            "icona": data.get("icona", "Settings"),
            "ordine": data.get("ordine", max_ord + 1),
            "attivo": 1 if data.get("attivo", True) else 0,
            "pt_ids": pt_ids_json
        })
        db.commit()
        new_id = db.execute(text("SELECT id FROM sezioni_configuratore WHERE codice = :c"), {"c": data["codice"]}).fetchone()[0]
        return {"status": "ok", "id": new_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/sezioni-configuratore/{sezione_id}")
def update_sezione_configuratore(sezione_id: int, data: dict, db: Session = Depends(get_db)):
    """Aggiorna una sezione"""
    _ensure_sezioni_table(db)
    try:
        import json as _json
        existing = db.execute(text("SELECT id, codice FROM sezioni_configuratore WHERE id = :id"), {"id": sezione_id}).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Sezione non trovata")
        old_codice = existing[1]
        
        field_map = {"codice": "codice", "etichetta": "etichetta", "descrizione": "descrizione",
                     "icona": "icona", "ordine": "ordine", "attivo": "attivo"}
        fields, params = [], {"id": sezione_id}
        for key, col in field_map.items():
            if key in data:
                val = data[key]
                if key == "attivo":
                    val = 1 if val else 0
                fields.append(f"{col}=:{col}")
                params[col] = val
        
        # Gestisci product_template_ids
        if "product_template_ids" in data:
            pt_ids = data["product_template_ids"]
            pt_ids_json = _json.dumps(pt_ids) if pt_ids else None
            fields.append("product_template_ids=:pt_ids")
            params["pt_ids"] = pt_ids_json
        
        if fields:
            fields.append("updated_at=CURRENT_TIMESTAMP")
            db.execute(text(f"UPDATE sezioni_configuratore SET {','.join(fields)} WHERE id=:id"), params)
            new_codice = data.get("codice")
            if new_codice and new_codice != old_codice:
                # ── CASCADE COMPLETO RINOMINA SEZIONE ──
                
                # 1. campi_configuratore.sezione
                db.execute(text("UPDATE campi_configuratore SET sezione=:new WHERE sezione=:old"),
                    {"new": new_codice, "old": old_codice})
                
                # 2. campi_configuratore.codice (prefisso sezione: "argano.trazione" → "motore.trazione")
                db.execute(text("""
                    UPDATE campi_configuratore 
                    SET codice = :new || SUBSTR(codice, LENGTH(:old) + 1)
                    WHERE codice LIKE :prefix
                """), {"new": new_codice, "old": old_codice, "prefix": old_codice + ".%"})
                
                # 3. valori_configurazione.sezione
                db.execute(text("UPDATE valori_configurazione SET sezione=:new WHERE sezione=:old"),
                    {"new": new_codice, "old": old_codice})
                
                # 4. valori_configurazione.codice_campo (prefisso)
                db.execute(text("""
                    UPDATE valori_configurazione
                    SET codice_campo = :new || SUBSTR(codice_campo, LENGTH(:old) + 1)
                    WHERE codice_campo LIKE :prefix
                """), {"new": new_codice, "old": old_codice, "prefix": old_codice + ".%"})
                
                # 5. template_data JSON keys nei template (rinomina chiave sezione e field_config keys)
                try:
                    templates = db.execute(text("SELECT id, template_data FROM product_templates WHERE template_data IS NOT NULL")).fetchall()
                    for tmpl_id, tmpl_json in templates:
                        if not tmpl_json:
                            continue
                        tmpl_data = _json.loads(tmpl_json)
                        changed = False
                        
                        # Rinomina chiave sezione (es: "argano" → "motore")
                        if old_codice in tmpl_data:
                            tmpl_data[new_codice] = tmpl_data.pop(old_codice)
                            changed = True
                        
                        # Rinomina chiavi in field_config (es: "argano.trazione" → "motore.trazione")
                        if "field_config" in tmpl_data:
                            fc = tmpl_data["field_config"]
                            new_fc = {}
                            for k, v in fc.items():
                                if k.startswith(old_codice + "."):
                                    new_fc[new_codice + k[len(old_codice):]] = v
                                    changed = True
                                else:
                                    new_fc[k] = v
                            if changed:
                                tmpl_data["field_config"] = new_fc
                        
                        if changed:
                            db.execute(text("UPDATE product_templates SET template_data=:data WHERE id=:id"),
                                {"data": _json.dumps(tmpl_data), "id": tmpl_id})
                except Exception as e_tmpl:
                    print(f"Warning: cascade template_data failed: {e_tmpl}")
                
                # 6. Rule JSON files (rinomina riferimenti campo nei file .json)
                try:
                    import glob
                    rules_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rules")
                    if os.path.exists(rules_dir):
                        for filepath in glob.glob(os.path.join(rules_dir, "*.json")):
                            with open(filepath, "r", encoding="utf-8") as rf:
                                rule_text = rf.read()
                            if f'"{old_codice}.' in rule_text:
                                rule_text = rule_text.replace(f'"{old_codice}.', f'"{new_codice}.')
                                with open(filepath, "w", encoding="utf-8") as wf:
                                    wf.write(rule_text)
                except Exception as e_rules:
                    print(f"Warning: cascade rule files failed: {e_rules}")

            db.commit()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/sezioni-configuratore/{sezione_id}")
def delete_sezione_configuratore(sezione_id: int, db: Session = Depends(get_db)):
    """Elimina una sezione"""
    _ensure_sezioni_table(db)
    try:
        existing = db.execute(text("SELECT id FROM sezioni_configuratore WHERE id = :id"), {"id": sezione_id}).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Sezione non trovata")
        db.execute(text("DELETE FROM sezioni_configuratore WHERE id=:id"), {"id": sezione_id})
        db.commit()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sezioni-configuratore/riordina")
def riordina_sezioni(data: dict, db: Session = Depends(get_db)):
    """Riordina le sezioni"""
    _ensure_sezioni_table(db)
    try:
        for i, sid in enumerate(data.get("ids", [])):
            db.execute(text("UPDATE sezioni_configuratore SET ordine=:o WHERE id=:id"), {"o": i, "id": sid})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sezioni-configuratore/{sezione_id}/associa-campi")
def associa_campi_a_sezione(sezione_id: int, data: dict, db: Session = Depends(get_db)):
    """Associa/sposta campi a una sezione"""
    _ensure_sezioni_table(db)
    try:
        sezione = db.execute(text("SELECT codice FROM sezioni_configuratore WHERE id = :id"), {"id": sezione_id}).fetchone()
        if not sezione:
            raise HTTPException(status_code=404, detail="Sezione non trovata")
        codice_sezione = sezione[0]
        for cid in data.get("campo_ids", []):
            db.execute(text("UPDATE campi_configuratore SET sezione=:s WHERE id=:id"), {"s": codice_sezione, "id": cid})
        db.commit()
        return {"status": "ok", "campi_associati": len(data.get("campo_ids", []))}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sezioni-configuratore/importa-esistenti")
def importa_sezioni_esistenti(db: Session = Depends(get_db)):
    """Importa sezioni esistenti come stringhe in campi_configuratore"""
    _ensure_sezioni_table(db)
    try:
        result = db.execute(text("SELECT DISTINCT sezione FROM campi_configuratore WHERE sezione IS NOT NULL AND sezione != ''"))
        sezioni_esistenti = [r[0] for r in result.fetchall()]
        importate = 0
        for i, codice in enumerate(sorted(sezioni_esistenti)):
            exists = db.execute(text("SELECT id FROM sezioni_configuratore WHERE codice = :c"), {"c": codice}).fetchone()
            if not exists:
                etichetta = codice.replace("_", " ").title()
                db.execute(text("INSERT INTO sezioni_configuratore (codice, etichetta, ordine, attivo) VALUES (:codice, :etichetta, :ordine, 1)"),
                          {"codice": codice, "etichetta": etichetta, "ordine": i * 10})
                importate += 1
        db.commit()
        return {"status": "ok", "importate": importate, "totale_trovate": len(sezioni_esistenti)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# CAMPI CONFIGURATORE
# ==========================================
@app.get("/campi-configuratore/sezioni")
def get_sezioni_configuratore(db: Session = Depends(get_db)):
    try:
        result = db.execute(text("""
            SELECT sezione,
                   COUNT(*) as totale,
                   SUM(CASE WHEN attivo = 1 THEN 1 ELSE 0 END) as attivi
            FROM campi_configuratore
            GROUP BY sezione
            ORDER BY sezione
        """))
        return [{"codice": r[0], "etichetta": r[0], "totale": r[1], "attivi": r[2]} for r in result.fetchall()]
    except:
        return []

@app.get("/campi-configuratore/schema.json")
def get_schema_campi(db: Session = Depends(get_db)):
    try:
        result = db.execute(text("SELECT codice, etichetta, tipo, sezione FROM campi_configuratore WHERE attivo=1 ORDER BY sezione, ordine"))
        return {"campi": [{"codice": r[0], "label": r[1], "tipo": r[2], "sezione": r[3]} for r in result.fetchall()]}
    except:
        return {"campi": []}

@app.get("/campi-configuratore/{sezione}")
def get_campi_sezione(sezione: str, solo_attivi: bool = True, include_opzioni: bool = False, template_id: int = None, preventivo_id: int = None, db: Session = Depends(get_db)):
    try:
        # Se preventivo_id fornito, ricava template_id automaticamente
        effective_template_id = template_id
        if not effective_template_id and preventivo_id:
            prev_row = db.execute(text("SELECT template_id FROM preventivi WHERE id = :pid"),
                                  {"pid": preventivo_id}).fetchone()
            if prev_row and prev_row[0]:
                effective_template_id = prev_row[0]

        q = """SELECT id, codice, etichetta, tipo, sezione, gruppo_dropdown, ordine, attivo, obbligatorio,
                      unita_misura, valore_min, valore_max, valore_default, descrizione, visibile_form, usabile_regole,
                      product_template_ids
               FROM campi_configuratore WHERE sezione=:s"""
        if solo_attivi:
            q += " AND attivo=1"
        q += " ORDER BY ordine"
        result = db.execute(text(q), {"s": sezione})
        campi = []
        for r in result.fetchall():
            pt_ids_raw = r[16]
            pt_ids = None
            if pt_ids_raw:
                try:
                    pt_ids = json.loads(pt_ids_raw)
                except (json.JSONDecodeError, TypeError):
                    pt_ids = None

            # Filtro per template_id: se campo ha restrizioni e template non è nella lista, salta
            if effective_template_id and pt_ids and effective_template_id not in pt_ids:
                continue

            campi.append({
                "id": r[0], "codice": r[1], "label": r[2], "tipo": r[3], "sezione": r[4],
                "gruppo_opzioni": r[5], "ordine": r[6], "attivo": bool(r[7]), "obbligatorio": bool(r[8]),
                "unita_misura": r[9], "valore_min": r[10], "valore_max": r[11],
                "valore_default": r[12], "descrizione": r[13],
                "visibile_form": bool(r[14]) if r[14] is not None else True,
                "usabile_regole": bool(r[15]) if r[15] is not None else False,
                "product_template_ids": pt_ids,
            })
        return campi
    except:
        return []

@app.post("/campi-configuratore")
def create_campo(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("""INSERT INTO campi_configuratore
            (codice, etichetta, tipo, sezione, gruppo_dropdown, ordine, attivo, obbligatorio,
             unita_misura, valore_min, valore_max, valore_default, descrizione, visibile_form, usabile_regole)
            VALUES (:codice, :etichetta, :tipo, :sezione, :gruppo, :ordine, 1, :obb,
                    :unita, :vmin, :vmax, :vdef, :desc, :vform, :uregole)"""),
            {"codice": data["codice"], "etichetta": data.get("label", data.get("etichetta", "")),
             "tipo": data.get("tipo", "text"), "sezione": data["sezione"],
             "gruppo": data.get("gruppo_opzioni", data.get("gruppo_dropdown")),
             "ordine": data.get("ordine", 0), "obb": 1 if data.get("obbligatorio") else 0,
             "unita": data.get("unita_misura"), "vmin": data.get("valore_min"),
             "vmax": data.get("valore_max"), "vdef": data.get("valore_default"),
             "desc": data.get("descrizione"), "vform": 1 if data.get("visibile_form", True) else 0,
             "uregole": 1 if data.get("usabile_regole") else 0})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/campi-configuratore/{campo_id}")
def update_campo(campo_id: int, data: dict, db: Session = Depends(get_db)):
    try:
        # Mappa nomi frontend -> nomi DB
        field_map = {
            "codice": "codice", "label": "etichetta", "etichetta": "etichetta",
            "tipo": "tipo", "sezione": "sezione",
            "gruppo_opzioni": "gruppo_dropdown", "gruppo_dropdown": "gruppo_dropdown",
            "ordine": "ordine", "attivo": "attivo", "obbligatorio": "obbligatorio",
            "unita_misura": "unita_misura", "valore_min": "valore_min",
            "valore_max": "valore_max", "valore_default": "valore_default",
            "descrizione": "descrizione", "visibile_form": "visibile_form",
            "usabile_regole": "usabile_regole",
            # Campi preventivo/PDF
            "includi_preventivo": "includi_preventivo",
            "mostra_default_preventivo": "mostra_default_preventivo",
            "mostra_default": "mostra_default_preventivo",
            "sezione_preventivo": "sezione_preventivo",
            "ordine_preventivo": "ordine_preventivo",
            "etichetta_preventivo": "etichetta_preventivo",
        }
        fields, params = [], {"id": campo_id}
        for frontend_key, db_col in field_map.items():
            if frontend_key in data:
                fields.append(f"{db_col}=:{db_col}")
                params[db_col] = data[frontend_key]
        # product_template_ids va serializzato come JSON
        if "product_template_ids" in data:
            pt_ids = data["product_template_ids"]
            fields.append("product_template_ids=:pt_ids")
            params["pt_ids"] = json.dumps(pt_ids) if pt_ids else None
        if fields:
            # Deduplica (label e etichetta mappano entrambi a etichetta)
            unique_fields = list(dict.fromkeys(fields))
            db.execute(text(f"UPDATE campi_configuratore SET {','.join(unique_fields)} WHERE id=:id"), params)
            db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/campi-configuratore/{campo_id}")
def delete_campo(campo_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM campi_configuratore WHERE id=:id"), {"id": campo_id})
        db.commit()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# SEZIONE B: GET/PUT valori_configurazione MODIFICATI
#            (SOSTITUISCONO gli endpoint esistenti ~linea 2619 e ~linea 2637)
# ============================================================================

@app.get("/preventivi/{preventivo_id}/configurazione/{sezione}")
def get_valori_sezione(preventivo_id: int, sezione: str, db: Session = Depends(get_db)):
    """
    Legge i valori di una sezione per un preventivo.
    AUTO-POPULATE: popola i default MANCANTI da campi_configuratore,
    anche se alcuni valori esistono già (es. da template).
    """
    try:
        # 1. Leggi valori esistenti
        result = db.execute(text("""
            SELECT codice_campo, valore, is_default, COALESCE(is_readonly, 0)
            FROM valori_configurazione
            WHERE preventivo_id = :pid AND sezione = :sez
        """), {"pid": preventivo_id, "sez": sezione})
        rows = result.fetchall()
        
        valori = {r[0]: r[1] for r in rows}
        defaults_info = {r[0]: bool(r[2]) for r in rows}
        readonly_info = {r[0]: bool(r[3]) for r in rows}
        
        # 2. Popola default MANCANTI da campi_configuratore (filtrati per prodotto)
        # Recupera template_id del preventivo per filtrare campi
        prev_row = db.execute(text("SELECT template_id FROM preventivi WHERE id = :pid"),
                              {"pid": preventivo_id}).fetchone()
        prev_template_id = prev_row[0] if prev_row and prev_row[0] else None

        campi = db.execute(text("""
            SELECT codice, valore_default, tipo, product_template_ids
            FROM campi_configuratore
            WHERE sezione = :sez AND attivo = 1
        """), {"sez": sezione}).fetchall()
        
        # Filtra campi per prodotto
        campi_filtrati = []
        for c in campi:
            pt_ids_raw = c[3]
            if pt_ids_raw and prev_template_id:
                try:
                    pt_ids = json.loads(pt_ids_raw)
                    if prev_template_id not in pt_ids:
                        continue
                except (json.JSONDecodeError, TypeError):
                    pass
            campi_filtrati.append(c)

        nuovi = 0
        for codice, valore_default, tipo, _ in campi_filtrati:
            if codice not in valori and valore_default is not None and str(valore_default).strip() != "":
                val = str(valore_default)
                db.execute(text("""
                    INSERT INTO valori_configurazione
                        (preventivo_id, sezione, codice_campo, valore, is_default)
                    VALUES (:pid, :sez, :campo, :val, 1)
                """), {"pid": preventivo_id, "sez": sezione, "campo": codice, "val": val})
                valori[codice] = val
                defaults_info[codice] = True
                readonly_info[codice] = False
                nuovi += 1
        
        if nuovi:
            db.commit()

            # Applica field_config dal template se esiste
            prev = db.execute(text(
                "SELECT template_id FROM preventivi WHERE id = :pid"
            ), {"pid": preventivo_id}).fetchone()
            if prev and prev[0]:
                _apply_template_field_config(db, preventivo_id, prev[0])
                db.commit()
                # Rileggi readonly aggiornati
                for codice in list(readonly_info.keys()):
                    row = db.execute(text("""
                        SELECT COALESCE(is_readonly, 0) FROM valori_configurazione
                        WHERE preventivo_id = :pid AND codice_campo = :campo
                    """), {"pid": preventivo_id, "campo": codice}).fetchone()
                    if row:
                        readonly_info[codice] = bool(row[0])

        return {
            "preventivo_id": preventivo_id,
            "sezione": sezione,
            "valori": valori,
            "is_default": defaults_info,
            "is_readonly": readonly_info,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/preventivi/{preventivo_id}/configurazione/{sezione}")
def save_valori_sezione(preventivo_id: int, sezione: str, data: dict, db: Session = Depends(get_db)):
    """
    Salva valori sezione (upsert) con tracking is_default.
    I campi con is_readonly=1 vengono SALTATI (non modificabili).
    """
    try:
        valori = data.get("valori", {})

        # Carica valori default dalla definizione campi (per confronto)
        defaults_result = db.execute(text("""
            SELECT codice, valore_default
            FROM campi_configuratore
            WHERE sezione = :sez AND attivo = 1
        """), {"sez": sezione})
        campo_defaults = {r[0]: r[1] for r in defaults_result.fetchall()}

        # Carica stato attuale (valore, is_default, is_readonly)
        existing_result = db.execute(text("""
            SELECT codice_campo, valore, is_default, COALESCE(is_readonly, 0)
            FROM valori_configurazione
            WHERE preventivo_id = :pid AND sezione = :sez
        """), {"pid": preventivo_id, "sez": sezione})
        existing = {r[0]: {"valore": r[1], "is_default": r[2], "is_readonly": r[3]}
                    for r in existing_result.fetchall()}

        skipped_readonly = []

        for codice_campo, valore in valori.items():
            # Salta campi readonly
            if codice_campo in existing and existing[codice_campo]["is_readonly"]:
                skipped_readonly.append(codice_campo)
                continue

            valore_str = str(valore) if valore is not None else None

            if codice_campo in existing:
                old = existing[codice_campo]
                if old["is_default"] == 0:
                    new_is_default = 0
                elif valore_str != old["valore"]:
                    new_is_default = 0
                else:
                    new_is_default = 1

                db.execute(text("""
                    UPDATE valori_configurazione
                    SET valore = :val, is_default = :isd, updated_at = CURRENT_TIMESTAMP
                    WHERE preventivo_id = :pid AND sezione = :sez AND codice_campo = :campo
                """), {"val": valore_str, "isd": new_is_default,
                       "pid": preventivo_id, "sez": sezione, "campo": codice_campo})
            else:
                default_val = campo_defaults.get(codice_campo)
                is_def = 1 if (default_val is not None and valore_str == str(default_val)) else 0

                db.execute(text("""
                    INSERT INTO valori_configurazione
                        (preventivo_id, sezione, codice_campo, valore, is_default)
                    VALUES (:pid, :sez, :campo, :val, :isd)
                """), {"pid": preventivo_id, "sez": sezione, "campo": codice_campo,
                       "val": valore_str, "isd": is_def})

        db.commit()

        # Trigger rule engine dopo salvataggio
        result = safe_evaluate_rules(preventivo_id, db)

        response = {
            "status": "ok",
            "sezione": sezione,
            "campi_salvati": len(valori) - len(skipped_readonly),
            "rules_result": result,
        }
        if skipped_readonly:
            response["campi_readonly_saltati"] = skipped_readonly

        return response
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/preventivi/{preventivo_id}/configurazione")
def get_tutti_valori_configurazione(preventivo_id: int, db: Session = Depends(get_db)):
    """Leggi tutti i valori di tutte le sezioni dinamiche per un preventivo"""
    try:
        result = db.execute(
            text("""
                SELECT sezione, codice_campo, valore 
                FROM valori_configurazione 
                WHERE preventivo_id = :pid
                ORDER BY sezione, codice_campo
            """),
            {"pid": preventivo_id}
        )
        
        sezioni = {}
        for row in result.fetchall():
            sezione, campo, valore = row[0], row[1], row[2]
            if sezione not in sezioni:
                sezioni[sezione] = {}
            sezioni[sezione][campo] = valore
        
        return {"preventivo_id": preventivo_id, "sezioni": sezioni}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/preventivi/{preventivo_id}/configurazione/{sezione}")
def delete_valori_sezione(preventivo_id: int, sezione: str, db: Session = Depends(get_db)):
    """Elimina tutti i valori di una sezione per un preventivo"""
    try:
        db.execute(
            text("DELETE FROM valori_configurazione WHERE preventivo_id = :pid AND sezione = :sez"),
            {"pid": preventivo_id, "sez": sezione}
        )
        db.commit()
        return {"status": "ok", "sezione": sezione}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# SEZIONE C: DATI PREVENTIVO STRUTTURATI PER PDF
# ============================================================================
# SEZIONE C: DATI DOCUMENTO PREVENTIVO
# ============================================================================

@app.get("/preventivi/{preventivo_id}/dati-documento")
def get_dati_documento_preventivo(preventivo_id: int, db: Session = Depends(get_db)):
    """
    Restituisce tutti i dati del preventivo strutturati per sezione del documento.
    Usato dal generatore PDF/DOCX per assemblare il documento.
    
    Logica flag (per-preventivo > globale):
    - Se valori_configurazione ha includi_preventivo/mostra_default_preventivo -> usa quelli
    - Altrimenti fallback su campi_configuratore (globale)
    """
    try:
        # 1. Carica sezioni preventivo (visibili, ordinate)
        sez_result = db.execute(text("""
            SELECT codice, titolo, ordine, tipo, mostra_titolo, nota_default, stile
            FROM sezioni_preventivo
            WHERE visibile = 1
            ORDER BY ordine
        """))
        sezioni_doc = [
            {"codice": r[0], "titolo": r[1], "ordine": r[2], "tipo": r[3],
             "mostra_titolo": bool(r[4]), "nota": r[5], "stile": r[6]}
            for r in sez_result.fetchall()
        ]

        # 2. Carica tutti i campi con mapping preventivo (globale)
        campi_result = db.execute(text("""
            SELECT codice, etichetta, tipo, sezione, unita_misura, valore_default,
                   COALESCE(includi_preventivo, 0) as includi_glob,
                   sezione_preventivo,
                   COALESCE(ordine_preventivo, 0) as ordine_prev,
                   etichetta_preventivo,
                   COALESCE(mostra_default_preventivo, 0) as mostra_def_glob
            FROM campi_configuratore
            WHERE attivo = 1
            ORDER BY sezione_preventivo, ordine_preventivo
        """))
        campi_global = {}
        for r in campi_result.fetchall():
            campi_global[r[0]] = {
                "codice": r[0], "etichetta": r[1], "tipo_campo": r[2],
                "sezione_config": r[3], "unita_misura": r[4],
                "valore_default": r[5],
                "includi_preventivo_glob": bool(r[6]),
                "sezione_preventivo": r[7],
                "ordine_preventivo": r[8],
                "etichetta_preventivo": r[9],
                "mostra_default_glob": bool(r[10]),
            }

        # 3. Carica tutti i valori per questo preventivo (con flag per-preventivo)
        val_result = db.execute(text("""
            SELECT sezione, codice_campo, valore,
                   COALESCE(is_default, 1) as is_default,
                   includi_preventivo,
                   mostra_default_preventivo
            FROM valori_configurazione
            WHERE preventivo_id = :pid
        """), {"pid": preventivo_id})
        valori_all = {}
        for sez, campo, valore, is_def, incl_prev, mostra_def in val_result.fetchall():
            valori_all[campo] = {
                "valore": valore,
                "is_default": bool(is_def),
                "sezione_config": sez,
                "includi_preventivo_prev": incl_prev,
                "mostra_default_prev": mostra_def,
            }

        # 4. Assembla il documento
        sezioni_output = []
        valori_standard = []

        for sez_doc in sezioni_doc:
            cod_sez = sez_doc["codice"]

            if cod_sez == "valori_standard":
                continue

            if cod_sez == "materiali":
                sezioni_output.append({**sez_doc, "campi": [], "_tipo_speciale": "materiali"})
                continue

            # Trova campi assegnati a questa sezione preventivo
            campi_sez = [
                v for v in campi_global.values()
                if v["sezione_preventivo"] == cod_sez
            ]
            campi_sez.sort(key=lambda x: x["ordine_preventivo"])

            campi_output = []
            for campo_def in campi_sez:
                codice = campo_def["codice"]
                val_info = valori_all.get(codice, {})
                valore = val_info.get("valore")
                is_def = val_info.get("is_default", True)

                if valore is None and campo_def["valore_default"]:
                    valore = campo_def["valore_default"]
                    is_def = True

                if valore is None or str(valore).strip() == "":
                    continue

                # Determina includi_preventivo: per-preventivo > globale
                incl_prev = val_info.get("includi_preventivo_prev")
                if incl_prev is not None:
                    includi = bool(incl_prev)
                else:
                    includi = campo_def["includi_preventivo_glob"]

                if not includi:
                    continue

                # Determina mostra_default: per-preventivo > globale
                mostra_def_prev = val_info.get("mostra_default_prev")
                if mostra_def_prev is not None:
                    mostra_default = bool(mostra_def_prev)
                else:
                    mostra_default = campo_def["mostra_default_glob"]

                campo_out = {
                    "codice": codice,
                    "etichetta": campo_def["etichetta_preventivo"] or campo_def["etichetta"],
                    "valore": valore,
                    "unita_misura": campo_def["unita_misura"],
                    "is_default": is_def,
                    "tipo_campo": campo_def["tipo_campo"],
                }

                if is_def and mostra_default:
                    valori_standard.append({
                        **campo_out,
                        "sezione_config": campo_def["sezione_config"],
                    })

                campi_output.append(campo_out)

            if campi_output:
                sezioni_output.append({**sez_doc, "campi": campi_output})

        # 5. Aggiungi sezione "valori_standard" se ci sono
        sez_standard = next((s for s in sezioni_doc if s["codice"] == "valori_standard"), None)
        if sez_standard and valori_standard:
            sezioni_output.append({**sez_standard, "campi": valori_standard})

        return {
            "preventivo_id": preventivo_id,
            "sezioni": sezioni_output,
            "valori_standard": valori_standard,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# SEZIONE D: INIZIALIZZA DEFAULTS PER UN PREVENTIVO
# ============================================================================

@app.post("/preventivi/{preventivo_id}/inizializza-defaults")
def inizializza_defaults_preventivo(preventivo_id: int, db: Session = Depends(get_db)):
    """
    Popola TUTTE le sezioni con i valori di default da campi_configuratore.
    Poi applica field_config dal template se associato.
    """
    try:
        campi = db.execute(text("""
            SELECT codice, sezione, valore_default, tipo
            FROM campi_configuratore
            WHERE attivo = 1 AND valore_default IS NOT NULL AND valore_default != ''
        """)).fetchall()

        existing = db.execute(text("""
            SELECT codice_campo FROM valori_configurazione WHERE preventivo_id = :pid
        """), {"pid": preventivo_id}).fetchall()
        existing_set = {r[0] for r in existing}

        count = 0
        sezioni_touched = set()

        for codice, sezione, valore_default, tipo in campi:
            if codice in existing_set:
                continue
            db.execute(text("""
                INSERT INTO valori_configurazione
                    (preventivo_id, sezione, codice_campo, valore, is_default)
                VALUES (:pid, :sez, :campo, :val, 1)
            """), {"pid": preventivo_id, "sez": sezione,
                   "campo": codice, "val": str(valore_default)})
            count += 1
            sezioni_touched.add(sezione)

        if count:
            db.commit()

        # Applica field_config dal template
        field_config_count = 0
        prev = db.execute(text(
            "SELECT template_id FROM preventivi WHERE id = :pid"
        ), {"pid": preventivo_id}).fetchone()
        if prev and prev[0]:
            field_config_count = _apply_template_field_config(db, preventivo_id, prev[0])
            if field_config_count:
                db.commit()

        return {
            "status": "ok",
            "campi_inizializzati": count,
            "sezioni": sorted(sezioni_touched),
            "field_config_applicati": field_config_count,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# RULE ENGINE - DEBUG CONTEXT
# ==========================================

@app.get("/preventivi/{preventivo_id}/rule-context")
def get_rule_context(preventivo_id: int, db: Session = Depends(get_db)):
    """Debug: mostra il contesto completo che il rule engine usa per valutare."""
    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    
    context = build_config_context(preventivo_id, db)
    rules = load_rules()
    
    # Mostra anche quali regole matcherebbero
    matching = []
    for rule in rules:
        if not rule.get("enabled", True):
            continue
        conditions = rule.get("conditions", [])
        all_met = all(evaluate_condition(c, context) for c in conditions) if conditions else False
        matching.append({
            "id": rule.get("id"),
            "name": rule.get("name"),
            "match": all_met,
            "conditions_detail": [
                {
                    "field": c.get("field"),
                    "operator": c.get("operator"),
                    "expected": c.get("value"),
                    "actual": context.get(c.get("field")),
                    "result": evaluate_condition(c, context)
                }
                for c in conditions
            ]
        })
    
    return {
        "preventivo_id": preventivo_id,
        "context": context,
        "context_keys_count": len(context),
        "rules_total": len(rules),
        "rules_matching": sum(1 for r in matching if r["match"]),
        "rules_detail": matching
    }


# ============================================================
# ORDINI E BOM - ENDPOINT
# ============================================================

@app.post("/preventivi/{preventivo_id}/conferma")
def api_conferma_preventivo(preventivo_id: int, body: dict = None, db: Session = Depends(get_db)):
    """
    Conferma preventivo e genera/aggiorna ordine.
    
    Body opzionale:
    - action: "new" (crea nuovo ordine, default), "update" (aggiorna ordine esistente), "check" (solo controlla)
    - ordine_id: ID ordine da aggiornare (se action=update)
    """
    action = (body or {}).get("action", "new")
    target_ordine_id = (body or {}).get("ordine_id")

    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()
    if not preventivo:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")

    try:
        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()

        # Ensure ordini table has revision columns
        _ensure_ordini_revisione_columns(db)

        # Check existing orders
        try:
            cursor.execute(
                "SELECT id, numero_ordine, stato, numero_revisione_origine, bom_esplosa, bom_numero_revisione "
                "FROM ordini WHERE preventivo_id = ? ORDER BY created_at DESC",
                (preventivo_id,)
            )
            existing_ordini = []
            for r in cursor.fetchall():
                existing_ordini.append({
                    "id": r[0], "numero_ordine": r[1], "stato": r[2],
                    "numero_revisione_origine": r[3], "bom_esplosa": bool(r[4]),
                    "bom_numero_revisione": r[5]
                })
        except Exception:
            cursor.execute(
                "SELECT id, numero_ordine, stato, bom_esplosa FROM ordini WHERE preventivo_id = ? ORDER BY created_at DESC",
                (preventivo_id,)
            )
            existing_ordini = [{"id": r[0], "numero_ordine": r[1], "stato": r[2], "bom_esplosa": bool(r[3])} for r in cursor.fetchall()]

        # Se action=check, restituisci solo info sugli ordini esistenti
        if action == "check":
            rev_corrente = getattr(preventivo, 'revisione_corrente', 0) or 0
            return {
                "existing_ordini": existing_ordini,
                "revisione_corrente": rev_corrente,
                "preventivo_id": preventivo_id,
            }

        # Auto-snapshot prima della conferma
        try:
            snap_result = _crea_snapshot_preventivo(preventivo_id, db, motivo="Auto-snapshot pre-conferma")
            rev_id = snap_result["id"] if snap_result else None
            num_rev = snap_result["numero_revisione"] if snap_result else 0
        except Exception as snap_err:
            print(f"[WARN] Snapshot pre-conferma fallito: {snap_err}")
            rev_id = None
            num_rev = getattr(preventivo, 'revisione_corrente', 0) or 0

        # Lead time
        try:
            lead_time = calcola_lead_time(preventivo_id, db)
        except Exception:
            lead_time = 15

        # Calcola totali
        materiali = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
        totale = _prev_totale(preventivo) or sum(m.prezzo_totale or 0 for m in materiali)
        config_snap = json.dumps({
            "preventivo": _prev_numero(preventivo),
            "revisione": num_rev,
            "materiali": [{"codice": m.codice, "descrizione": m.descrizione,
                           "quantita": m.quantita, "prezzo_totale": m.prezzo_totale} for m in materiali]
        }, ensure_ascii=False)

        if action == "update" and target_ordine_id:
            # AGGIORNA ordine esistente
            cursor.execute(
                "UPDATE ordini SET stato='confermato', tipo_impianto=?, configurazione_json=?, "
                "totale_materiali=?, totale_netto=?, lead_time_giorni=?, "
                "data_consegna_prevista=?, revisione_id=?, numero_revisione_origine=?, "
                "updated_at=datetime('now') WHERE id=?",
                (_prev_tipo(preventivo), config_snap, totale,
                 _prev_netto(preventivo) or totale, lead_time,
                 (datetime.now() + timedelta(days=lead_time)).isoformat(),
                 rev_id, num_rev, target_ordine_id)
            )
            conn.commit()

            # Aggiorna preventivo
            _prev_set(preventivo, "confermato", 'stato', 'status')
            db.commit()

            cursor.execute("SELECT numero_ordine FROM ordini WHERE id=?", (target_ordine_id,))
            numero_ordine = cursor.fetchone()[0]

            return {
                "status": "aggiornato", "preventivo_id": preventivo_id,
                "ordine_id": target_ordine_id, "numero_ordine": numero_ordine,
                "revisione_origine": num_rev,
                "lead_time_giorni": lead_time,
                "data_consegna_prevista": (datetime.now() + timedelta(days=lead_time)).isoformat(),
                "totale": totale
            }

        # action == "new" — CREA nuovo ordine
        if _prev_stato(preventivo) not in ("draft", "bozza", "inviato", "confermato", None):
            raise HTTPException(status_code=400, detail=f"Stato attuale: {_prev_stato(preventivo)}")

        _prev_set(preventivo, "confermato", 'stato', 'status')
        db.commit()

        anno = datetime.now().year
        
        # Auto-crea/aggiorna tabella ordini
        try:
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='ordini'")
            row = cursor.fetchone()
            if row and 'NOT NULL' in (row[0] or '') and 'cliente_id' in (row[0] or ''):
                cursor.execute("DROP TABLE ordini")
                conn.commit()
        except Exception:
            pass
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ordini (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_ordine TEXT UNIQUE NOT NULL,
                preventivo_id INTEGER,
                cliente_id INTEGER DEFAULT NULL,
                stato TEXT DEFAULT 'confermato',
                tipo_impianto TEXT,
                configurazione_json TEXT,
                totale_materiali REAL DEFAULT 0,
                totale_netto REAL DEFAULT 0,
                lead_time_giorni INTEGER DEFAULT 15,
                data_consegna_prevista TEXT,
                bom_esplosa INTEGER DEFAULT 0,
                data_esplosione_bom TEXT,
                revisione_id INTEGER,
                numero_revisione_origine INTEGER,
                bom_revisione_id INTEGER,
                bom_numero_revisione INTEGER,
                bom_esplosa_at TEXT,
                created_by TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()
        _ensure_ordini_revisione_columns(db)

        cursor.execute("SELECT COUNT(*) FROM ordini WHERE numero_ordine LIKE ?", (f"ORD-{anno}-%",))
        count = cursor.fetchone()[0]
        numero_ordine = f"ORD-{anno}-{count + 1:04d}"

        cursor.execute(
            "INSERT INTO ordini (numero_ordine, preventivo_id, cliente_id, stato, tipo_impianto, "
            "configurazione_json, totale_materiali, totale_netto, lead_time_giorni, "
            "data_consegna_prevista, revisione_id, numero_revisione_origine, created_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (numero_ordine, preventivo_id, getattr(preventivo, 'cliente_id', 0) or 0, "confermato",
             _prev_tipo(preventivo), config_snap, totale,
             _prev_netto(preventivo) or totale, lead_time,
             (datetime.now() + timedelta(days=lead_time)).isoformat(),
             rev_id, num_rev, "admin")
        )
        ordine_id = cursor.lastrowid
        conn.commit()

        # Salva ordine_id nel preventivo
        try:
            cursor.execute(
                "UPDATE preventivi SET ordine_id = ?, data_conferma = datetime('now'), "
                "lead_time_giorni = ? WHERE id = ?",
                (ordine_id, lead_time, preventivo_id)
            )
            conn.commit()
        except Exception as e:
            print(f"[WARN] ordine_id backlink failed: {e}")

        return {
            "status": "confermato", "preventivo_id": preventivo_id,
            "ordine_id": ordine_id, "numero_ordine": numero_ordine,
            "revisione_origine": num_rev,
            "lead_time_giorni": lead_time,
            "data_consegna_prevista": (datetime.now() + timedelta(days=lead_time)).isoformat(),
            "totale": totale
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Errore conferma: {str(e)}")


@app.get("/ordini")
def api_get_ordini(db: Session = Depends(get_db)):
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT o.id, o.numero_ordine, o.preventivo_id, o.stato, o.tipo_impianto, "
        "o.totale_materiali, o.lead_time_giorni, o.data_consegna_prevista, "
        "o.bom_esplosa, o.created_at, c.ragione_sociale "
        "FROM ordini o LEFT JOIN clienti c ON c.id = o.cliente_id "
        "ORDER BY o.created_at DESC"
    )
    return [{"id": r[0], "numero_ordine": r[1], "preventivo_id": r[2], "stato": r[3],
             "tipo_impianto": r[4], "totale_materiali": r[5], "lead_time_giorni": r[6],
             "data_consegna_prevista": r[7], "bom_esplosa": bool(r[8]),
             "created_at": r[9], "cliente": r[10]} for r in cursor.fetchall()]


@app.get("/ordini/{ordine_id}")
def api_get_ordine(ordine_id: int, db: Session = Depends(get_db)):
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ordini WHERE id = ?", (ordine_id,))
    r = cursor.fetchone()
    if not r:
        raise HTTPException(status_code=404, detail="Ordine non trovato")
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, r))


@app.post("/ordini/{ordine_id}/esplodi-bom")

def api_esplodi_bom(ordine_id: int, body: dict = None, db: Session = Depends(get_db)):
    """
    Esplode BOM per un ordine.
    Body opzionale:
    - action: "esplodi" (default), "check" (controlla se gia esplosa)
    """
    action = (body or {}).get("action", "esplodi")
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, preventivo_id FROM ordini WHERE id = ?", (ordine_id,))
    ordine = cursor.fetchone()
    if not ordine:
        raise HTTPException(status_code=404, detail="Ordine non trovato")

    preventivo_id = ordine[1]

    # Check se BOM gia esplosa
    if action == "check":
        try:
            cursor.execute(
                "SELECT bom_esplosa, bom_numero_revisione, bom_esplosa_at FROM ordini WHERE id=?",
                (ordine_id,)
            )
            r = cursor.fetchone()
            # Revisione corrente del preventivo
            cursor.execute("SELECT revisione_corrente FROM preventivi WHERE id=?", (preventivo_id,))
            pr = cursor.fetchone()
            rev_corrente = pr[0] if pr and pr[0] else 0
            return {
                "bom_esplosa": bool(r[0]) if r else False,
                "bom_numero_revisione": r[1] if r else None,
                "bom_esplosa_at": r[2] if r else None,
                "revisione_corrente": rev_corrente,
                "outdated": rev_corrente > (r[1] or 0) if r and r[0] else False,
            }
        except Exception:
            return {"bom_esplosa": False}

    preventivo = db.query(Preventivo).filter(Preventivo.id == preventivo_id).first()

    # Revisione corrente
    rev_corrente = getattr(preventivo, 'revisione_corrente', 0) or 0
    # Ultimo snapshot ID
    cursor.execute(
        "SELECT id FROM revisioni_preventivo WHERE preventivo_id=? ORDER BY id DESC LIMIT 1",
        (preventivo_id,)
    )
    last_snap = cursor.fetchone()
    rev_id = last_snap[0] if last_snap else None

    # Costruisci contesto usando build_config_context (stesso usato da regole)
    contesto = build_config_context(preventivo_id, db)
    if not contesto.get("numero_fermate") and not contesto.get("NUM_FERMATE"):
        contesto.update({"NUM_FERMATE": 6, "CORSA": 18, "TIPO_DISPLAY": "LCD",
                         "TIPO_PORTE": "Automatiche", "TENSIONE": "400V", "POTENZA_KW": 7.5})

    contesto_norm = {}
    for k, v in contesto.items():
        contesto_norm[k] = v
        contesto_norm[k.upper()] = v

    cursor.execute("DELETE FROM esplosi WHERE ordine_id = ?", (ordine_id,))

    materiali = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
    if not materiali:
        raise HTTPException(status_code=400, detail="Nessun materiale nel preventivo")

    tutti = []
    for mat in materiali:
        tutti.extend(esplodi_bom_ricorsiva(cursor, mat.codice, mat.quantita,
                                            contesto_norm, ordine_id, 0, mat.codice, []))

    aggregati = {}
    for esp in tutti:
        cod = esp["codice"]
        if cod in aggregati:
            aggregati[cod]["quantita"] += esp["quantita"]
            aggregati[cod]["costo_totale"] += esp["costo_totale"]
        else:
            aggregati[cod] = esp.copy()

    for esp in aggregati.values():
        cursor.execute(
            "INSERT INTO esplosi (ordine_id, codice, descrizione, tipo, categoria, "
            "quantita, unita_misura, costo_unitario, costo_totale, padre_codice, "
            "livello_esplosione, percorso, lead_time_giorni, "
            "parametro1_nome, parametro1_valore, parametro2_nome, parametro2_valore) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (ordine_id, esp["codice"], esp.get("descrizione"), esp.get("tipo"), esp.get("categoria"),
             esp["quantita"], esp.get("unita_misura", "PZ"), esp.get("costo_unitario", 0),
             esp.get("costo_totale", 0), esp.get("padre_codice"), esp.get("livello_esplosione", 0),
             esp.get("percorso"), esp.get("lead_time_giorni", 0),
             esp.get("parametro1_nome"), esp.get("parametro1_valore"),
             esp.get("parametro2_nome"), esp.get("parametro2_valore"))
        )

    now_iso = datetime.now().isoformat()
    try:
        cursor.execute(
            "UPDATE ordini SET bom_esplosa=1, data_esplosione_bom=?, stato='in_produzione', "
            "bom_revisione_id=?, bom_numero_revisione=?, bom_esplosa_at=? WHERE id=?",
            (now_iso, rev_id, rev_corrente, now_iso, ordine_id)
        )
    except Exception:
        # Fallback se colonne nuove non esistono
        cursor.execute("UPDATE ordini SET bom_esplosa=1, data_esplosione_bom=?, stato='in_produzione' WHERE id=?",
                       (now_iso, ordine_id))
    conn.commit()

    costo_tot = sum(e.get("costo_totale", 0) for e in aggregati.values())
    max_lt = max((e.get("lead_time_giorni", 0) for e in aggregati.values()), default=0)

    return {
        "ordine_id": ordine_id,
        "revisione_bom": rev_corrente,
        "componenti_master": len(materiali),
        "componenti_esplosi": len(tutti),
        "componenti_aggregati": len(aggregati),
        "costo_totale_componenti": round(costo_tot, 2),
        "lead_time_max_giorni": max_lt,
        "dettaglio": [{"codice": e["codice"], "descrizione": e.get("descrizione"),
                       "tipo": e.get("tipo"), "quantita": e["quantita"],
                       "unita": e.get("unita_misura"), "costo_totale": round(e.get("costo_totale", 0), 2)}
                      for e in sorted(aggregati.values(), key=lambda x: x.get("categoria", ""))]
    }


@app.get("/ordini/{ordine_id}/esplosi")
def api_get_esplosi(ordine_id: int, tipo: str = None, db: Session = Depends(get_db)):
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    q = "SELECT * FROM esplosi WHERE ordine_id = ?"
    params = [ordine_id]
    if tipo:
        q += " AND tipo = ?"
        params.append(tipo)
    q += " ORDER BY categoria, codice"
    cursor.execute(q, params)
    cols = [d[0] for d in cursor.description]
    esplosi = [dict(zip(cols, r)) for r in cursor.fetchall()]
    totale = sum(e.get("costo_totale", 0) or 0 for e in esplosi)
    return {"ordine_id": ordine_id, "totale_componenti": len(esplosi),
            "costo_totale": round(totale, 2), "esplosi": esplosi}


@app.get("/ordini/{ordine_id}/lista-acquisti")
def api_lista_acquisti(ordine_id: int, db: Session = Depends(get_db)):
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT e.codice, e.descrizione, e.quantita, e.unita_misura, e.costo_unitario, e.costo_totale, "
        "ab.fornitore, ab.codice_fornitore, ab.lead_time_acquisto "
        "FROM esplosi e LEFT JOIN articoli_bom ab ON ab.codice = e.codice "
        "WHERE e.ordine_id = ? AND e.tipo = 'ACQUISTO' ORDER BY ab.fornitore, e.codice",
        (ordine_id,)
    )
    fornitori = {}
    for r in cursor.fetchall():
        f = r[6] or "Non assegnato"
        if f not in fornitori:
            fornitori[f] = {"articoli": [], "totale": 0}
        fornitori[f]["articoli"].append({"codice": r[0], "descrizione": r[1], "quantita": r[2],
            "unita": r[3], "costo_unitario": r[4], "costo_totale": r[5],
            "codice_fornitore": r[7], "lead_time_giorni": r[8]})
        fornitori[f]["totale"] += r[5] or 0
    return {
        "ordine_id": ordine_id, "num_fornitori": len(fornitori),
        "costo_totale_acquisti": round(sum(f["totale"] for f in fornitori.values()), 2),
        "fornitori": {n: {"num_articoli": len(d["articoli"]), "totale": round(d["totale"], 2),
                          "articoli": d["articoli"]} for n, d in fornitori.items()}
    }


@app.get("/preventivi/{preventivo_id}/lead-time")
def api_lead_time(preventivo_id: int, db: Session = Depends(get_db)):
    lt = calcola_lead_time(preventivo_id, db)
    return {"preventivo_id": preventivo_id, "lead_time_giorni": lt}

# ==========================================
# STARTUP
# ==========================================
@app.on_event("startup")
def startup_event():
    """Inizializzazione all'avvio"""
    db = SessionLocal()
    try:
        create_default_admin(db)
        _ensure_revisioni_table(db)
        _ensure_revisione_corrente_column(db)
        _ensure_ordini_revisione_columns(db)
        # Migrazione: product_template_ids su campi_configuratore
        try:
            db.execute(text("SELECT product_template_ids FROM campi_configuratore LIMIT 1"))
        except Exception:
            try:
                db.execute(text("ALTER TABLE campi_configuratore ADD COLUMN product_template_ids TEXT"))
                db.commit()
                print("[STARTUP] Aggiunta colonna product_template_ids a campi_configuratore")
            except Exception:
                db.rollback()
    except Exception as e:
        print(f"Errore startup: {e}")
    finally:
        db.close()

# ============================================================
# REVISIONI PREVENTIVO
# ============================================================

def _ensure_revisioni_table(db):
    """Crea tabella revisioni_preventivo se non esiste"""
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS revisioni_preventivo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            preventivo_id INTEGER NOT NULL,
            numero_revisione INTEGER NOT NULL,
            motivo TEXT,
            snapshot_configurazione TEXT,
            snapshot_materiali TEXT,
            snapshot_totali TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (preventivo_id) REFERENCES preventivi(id)
        )
    """)
    conn.commit()

def _ensure_revisione_corrente_column(db):
    """Aggiunge colonna revisione_corrente a preventivi se non esiste"""
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT revisione_corrente FROM preventivi LIMIT 1")
    except Exception:
        cursor.execute("ALTER TABLE preventivi ADD COLUMN revisione_corrente INTEGER DEFAULT 0")
        conn.commit()
        print("[MIGRATION] Aggiunta colonna revisione_corrente a preventivi")

def _ensure_ordini_revisione_columns(db):
    """Aggiunge colonne di tracciamento revisione alla tabella ordini"""
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ordini'")
        if not cursor.fetchone():
            return  # Tabella ordini non esiste ancora
    except Exception:
        return

    new_cols = {
        "revisione_id": "INTEGER",
        "numero_revisione_origine": "INTEGER",
        "bom_revisione_id": "INTEGER",
        "bom_numero_revisione": "INTEGER",
        "bom_esplosa_at": "TEXT",
    }
    for col_name, col_type in new_cols.items():
        try:
            cursor.execute(f"SELECT {col_name} FROM ordini LIMIT 1")
        except Exception:
            try:
                cursor.execute(f"ALTER TABLE ordini ADD COLUMN {col_name} {col_type}")
                conn.commit()
                print(f"[MIGRATION] Aggiunta colonna {col_name} a ordini")
            except Exception as e:
                print(f"[WARN] Migration {col_name}: {e}")

def _crea_snapshot_preventivo(preventivo_id: int, db, motivo: str = None, created_by: str = "admin"):
    """Crea uno snapshot/revisione del preventivo corrente"""
    _ensure_revisioni_table(db)

    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()

    # Prossimo numero revisione
    cursor.execute(
        "SELECT COALESCE(MAX(numero_revisione), 0) + 1 FROM revisioni_preventivo WHERE preventivo_id = ?",
        (preventivo_id,)
    )
    num_rev = cursor.fetchone()[0]

    # Snapshot configurazione
    cursor.execute("SELECT * FROM preventivi WHERE id = ?", (preventivo_id,))
    prev_row = cursor.fetchone()
    if not prev_row:
        return None
    prev_cols = [d[0] for d in cursor.description]
    prev_dict = dict(zip(prev_cols, prev_row))

    # Snapshot configurazione da sezioni — PRIMA di usare config_snap
    config_snap = {}
    for table in ['dati_commessa', 'dati_principali', 'normative', 'disposizione_vano', 'porte', 'argano']:
        try:
            cursor.execute(f"SELECT * FROM {table} WHERE preventivo_id = ?", (preventivo_id,))
            row = cursor.fetchone()
            if row:
                cols = [d[0] for d in cursor.description]
                config_snap[table] = dict(zip(cols, row))
        except Exception:
            pass

    # Valori configurazione (tabella chiave/valore) — ORA config_snap esiste
    try:
        cursor.execute(
            "SELECT codice_campo, valore, sezione FROM valori_configurazione WHERE preventivo_id = ?",
            (preventivo_id,)
        )
        valori_config = {}
        for row in cursor.fetchall():
            valori_config[row[0]] = {"valore": row[1], "sezione": row[2]}
        if valori_config:
            config_snap["valori_configurazione"] = valori_config
    except Exception:
        pass

    # Carica configurazione JSON dalle sezioni dinamiche
    try:
        cursor.execute(
            "SELECT sezione, valori FROM configurazione_preventivo WHERE preventivo_id = ?",
            (preventivo_id,)
        )
        for row in cursor.fetchall():
            config_snap[f"config_{row[0]}"] = row[1]
    except Exception:
        pass

    # Snapshot materiali
    materiali = db.query(Materiale).filter(Materiale.preventivo_id == preventivo_id).all()
    mat_list = []
    for m in materiali:
        mat_dict = {}
        for col in m.__table__.columns:
            val = getattr(m, col.name, None)
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            mat_dict[col.name] = val
        mat_list.append(mat_dict)

    # Snapshot totali
    totali = {
        "total_price": prev_dict.get("total_price", 0),
        "sconto_cliente": prev_dict.get("sconto_cliente", 0),
        "sconto_extra_admin": prev_dict.get("sconto_extra_admin", 0),
        "total_price_finale": prev_dict.get("total_price_finale", 0),
    }

    cursor.execute(
        "INSERT INTO revisioni_preventivo "
        "(preventivo_id, numero_revisione, motivo, snapshot_configurazione, snapshot_materiali, snapshot_totali, created_by) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            preventivo_id, num_rev, motivo,
            json.dumps(config_snap, ensure_ascii=False, default=str),
            json.dumps(mat_list, ensure_ascii=False, default=str),
            json.dumps(totali, ensure_ascii=False, default=str),
            created_by
        )
    )
    rev_id = cursor.lastrowid

    # Aggiorna revisione_corrente sul preventivo
    cursor.execute(
        "UPDATE preventivi SET revisione_corrente = ?, updated_at = datetime('now') WHERE id = ?",
        (num_rev, preventivo_id)
    )
    conn.commit()

    return {
        "id": rev_id,
        "preventivo_id": preventivo_id,
        "numero_revisione": num_rev,
        "motivo": motivo,
        "created_at": datetime.now().isoformat()
    }


@app.get("/preventivi/{preventivo_id}/revisioni")
def get_revisioni(preventivo_id: int, db: Session = Depends(get_db)):
    """Lista revisioni di un preventivo"""
    _ensure_revisioni_table(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, preventivo_id, numero_revisione, motivo, created_by, created_at "
        "FROM revisioni_preventivo WHERE preventivo_id = ? ORDER BY numero_revisione DESC",
        (preventivo_id,)
    )
    rows = cursor.fetchall()
    cols = [d[0] for d in cursor.description]
    return [dict(zip(cols, r)) for r in rows]


@app.get("/preventivi/{preventivo_id}/revisioni/{revisione_id}")
def get_revisione_dettaglio(preventivo_id: int, revisione_id: int, db: Session = Depends(get_db)):
    """Dettaglio di una revisione con snapshot completo"""
    _ensure_revisioni_table(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM revisioni_preventivo WHERE id = ? AND preventivo_id = ?",
        (revisione_id, preventivo_id)
    )
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Revisione non trovata")
    cols = [d[0] for d in cursor.description]
    result = dict(zip(cols, row))
    # Parse JSON fields
    for field in ['snapshot_configurazione', 'snapshot_materiali', 'snapshot_totali']:
        if result.get(field):
            try:
                result[field] = json.loads(result[field])
            except Exception:
                pass
    return result


@app.post("/preventivi/{preventivo_id}/revisioni")
def crea_revisione(preventivo_id: int, body: dict = None, db: Session = Depends(get_db)):
    """Crea manualmente una revisione/snapshot del preventivo"""
    motivo = (body or {}).get("motivo", "Snapshot manuale")
    result = _crea_snapshot_preventivo(preventivo_id, db, motivo=motivo)
    if not result:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")
    return result


@app.post("/preventivi/{preventivo_id}/auto-snapshot")
def auto_snapshot(preventivo_id: int, db: Session = Depends(get_db)):
    """Crea snapshot solo se il preventivo è dirty (modificato dopo ultimo snapshot)"""
    _ensure_revisioni_table(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT updated_at FROM preventivi WHERE id = ?", (preventivo_id,))
    row = cursor.fetchone()
    if not row or not row[0]:
        return {"created": False, "reason": "preventivo non trovato o senza updated_at"}

    prev_updated = row[0]

    cursor.execute(
        "SELECT created_at FROM revisioni_preventivo WHERE preventivo_id = ? ORDER BY id DESC LIMIT 1",
        (preventivo_id,)
    )
    row = cursor.fetchone()
    if row and str(prev_updated) <= str(row[0]):
        return {"created": False, "reason": "nessuna modifica dall'ultimo snapshot"}

    result = _crea_snapshot_preventivo(preventivo_id, db, motivo="Auto-save all'uscita")
    return {"created": True, "revisione": result}


@app.post("/preventivi/{preventivo_id}/revisioni/{revisione_id}/ripristina")
def ripristina_revisione(preventivo_id: int, revisione_id: int, db: Session = Depends(get_db)):
    """Ripristina un preventivo da una revisione precedente"""
    _ensure_revisioni_table(db)

    try:
        # Auto-snapshot stato corrente
        try:
            _crea_snapshot_preventivo(preventivo_id, db, motivo="Auto-snapshot prima di ripristino")
        except Exception as e:
            print(f"[WARN] Auto-snapshot fallito: {e}")

        conn = db.get_bind().raw_connection()
        cursor = conn.cursor()

        # Carica la revisione (incluso numero_revisione)
        cursor.execute(
            "SELECT numero_revisione, snapshot_configurazione, snapshot_materiali, snapshot_totali "
            "FROM revisioni_preventivo WHERE id = ? AND preventivo_id = ?",
            (revisione_id, preventivo_id)
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Revisione non trovata")

        num_rev_ripristinata = row[0]
        snap_config = json.loads(row[1]) if row[1] else {}
        snap_materiali = json.loads(row[2]) if row[2] else []
        snap_totali = json.loads(row[3]) if row[3] else {}

        # Ripristina totali + revisione_corrente
        totali_updates = ["revisione_corrente = ?", "updated_at = datetime('now')"]
        totali_params = [num_rev_ripristinata]
        if snap_totali:
            for k, v in snap_totali.items():
                totali_updates.append(f"{k} = ?")
                totali_params.append(v)
        totali_params.append(preventivo_id)
        cursor.execute(
            f"UPDATE preventivi SET {', '.join(totali_updates)} WHERE id = ?",
            totali_params
        )

        # Ripristina materiali
        cursor.execute("DELETE FROM materiali WHERE preventivo_id = ?", (preventivo_id,))
        for mat in snap_materiali:
            mat.pop('id', None)
            mat.pop('created_at', None)
            mat['preventivo_id'] = preventivo_id
            cols = list(mat.keys())
            vals = list(mat.values())
            placeholders = ', '.join(['?'] * len(cols))
            try:
                cursor.execute(
                    f"INSERT INTO materiali ({', '.join(cols)}) VALUES ({placeholders})",
                    vals
                )
            except Exception as e:
                print(f"[WARN] Skip materiale restore: {e}")

        # Ripristina sezioni configurazione
        for table_name, data in snap_config.items():
            if table_name == "valori_configurazione":
                try:
                    cursor.execute(
                        "DELETE FROM valori_configurazione WHERE preventivo_id = ?",
                        (preventivo_id,)
                    )
                    if isinstance(data, dict):
                        for campo, info in data.items():
                            valore = info.get("valore", "") if isinstance(info, dict) else str(info)
                            sezione = info.get("sezione", "") if isinstance(info, dict) else ""
                            cursor.execute(
                                "INSERT INTO valori_configurazione "
                                "(preventivo_id, sezione, codice_campo, valore) "
                                "VALUES (?, ?, ?, ?)",
                                (preventivo_id, sezione, campo, valore)
                            )
                except Exception as e:
                    print(f"[WARN] Skip valori_configurazione restore: {e}")

            elif table_name.startswith('config_'):
                sezione = table_name[7:]
                try:
                    cursor.execute(
                        "UPDATE configurazione_preventivo SET valori = ?, updated_at = datetime('now') "
                        "WHERE preventivo_id = ? AND sezione = ?",
                        (data if isinstance(data, str) else json.dumps(data), preventivo_id, sezione)
                    )
                except Exception:
                    pass
            elif isinstance(data, dict):
                try:
                    data_copy = {k: v for k, v in data.items() if k not in ('id', 'preventivo_id')}
                    if data_copy:
                        updates = [f"{k} = ?" for k in data_copy.keys()]
                        params = list(data_copy.values()) + [preventivo_id]
                        cursor.execute(
                            f"UPDATE {table_name} SET {', '.join(updates)} WHERE preventivo_id = ?",
                            params
                        )
                except Exception as e:
                    print(f"[WARN] Skip config restore {table_name}: {e}")

        conn.commit()

        # Riesegui regole dopo ripristino
        try:
            safe_evaluate_rules(preventivo_id, db)
        except Exception:
            pass

        return {
            "success": True,
            "message": f"Preventivo ripristinato dalla revisione #{num_rev_ripristinata}",
            "preventivo_id": preventivo_id,
            "revisione_corrente": num_rev_ripristinata
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Errore ripristino: {str(e)}")
    
@app.get("/preventivi/{preventivo_id}/dirty")
def check_dirty(preventivo_id: int, db: Session = Depends(get_db)):
    """Controlla se il preventivo ha modifiche non snapshot-ate."""
    _ensure_revisioni_table(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT updated_at FROM preventivi WHERE id = ?", (preventivo_id,))
    row = cursor.fetchone()
    if not row or not row[0]:
        return {"dirty": False, "preventivo_id": preventivo_id}

    prev_updated = row[0]

    cursor.execute(
        "SELECT created_at FROM revisioni_preventivo WHERE preventivo_id = ? ORDER BY id DESC LIMIT 1",
        (preventivo_id,)
    )
    row = cursor.fetchone()
    if not row:
        return {"dirty": True, "preventivo_id": preventivo_id}

    dirty = str(prev_updated) > str(row[0])
    return {"dirty": dirty, "preventivo_id": preventivo_id}


@app.get("/preventivi/{preventivo_id}/filiera")
def get_filiera(preventivo_id: int, db: Session = Depends(get_db)):
    """Restituisce la filiera compatta: preventivo -> ordini -> BOM con info revisioni"""
    _ensure_revisioni_table(db)
    conn = db.get_bind().raw_connection()
    cursor = conn.cursor()

    # Preventivo — prova entrambi i nomi colonna
    cursor.execute("SELECT * FROM preventivi WHERE id = ?", (preventivo_id,))
    prev = cursor.fetchone()
    if not prev:
        raise HTTPException(404, "Preventivo non trovato")
    prev_cols = [d[0] for d in cursor.description]
    prev_full = dict(zip(prev_cols, prev))
    
    prev_data = {
        "id": prev_full.get("id"),
        "revisione_corrente": prev_full.get("revisione_corrente", 0) or 0,
        "updated_at": prev_full.get("updated_at"),
        "status": prev_full.get("status") or prev_full.get("stato") or "draft",
    }

    # Ultima revisione
    cursor.execute(
        "SELECT numero_revisione, created_at FROM revisioni_preventivo WHERE preventivo_id = ? ORDER BY id DESC LIMIT 1",
        (preventivo_id,)
    )
    last_rev = cursor.fetchone()
    prev_data["ultima_revisione"] = last_rev[0] if last_rev else 0
    prev_data["ultima_revisione_data"] = last_rev[1] if last_rev else None

    # Ordini — con fallback robusto
    ordini = []
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ordini'")
        if cursor.fetchone():
            # Prova prima con colonne nuove
            try:
                cursor.execute(
                    "SELECT id, numero_ordine, stato, revisione_id, numero_revisione_origine, "
                    "bom_esplosa, bom_numero_revisione, bom_esplosa_at, created_at, "
                    "totale_materiali, totale_netto, lead_time_giorni, data_consegna_prevista "
                    "FROM ordini WHERE preventivo_id = ? ORDER BY created_at DESC",
                    (preventivo_id,)
                )
                ordini_cols = [d[0] for d in cursor.description]
                ordini = [dict(zip(ordini_cols, r)) for r in cursor.fetchall()]
            except Exception:
                # Fallback senza colonne nuove
                cursor.execute(
                    "SELECT id, numero_ordine, stato, bom_esplosa, created_at, "
                    "totale_materiali, totale_netto, lead_time_giorni, data_consegna_prevista "
                    "FROM ordini WHERE preventivo_id = ? ORDER BY created_at DESC",
                    (preventivo_id,)
                )
                ordini_cols = [d[0] for d in cursor.description]
                ordini = [dict(zip(ordini_cols, r)) for r in cursor.fetchall()]
    except Exception as e:
        print(f"[WARN] filiera ordini: {e}")
        ordini_cols = [d[0] for d in cursor.description]
        ordini = [dict(zip(ordini_cols, r)) for r in cursor.fetchall()]
    except Exception:
        # Fallback se colonne nuove non esistono ancora
        cursor.execute(
            "SELECT id, numero_ordine, stato, bom_esplosa, created_at, "
            "totale_materiali, totale_netto, lead_time_giorni, data_consegna_prevista "
            "FROM ordini WHERE preventivo_id = ? ORDER BY created_at DESC",
            (preventivo_id,)
        )
        ordini_cols = [d[0] for d in cursor.description]
        ordini = [dict(zip(ordini_cols, r)) for r in cursor.fetchall()]

    # Per ogni ordine, calcola se è outdated
    rev_corrente = prev_data.get("revisione_corrente", 0) or 0
    for o in ordini:
        rev_origine = o.get("numero_revisione_origine") or 0
        o["outdated"] = rev_corrente > rev_origine if rev_origine > 0 else False
        o["revisioni_dietro"] = max(0, rev_corrente - rev_origine) if rev_origine > 0 else 0

        bom_rev = o.get("bom_numero_revisione") or 0
        o["bom_outdated"] = rev_corrente > bom_rev if bom_rev > 0 and o.get("bom_esplosa") else False
        o["bom_revisioni_dietro"] = max(0, rev_corrente - bom_rev) if bom_rev > 0 else 0

    return {
        "preventivo": prev_data,
        "ordini": ordini,
        "revisione_corrente": rev_corrente,
    }

# ============================================================================
# ENDPOINT: Parse Excel (usato dal wizard del Rule Designer)
# ============================================================================

@app.post("/api/v2/import/excel/parse")
async def parse_excel(
    file: UploadFile = File(...),
    sheet: str = None,
    header_row: int = 1,
    limit: int = 200,
):
    """
    Parsa un file Excel e restituisce dati strutturati.
    Usato dal wizard Excel del Rule Designer.
    
    - Se sheet non specificato: restituisce lista fogli + dati primo foglio
    - Se sheet specificato: restituisce colonne e righe di quel foglio
    """
    import tempfile, shutil
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installato")

    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        all_sheets = wb.sheetnames

        target_sheet = sheet or all_sheets[0]
        if target_sheet not in all_sheets:
            raise HTTPException(400, f"Foglio '{target_sheet}' non trovato. Disponibili: {all_sheets}")

        ws = wb[target_sheet]
        rows_raw = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= header_row - 1 + limit + 1:  # header + limit righe dati
                break
            rows_raw.append(list(row))

        wb.close()

        if len(rows_raw) < header_row:
            return {"sheets": all_sheets, "columns": [], "rows": []}

        # Header dalla riga indicata
        headers = []
        for j, h in enumerate(rows_raw[header_row - 1]):
            if h is not None:
                headers.append(str(h).strip())
            else:
                headers.append(f"col_{j}")

        # Righe dati
        data_rows = []
        for row in rows_raw[header_row:]:
            d = {}
            for j, h in enumerate(headers):
                val = row[j] if j < len(row) else None
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                d[h] = val
            data_rows.append(d)

        return {
            "sheets": all_sheets,
            "columns": headers,
            "rows": data_rows[:limit],
        }
    finally:
        os.unlink(tmp_path)



# ============================================================================
# ENDPOINT: Data Tables (gestione tabelle dati da Excel)
# ============================================================================

@app.post("/data-tables/save")
def save_data_table(data: dict):
    """
    Salva una data table JSON direttamente (usato dal wizard Excel).
    
    Il wizard nel Rule Designer manda il JSON già processato.
    Lo salviamo in ./data/{nome_tabella}.json.
    """
    import os
    meta = data.get("_meta", {})
    nome = meta.get("nome") or data.get("nome_tabella")
    if not nome:
        raise HTTPException(400, "Campo _meta.nome obbligatorio")

    # Sanitizza nome file
    nome_safe = "".join(c for c in nome if c.isalnum() or c == "_").lower()
    if not nome_safe:
        raise HTTPException(400, "Nome tabella non valido")

    data_dir = "./data"
    os.makedirs(data_dir, exist_ok=True)
    outpath = os.path.join(data_dir, f"{nome_safe}.json")

    with open(outpath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return {"saved": nome_safe, "path": outpath}


@app.get("/data-tables")
def list_data_tables():
    """Lista tutte le data tables disponibili in ./data/"""
    loader = ExcelDataLoader()
    tables = loader.list_tables()
    return {"tables": tables, "data_dir": "./data"}


@app.get("/data-tables/{nome_tabella}")
def get_data_table(nome_tabella: str):
    """Restituisce il contenuto di una data table."""
    loader = ExcelDataLoader()
    table = loader.load_table(nome_tabella)
    if not table:
        raise HTTPException(404, f"Tabella '{nome_tabella}' non trovata")
    return table


@app.delete("/data-tables/{nome_tabella}")
def delete_data_table(nome_tabella: str):
    """Elimina una data table."""
    import os
    filepath = os.path.join("./data", f"{nome_tabella}.json")
    if not os.path.exists(filepath):
        raise HTTPException(404, f"Tabella '{nome_tabella}' non trovata")
    os.remove(filepath)
    return {"deleted": nome_tabella}


@app.post("/data-tables/upload")
async def upload_excel_data(file: UploadFile = File(...), overwrite: bool = True):
    """
    Carica un file Excel con foglio _META → genera data tables JSON.
    
    Il file Excel deve seguire la convenzione:
    - Foglio _META con colonne: foglio, tipo, nome_tabella, ...
    - Fogli dati con riga 1 = header, righe successive = dati
    
    Returns: 
        Lista tabelle generate con eventuali errori/warning
    """
    import tempfile
    import shutil

    # Salva file temporaneo
    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        loader = ExcelDataLoader()
        result = loader.load_excel(tmp_path, overwrite=overwrite)
        
        # Salva anche il file Excel originale in ./data/excel_originals/
        originals_dir = os.path.join("./data", "excel_originals")
        os.makedirs(originals_dir, exist_ok=True)
        dest = os.path.join(originals_dir, file.filename)
        shutil.copy2(tmp_path, dest)
        result["original_saved"] = dest
        
        return result
    finally:
        os.unlink(tmp_path)


@app.post("/data-tables/validate")
async def validate_excel_data(file: UploadFile = File(...)):
    """
    Valida un file Excel senza generare tabelle.
    Utile per preview prima del caricamento.
    """
    import tempfile
    import shutil

    suffix = os.path.splitext(file.filename)[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        loader = ExcelDataLoader()
        return loader.validate_excel(tmp_path)
    finally:
        os.unlink(tmp_path)


@app.post("/data-tables/merge")
def merge_data_tables(data: dict):
    """
    Unisce più tabelle lookup_range partizionate in una sola.
    
    Body JSON:
    {
        "table_names": ["contattori_50hz", "contattori_60hz"],
        "merged_name": "contattori_oleo",
        "partition_field": "frequenza_tensione"
    }
    """
    table_names = data.get("table_names", [])
    merged_name = data.get("merged_name", "")
    partition_field = data.get("partition_field", "")

    if not table_names or not merged_name:
        raise HTTPException(400, "Servono table_names e merged_name")

    loader = ExcelDataLoader()
    result = loader.merge_partitioned_tables(table_names, merged_name, partition_field)
    
    if loader.errors:
        raise HTTPException(400, {"errors": loader.errors})
    
    return {"merged": merged_name, "partitions": list(result.get("partizioni", {}).keys())}

if __name__ == "__main__":
    import uvicorn
    print("Avvio Configuratore Elettroquadri API v0.11.0")
    print("Server in ascolto su http://0.0.0.0:8000")
    print("Documentazione API: http://0.0.0.0:8000/docs")
    uvicorn.run(app, host="0.0.0.0", port=8000)
