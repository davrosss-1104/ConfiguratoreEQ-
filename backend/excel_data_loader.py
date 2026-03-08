"""
excel_data_loader.py — Loader generico Excel → JSON per il Rule Engine

CONVENZIONE EXCEL:
Ogni file Excel DEVE avere un foglio "_MAPPA" che descrive i fogli dati.
Il loader legge _MAPPA, valida la struttura, e produce file JSON in ./data/

TIPI DI TABELLA SUPPORTATI:
- lookup_range:   input numerico → output multipli, partizionabile (es: contattori)
- catalogo:       catalogo prodotti, matching multi-criterio (es: trasformatori)
- costanti:       tabella semplice chiave→valore (es: ponti raddrizzatori)

FORMATO FOGLIO _MAPPA:
Riga 1: headers
  foglio | tipo | nome_tabella | colonna_chiave | tipo_chiave | partizionato_per | valore_partizione | riga_intestazioni | note
Righe successive: una riga per ogni foglio dati

COLONNE ART:
Le colonne il cui nome inizia con "ART:" vengono trattate come codici articolo.
Vengono portate nell'output con prefisso "art_" (es. "ART: Cont. dir." → "art_cont._dir.").
"""

import json
import os
import logging
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelDataLoader:
    """Parser generico: Excel con _MAPPA → JSON data tables."""

    TIPI_VALIDI = {"lookup_range", "catalogo", "costanti"}
    
    # Tipi interni per il rule engine (mapping italiano → engine)
    TIPO_ENGINE_MAP = {
        "lookup_range": "lookup_range",
        "catalogo": "catalog",
        "costanti": "constants",
    }

    MAPPA_COLUMNS = [
        "foglio", "tipo", "nome_tabella", "colonna_chiave",
        "tipo_chiave", "partizionato_per", "valore_partizione",
        "riga_intestazioni", "colonna_suffisso", "note"
    ]

    @staticmethod
    def _normalize_key(s: str) -> str:
        """
        Normalizza un nome colonna per produrre chiavi JSON valide
        e compatibili col sistema placeholder del rule engine.
        
        Regole:
        - lowercase
        - spazi → underscore
        - caratteri non alfanumerici (: . ( ) ecc.) → underscore
        - underscore multipli → singolo
        - rimuovi underscore iniziali/finali
        
        Esempi:
          "kW"             → "kw"
          "IN (A)"         → "in_a"
          "DIR: Cont."     → "dir_cont"
          "ART: Cont. Dir." → "art_cont_dir"
          "ST: Mors RST"   → "st_mors_rst"
        """
        import re
        if not s:
            return ""
        n = str(s).strip().lower()
        n = re.sub(r'[^a-zA-Z0-9_]', '_', n)
        n = re.sub(r'_+', '_', n)
        n = n.strip('_')
        return n

    def __init__(self, data_dir: str = "./data"):
        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)
        self.errors: List[str] = []
        self.warnings: List[str] = []

    # ========================================================================
    # ENTRY POINT
    # ========================================================================
    def load_excel(self, filepath: str, overwrite: bool = True) -> Dict[str, Any]:
        """
        Carica un Excel con _MAPPA, valida, genera JSON in ./data/.
        
        Gestisce automaticamente le partizioni: se più righe in _MAPPA hanno 
        lo stesso nome_tabella ma valore_partizione diverso, produce un'unica
        tabella partizionata.
        
        Returns:
            {
                "success": bool,
                "tables_generated": ["nome_tabella1", ...],
                "files_written": ["data/xxx.json", ...],
                "errors": [...],
                "warnings": [...]
            }
        """
        self.errors = []
        self.warnings = []

        try:
            import openpyxl
        except ImportError:
            return {"success": False, "errors": ["openpyxl non installato"]}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            return {"success": False, "errors": [f"Errore apertura file: {e}"]}

        # 1. Leggi e valida _MAPPA
        mappa_entries = self._read_mappa(wb)
        if not mappa_entries:
            wb.close()
            return {
                "success": False,
                "tables_generated": [],
                "files_written": [],
                "errors": self.errors or ["Foglio _MAPPA mancante o vuoto"],
                "warnings": self.warnings
            }

        # 2. Raggruppa entries per nome_tabella (per gestire partizioni automatiche)
        tabelle_group: Dict[str, List[Dict]] = {}
        for entry in mappa_entries:
            nome = entry["nome_tabella"]
            tabelle_group.setdefault(nome, []).append(entry)

        # 3. Per ogni tabella (eventualmente multi-foglio), genera JSON
        tables_generated = []
        files_written = []

        for nome_tabella, entries in tabelle_group.items():
            try:
                json_data = self._build_table(wb, entries, filepath)
                if json_data is None:
                    continue
            except Exception as e:
                self.errors.append(f"Errore generazione '{nome_tabella}': {e}")
                continue

            # Scrivi file
            outpath = os.path.join(self.data_dir, f"{nome_tabella}.json")
            if os.path.exists(outpath) and not overwrite:
                self.warnings.append(f"File '{outpath}' già esiste, saltato (overwrite=False)")
                continue

            with open(outpath, "w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False)

            tables_generated.append(nome_tabella)
            files_written.append(outpath)
            logger.info(f"Generato: {outpath} ({entries[0]['tipo']}, {len(entries)} fogli)")

        wb.close()

        return {
            "success": len(self.errors) == 0,
            "tables_generated": tables_generated,
            "files_written": files_written,
            "errors": self.errors,
            "warnings": self.warnings,
        }

    # ========================================================================
    # DISPATCHER: costruisce tabella da uno o più fogli
    # ========================================================================
    def _build_table(self, wb, entries: List[Dict], filepath: str) -> Optional[Dict]:
        """
        Costruisce la tabella JSON da uno o più entries (fogli).
        Gestisce automaticamente il merge delle partizioni.
        """
        nome_tabella = entries[0]["nome_tabella"]
        tipo = entries[0]["tipo"]
        
        # Verifica che tutti gli entries abbiano lo stesso tipo
        for e in entries:
            if e["tipo"] != tipo:
                self.errors.append(
                    f"Tabella '{nome_tabella}': fogli con tipi diversi "
                    f"({e['tipo']} vs {tipo}). Tutti devono avere lo stesso tipo."
                )
                return None

        # Leggi i dati da ogni foglio
        fogli_data = []
        total_rows = 0
        for entry in entries:
            foglio_nome = entry["foglio"]
            if foglio_nome not in wb.sheetnames:
                self.errors.append(f"Foglio '{foglio_nome}' dichiarato in _MAPPA ma non trovato nel file")
                continue

            ws = wb[foglio_nome]
            riga_intestazioni = int(entry.get("riga_intestazioni") or 1)
            rows = self._read_sheet_data(ws, riga_intestazioni)

            if not rows:
                self.warnings.append(f"Foglio '{foglio_nome}' è vuoto, saltato")
                continue

            fogli_data.append({"entry": entry, "rows": rows})
            total_rows += len(rows) - 1  # escludi header

        if not fogli_data:
            self.warnings.append(f"Nessun dato trovato per tabella '{nome_tabella}'")
            return None

        # Genera JSON in base al tipo
        if tipo == "lookup_range":
            json_data = self._build_lookup_range(fogli_data)
        elif tipo == "catalogo":
            json_data = self._build_catalog(fogli_data)
        elif tipo == "costanti":
            json_data = self._build_constants(fogli_data)
        else:
            self.errors.append(f"Tipo '{tipo}' non supportato per '{nome_tabella}'")
            return None

        # Aggiungi metadata
        json_data["_meta"] = {
            "nome": nome_tabella,
            "tipo": tipo,
            "tipo_engine": self.TIPO_ENGINE_MAP.get(tipo, tipo),
            "fogli_origine": [e["foglio"] for e in entries],
            "file_origine": os.path.basename(filepath),
            "generato_il": datetime.now().isoformat(),
            "righe_totali": total_rows,
            # Info dalla _MAPPA per generazione regole
            "colonna_chiave": entries[0].get("colonna_chiave", ""),
            "tipo_chiave": entries[0].get("tipo_chiave", ""),
            "partizionato_per": entries[0].get("partizionato_per", ""),
            "valori_partizione": [e.get("valore_partizione", "") for e in entries if e.get("valore_partizione")],
        }

        return json_data

    # ========================================================================
    # LETTURA _MAPPA
    # ========================================================================
    def _read_mappa(self, wb) -> List[Dict[str, str]]:
        """Legge il foglio _MAPPA e restituisce lista di entry."""
        if "_MAPPA" not in wb.sheetnames:
            self.errors.append("Foglio '_MAPPA' non trovato. Ogni Excel deve avere un foglio _MAPPA.")
            return []

        ws = wb["_MAPPA"]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            self.errors.append("Foglio _MAPPA vuoto (serve almeno header + 1 riga dati)")
            return []

        # Valida headers — mapping flessibile
        headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

        col_map = {}
        for i, h in enumerate(headers):
            for expected in self.MAPPA_COLUMNS:
                if h == expected or h.replace(" ", "_") == expected:
                    col_map[expected] = i
                    break

        missing = {"foglio", "tipo", "nome_tabella"} - set(col_map.keys())
        if missing:
            self.errors.append(f"Colonne obbligatorie mancanti in _MAPPA: {missing}")
            return []

        # Leggi entries
        entries = []
        for row_idx, row in enumerate(rows[1:], start=2):
            entry = {}
            for col_name, col_idx in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                entry[col_name] = str(val).strip() if val is not None else ""

            # Validazioni
            if not entry.get("foglio"):
                self.warnings.append(f"_MAPPA riga {row_idx}: campo 'foglio' vuoto, saltata")
                continue
            if entry.get("tipo") not in self.TIPI_VALIDI:
                self.errors.append(
                    f"_MAPPA riga {row_idx}: tipo '{entry.get('tipo')}' non valido. "
                    f"Valori ammessi: {', '.join(sorted(self.TIPI_VALIDI))}"
                )
                continue
            if not entry.get("nome_tabella"):
                entry["nome_tabella"] = entry["foglio"].lower().replace(" ", "_")

            entries.append(entry)

        return entries

    # ========================================================================
    # LETTURA DATI FOGLIO
    # ========================================================================
    def _read_sheet_data(self, ws, riga_intestazioni: int = 1) -> List[List[Any]]:
        """
        Legge le righe di un foglio a partire dalla riga intestazioni.
        riga_intestazioni=1 → header è la prima riga (default).
        riga_intestazioni=3 → salta le prime 2 righe decorative.
        """
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < riga_intestazioni:
                continue  # Salta righe prima delle intestazioni
            # Salta righe completamente vuote
            if all(v is None for v in row):
                continue
            rows.append(list(row))
        return rows

    def _rows_to_dicts(self, rows: List[List]) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Converte righe (con header in riga 0) in lista di dict."""
        if len(rows) < 2:
            return [], []
        headers = [self._normalize_key(h) if h else f"col_{i}"
                   for i, h in enumerate(rows[0])]
        dicts = []
        for row in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                if val is not None:
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                d[h] = val
            dicts.append(d)
        return headers, dicts

    # ========================================================================
    # BUILDER: lookup_range
    # ========================================================================
    def _build_lookup_range(self, fogli_data: List[Dict]) -> Dict[str, Any]:
        """
        Genera lookup_range da uno o più fogli.
        
        Calcolo fasce: il punto di separazione tra una fascia e la successiva 
        è il punto medio tra i due valori adiacenti della colonna chiave.
        Prima fascia parte da 0, ultima non ha limite superiore.
        
        Se ci sono più fogli con lo stesso nome_tabella e valore_partizione diverso,
        genera automaticamente una tabella partizionata.
        """
        first_entry = fogli_data[0]["entry"]
        col_chiave_raw = first_entry.get("colonna_chiave", "").strip()
        part_field = first_entry.get("partizionato_per", "").strip()
        is_partitioned = len(fogli_data) > 1 or bool(part_field)

        result = {
            "tipo": "lookup_range",
        }

        if is_partitioned:
            result["partizionato_per"] = part_field
            result["partizioni"] = {}

        all_ranges = []

        for foglio_info in fogli_data:
            entry = foglio_info["entry"]
            rows = foglio_info["rows"]
            headers, data = self._rows_to_dicts(rows)

            # Trova colonna chiave
            col_chiave = self._normalize_key(col_chiave_raw)
            if col_chiave not in headers:
                # Fallback: cerca match parziale
                for h in headers:
                    if col_chiave in h or h in col_chiave:
                        col_chiave = h
                        break
                else:
                    col_chiave = headers[0] if headers else ""
                    self.warnings.append(
                        f"Colonna chiave '{col_chiave_raw}' non trovata in foglio "
                        f"'{entry['foglio']}', uso '{col_chiave}'"
                    )

            result["parametro_lookup"] = col_chiave

            # Tutte le colonne non-chiave sono output
            output_cols = [h for h in headers if h != col_chiave]

            # Ordina per valore chiave
            data_sorted = sorted(data, key=lambda r: float(r.get(col_chiave, 0) or 0))

            # Costruisci ranges con midpoint
            ranges = self._build_ranges(data_sorted, col_chiave, output_cols)

            if is_partitioned:
                # Usa valore_partizione se disponibile, altrimenti nome foglio
                part_value = entry.get("valore_partizione", "").strip()
                if not part_value:
                    part_value = entry["foglio"]
                result["partizioni"][part_value] = ranges
            else:
                all_ranges = ranges

        if not is_partitioned:
            result["ranges"] = all_ranges

        return result

    def _build_ranges(self, data_sorted: List[Dict], col_chiave: str,
                      output_cols: List[str]) -> List[Dict]:
        """
        Costruisce le fasce (ranges) con punto medio come separatore.
        
        Per valori [4.4, 5.9, 7.7, 9.6]:
          fascia 0: [0, 5.15)        — midpoint(4.4, 5.9) = 5.15
          fascia 1: [5.15, 6.8)      — midpoint(5.9, 7.7) = 6.8
          fascia 2: [6.8, 8.65)      — midpoint(7.7, 9.6) = 8.65
          fascia 3: [8.65, None)     — ultimo, senza limite
        """
        ranges = []
        values = []

        # Raccogli valori numerici validi
        for row in data_sorted:
            val = row.get(col_chiave)
            if val is not None:
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    continue

        if not values:
            return ranges

        # Calcola midpoints
        midpoints = []
        for i in range(len(values) - 1):
            midpoints.append(round((values[i] + values[i + 1]) / 2, 4))

        # Costruisci ranges
        for i, row in enumerate(data_sorted):
            val = row.get(col_chiave)
            if val is None:
                continue
            try:
                float(val)
            except (ValueError, TypeError):
                continue

            # Limiti fascia
            da = 0 if i == 0 else midpoints[i - 1]
            a = midpoints[i] if i < len(midpoints) else None

            # Tutti i valori non-chiave vanno in output
            output = {}
            for col in output_cols:
                v = row.get(col)
                if v is not None:
                    output[col] = v

            ranges.append({"da": da, "a": a, "output": output})

        return ranges

    # ========================================================================
    # BUILDER: catalogo
    # ========================================================================
    def _build_catalog(self, fogli_data: List[Dict]) -> Dict[str, Any]:
        """
        Genera catalogo: tabella piatta di prodotti per matching multi-criterio.
        Se colonna_suffisso è definita nella _MAPPA, i componenti con righe multiple
        ricevono il suffisso _{valore}v nella colonna chiave.
        """
        from collections import Counter
        all_records = []
        all_headers = []

        for foglio_info in fogli_data:
            rows = foglio_info["rows"]
            headers, data = self._rows_to_dicts(rows)
            if not all_headers:
                all_headers = headers
            all_records.extend(data)

        entry = fogli_data[0]["entry"]
        col_chiave_raw = entry.get("colonna_chiave", "").strip()
        col_chiave = self._normalize_key(col_chiave_raw)
        if col_chiave not in all_headers:
            col_chiave = all_headers[0] if all_headers else "codice"

        col_suffisso_raw = entry.get("colonna_suffisso") or ""
        col_suffisso = self._normalize_key(col_suffisso_raw) if col_suffisso_raw else ""

        if col_suffisso:
            conteggio = Counter(str(r.get(col_chiave, "")).strip() for r in all_records)
            for record in all_records:
                chiave_val = str(record.get(col_chiave, "")).strip()
                if conteggio.get(chiave_val, 1) > 1:
                    suf_raw = record.get(col_suffisso)
                    if suf_raw is not None:
                        try:
                            suf = str(int(float(str(suf_raw))))
                        except (ValueError, TypeError):
                            suf = str(suf_raw).replace(" ", "_").lower()
                        record[col_chiave] = f"{chiave_val}_{suf}v"

        return {
            "tipo": "catalog",
            "colonna_id": col_chiave,
            "colonne": all_headers,
            "records": all_records,
            "_meta": {
                "colonna_suffisso": col_suffisso or None,
            }
        }

    # ========================================================================
    # BUILDER: costanti
    # ========================================================================
    def _build_constants(self, fogli_data: List[Dict]) -> Dict[str, Any]:
        """
        Genera costanti: tabella semplice chiave → valori.
        
        Output JSON:
        {
            "tipo": "constants",
            "parametro_lookup": "tensione_freno",
            "valori": {
                "180": {"ponte": "tipo_x", ...},
                ...
            }
        }
        """
        all_valori = {}
        col_chiave_final = ""

        for foglio_info in fogli_data:
            entry = foglio_info["entry"]
            rows = foglio_info["rows"]
            headers, data = self._rows_to_dicts(rows)

            col_chiave_raw = entry.get("colonna_chiave", "").strip()
            col_chiave = self._normalize_key(col_chiave_raw)
            if col_chiave not in headers:
                col_chiave = headers[0] if headers else ""

            if not col_chiave_final:
                col_chiave_final = col_chiave

            output_cols = [h for h in headers if h != col_chiave]

            for row in data:
                key = row.get(col_chiave)
                if key is None:
                    continue
                key_str = self._normalize_key(str(key))
                output = {col: row.get(col) for col in output_cols if row.get(col) is not None}
                if key_str in all_valori:
                    self.warnings.append(
                        f"Chiave duplicata '{key_str}' nel foglio '{entry['foglio']}'. "
                        f"Il tipo 'costanti' richiede chiavi univoche. "
                        f"Se servono più righe per la stessa chiave, usare tipo 'catalogo'."
                    )
                all_valori[key_str] = output

        return {
            "tipo": "constants",
            "parametro_lookup": col_chiave_final,
            "valori": all_valori,
        }

    # ========================================================================
    # UTILITY: merge partizioni (legacy, per API /data-tables/merge)
    # ========================================================================
    def merge_partitioned_tables(self, table_names: List[str],
                                 merged_name: str,
                                 partition_field: str) -> Dict[str, Any]:
        """
        Unisce più tabelle lookup_range partizionate in una sola.
        
        NOTA: Con il nuovo loader, le partizioni vengono gestite automaticamente
        in fase di import quando più righe in _MAPPA hanno lo stesso nome_tabella.
        Questo metodo resta per compatibilità.
        """
        merged = {
            "tipo": "lookup_range",
            "partizionato_per": partition_field,
            "partizioni": {},
        }

        parametro_lookup = None

        for name in table_names:
            filepath = os.path.join(self.data_dir, f"{name}.json")
            if not os.path.exists(filepath):
                self.errors.append(f"File '{filepath}' non trovato per merge")
                continue

            with open(filepath, "r", encoding="utf-8") as f:
                table = json.load(f)

            if parametro_lookup is None:
                parametro_lookup = table.get("parametro_lookup")

            merged["parametro_lookup"] = parametro_lookup

            if "partizioni" in table:
                merged["partizioni"].update(table["partizioni"])
            elif "ranges" in table:
                merged["partizioni"][name] = table["ranges"]

        merged["_meta"] = {
            "nome": merged_name,
            "tipo": "lookup_range",
            "merged_from": table_names,
            "generato_il": datetime.now().isoformat(),
        }

        outpath = os.path.join(self.data_dir, f"{merged_name}.json")
        with open(outpath, "w", encoding="utf-8") as f:
            json.dump(merged, f, indent=2, ensure_ascii=False)

        return merged

    # ========================================================================
    # UTILITY: load / list tables
    # ========================================================================
    def load_table(self, nome_tabella: str) -> Optional[Dict[str, Any]]:
        """Carica una tabella JSON dal data_dir."""
        filepath = os.path.join(self.data_dir, f"{nome_tabella}.json")
        if not os.path.exists(filepath):
            return None
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)

    def list_tables(self) -> List[Dict[str, Any]]:
        """Lista tutte le tabelle disponibili in data_dir."""
        tables = []
        if not os.path.exists(self.data_dir):
            return tables
        for fname in sorted(os.listdir(self.data_dir)):
            if not fname.endswith(".json"):
                continue
            filepath = os.path.join(self.data_dir, fname)
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                meta = data.get("_meta", {})
                tables.append({
                    "nome": meta.get("nome", fname.replace(".json", "")),
                    "tipo": data.get("tipo", meta.get("tipo", "unknown")),
                    "file": fname,
                    "righe": meta.get("righe_totali", meta.get("righe", 0)),
                    "file_origine": meta.get("file_origine", ""),
                    "generato_il": meta.get("generato_il", ""),
                })
            except Exception:
                pass
        return tables

    # ========================================================================
    # VALIDAZIONE EXCEL
    # ========================================================================
    def validate_excel(self, filepath: str) -> Dict[str, Any]:
        """
        Valida un Excel senza generare JSON.
        Utile per preview / anteprima prima del caricamento.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            return {"valid": False, "errors": [str(e)]}

        self.errors = []
        self.warnings = []

        result = {
            "fogli": wb.sheetnames,
            "ha_mappa": "_MAPPA" in wb.sheetnames,
        }

        if not result["ha_mappa"]:
            result["valid"] = False
            result["errors"] = ["Foglio '_MAPPA' non trovato"]
            wb.close()
            return result

        mappa_entries = self._read_mappa(wb)

        # Raggruppa per nome_tabella
        tabelle_group: Dict[str, List[Dict]] = {}
        for entry in mappa_entries:
            nome = entry["nome_tabella"]
            tabelle_group.setdefault(nome, []).append(entry)

        tabelle_info = []
        for nome_tabella, entries in tabelle_group.items():
            tipo = entries[0]["tipo"]
            part_field = entries[0].get("partizionato_per", "")
            
            fogli_info = []
            col_output_all = []
            
            for entry in entries:
                foglio = entry["foglio"]
                info = {
                    "foglio": foglio,
                    "valore_partizione": entry.get("valore_partizione", ""),
                    "esiste": foglio in wb.sheetnames,
                    "righe": 0,
                    "colonne": [],
                }
                if info["esiste"]:
                    ws = wb[foglio]
                    riga_int = int(entry.get("riga_intestazioni") or 1)
                    rows = self._read_sheet_data(ws, riga_int)
                    info["righe"] = len(rows) - 1 if rows else 0
                    if rows:
                        headers_raw = [str(h) for h in rows[0] if h is not None]
                        info["colonne"] = headers_raw
                        headers_norm = [self._normalize_key(h)
                                       for h in rows[0] if h is not None]
                        col_chiave = self._normalize_key(
                            entry.get("colonna_chiave", ""))
                        for h in headers_norm:
                            if h != col_chiave and h not in col_output_all:
                                col_output_all.append(h)
                fogli_info.append(info)

            tabella_info = {
                "nome_tabella": nome_tabella,
                "tipo": tipo,
                "colonna_chiave": entries[0].get("colonna_chiave", ""),
                "tipo_chiave": entries[0].get("tipo_chiave", ""),
                "fogli": fogli_info,
                "colonne_output": col_output_all,
            }
            if part_field:
                tabella_info["partizionato_per"] = part_field
                tabella_info["valori_partizione"] = [
                    e.get("valore_partizione", e["foglio"]) for e in entries
                ]
            
            tabelle_info.append(tabella_info)

        result["tabelle"] = tabelle_info
        result["valid"] = len(self.errors) == 0
        result["errors"] = self.errors
        result["warnings"] = self.warnings

        wb.close()
        return result


# ============================================================================
# CLI per test rapido
# ============================================================================
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python excel_data_loader.py <file.xlsx> [--validate]")
        sys.exit(1)

    filepath = sys.argv[1]
    loader = ExcelDataLoader()

    if "--validate" in sys.argv:
        result = loader.validate_excel(filepath)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        result = loader.load_excel(filepath)
        print(json.dumps(result, indent=2, ensure_ascii=False))
