"""
excel_import.py — Import Excel con foglio _MAPPA → data tables + regole JSON

Flusso:
  1. Legge il foglio _MAPPA per capire la struttura dell'Excel
  2. Per ogni riga della mappa, legge il foglio dati corrispondente
  3. Genera data tables JSON in ./data/
  4. Genera regole lookup JSON in ./rules/

Tipi di tabella supportati:
  - lookup_range:  chiave numerica → range automatici (es. contattori per potenza)
  - catalogo:      elenco prodotti per matching multi-criterio (es. trasformatori)
  - costanti:      tabella chiave → valore diretto (es. ponti raddrizzatori)

Convenzione colonne ART:
  Colonne il cui nome inizia con "ART:" contengono codici articolo del gestionale.
  Vengono separate dalle colonne tecniche e portate nelle regole materiali.
"""

import json
import os
import logging
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)

DATA_DIR = "./data"
RULES_DIR = "./rules"

# Colonne del foglio _MAPPA
MAPPA_COLUMNS = [
    "foglio", "tipo", "nome_tabella", "colonna_chiave",
    "tipo_chiave", "partizionato_per", "valore_partizione",
    "riga_intestazioni", "colonna_suffisso", "note"
]

TIPI_VALIDI = {"lookup_range", "catalogo", "costanti"}


# ============================================================================
# CLASSE PRINCIPALE
# ============================================================================

class ExcelImporter:
    """Importa Excel con _MAPPA → data tables JSON + regole lookup."""

    def __init__(self, data_dir: str = DATA_DIR, rules_dir: str = RULES_DIR):
        self.data_dir = data_dir
        self.rules_dir = rules_dir
        self.errors: List[str] = []
        self.warnings: List[str] = []

    # ========================================================================
    # ENTRY POINT: PREVIEW (validazione senza generare)
    # ========================================================================
    def preview(self, filepath: str) -> Dict[str, Any]:
        """
        Legge l'Excel, valida _MAPPA, ritorna preview dei dati senza generare file.
        
        Returns:
            {
                "success": bool,
                "mappa": [ { info per ogni tabella trovata } ],
                "errors": [...],
                "warnings": [...]
            }
        """
        self.errors = []
        self.warnings = []

        wb = self._open_workbook(filepath)
        if wb is None:
            return self._result(False)

        mappa_entries = self._read_mappa(wb)
        if not mappa_entries:
            wb.close()
            return self._result(False)

        preview_tables = []
        for entry in mappa_entries:
            table_preview = self._preview_table(wb, entry)
            if table_preview:
                preview_tables.append(table_preview)

        wb.close()
        result = self._result(len(self.errors) == 0)
        result["mappa"] = preview_tables
        return result

    # ========================================================================
    # ENTRY POINT: GENERA (produce file JSON)
    # ========================================================================
    def genera(self, filepath: str) -> Dict[str, Any]:
        """
        Legge l'Excel, genera data tables e regole lookup.
        
        Returns:
            {
                "success": bool,
                "tables_generated": ["nome_tabella1", ...],
                "rules_generated": ["rule_LOOKUP_xxx.json", ...],
                "files_written": ["data/xxx.json", "rules/yyy.json", ...],
                "errors": [...],
                "warnings": [...]
            }
        """
        self.errors = []
        self.warnings = []

        wb = self._open_workbook(filepath)
        if wb is None:
            return self._result(False)

        mappa_entries = self._read_mappa(wb)
        if not mappa_entries:
            wb.close()
            return self._result(False)

        os.makedirs(self.data_dir, exist_ok=True)
        os.makedirs(self.rules_dir, exist_ok=True)

        # Raggruppa entries per nome_tabella (per le partizioni)
        grouped = self._group_by_table(mappa_entries)

        tables_generated = []
        rules_generated = []
        files_written = []

        for nome_tabella, entries in grouped.items():
            # 1. Leggi dati da tutti i fogli di questa tabella
            table_data = self._build_table(wb, entries)
            if table_data is None:
                continue

            # 2. Salva data table JSON
            data_file = os.path.join(self.data_dir, f"{nome_tabella}.json")
            self._save_json(table_data, data_file)
            tables_generated.append(nome_tabella)
            files_written.append(data_file)

            # 3. Genera regola lookup
            rule = self._build_lookup_rule(nome_tabella, entries, table_data)
            if rule:
                rule_file = os.path.join(self.rules_dir, f"rule_LOOKUP_{nome_tabella.upper()}.json")
                self._save_json(rule, rule_file)
                rules_generated.append(f"rule_LOOKUP_{nome_tabella.upper()}.json")
                files_written.append(rule_file)

        wb.close()

        result = self._result(len(self.errors) == 0)
        result["tables_generated"] = tables_generated
        result["rules_generated"] = rules_generated
        result["files_written"] = files_written
        return result

    # ========================================================================
    # LETTURA _MAPPA
    # ========================================================================
    def _read_mappa(self, wb) -> List[Dict[str, Any]]:
        """Legge il foglio _MAPPA e ritorna lista di entry validate."""
        if "_MAPPA" not in wb.sheetnames:
            self.errors.append(
                "Foglio '_MAPPA' non trovato. "
                "Aggiungete un foglio chiamato esattamente '_MAPPA' "
                "con le colonne: " + ", ".join(MAPPA_COLUMNS[:5])
            )
            return []

        ws = wb["_MAPPA"]

        # Leggi intestazioni dalla riga 1
        headers_raw = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            headers_raw.append(str(v).strip().lower() if v else "")

        # Mappa colonne per nome
        col_map = {}
        for i, h in enumerate(headers_raw):
            # Normalizza: rimuovi spazi, lowercase
            h_norm = h.replace(" ", "_")
            if h_norm in MAPPA_COLUMNS:
                col_map[h_norm] = i

        # Verifica colonne obbligatorie
        required = ["foglio", "tipo", "nome_tabella", "colonna_chiave", "tipo_chiave"]
        missing = [c for c in required if c not in col_map]
        if missing:
            self.errors.append(
                f"Foglio _MAPPA: colonne obbligatorie mancanti: {', '.join(missing)}. "
                f"Intestazioni trovate: {headers_raw}"
            )
            return []

        # Leggi righe dati
        entries = []
        for row in range(2, ws.max_row + 1):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                row_vals.append(v)

            # Salta righe vuote
            foglio_val = row_vals[col_map["foglio"]] if col_map["foglio"] < len(row_vals) else None
            if not foglio_val or (isinstance(foglio_val, str) and not foglio_val.strip()):
                continue

            entry = {}
            for col_name, idx in col_map.items():
                v = row_vals[idx] if idx < len(row_vals) else None
                if v is not None:
                    entry[col_name] = str(v).strip() if isinstance(v, str) else v
                else:
                    entry[col_name] = None

            # Valida entry
            err = self._validate_entry(entry, row)
            if err:
                self.errors.append(err)
                continue

            # Normalizza
            entry["tipo"] = entry["tipo"].lower().strip()
            entry["nome_tabella"] = self._normalize_name(entry["nome_tabella"])
            entry["riga_intestazioni"] = int(entry.get("riga_intestazioni") or 1)

            entries.append(entry)

        if not entries:
            self.errors.append("Foglio _MAPPA: nessuna riga dati valida trovata")

        return entries

    def _validate_entry(self, entry: Dict, row: int) -> Optional[str]:
        """Valida una singola riga della mappa."""
        foglio = entry.get("foglio")
        tipo = entry.get("tipo", "")
        nome = entry.get("nome_tabella")
        chiave = entry.get("colonna_chiave")
        tipo_chiave = entry.get("tipo_chiave", "")

        if not foglio:
            return f"_MAPPA riga {row}: colonna 'foglio' vuota"
        if not tipo:
            return f"_MAPPA riga {row}: colonna 'tipo' vuota"
        if str(tipo).lower().strip() not in TIPI_VALIDI:
            return (
                f"_MAPPA riga {row}: tipo '{tipo}' non riconosciuto. "
                f"Valori ammessi: {', '.join(sorted(TIPI_VALIDI))}"
            )
        if not nome:
            return f"_MAPPA riga {row}: colonna 'nome_tabella' vuota"
        if not chiave:
            return f"_MAPPA riga {row}: colonna 'colonna_chiave' vuota"
        if str(tipo_chiave).lower().strip() not in ("numero", "testo"):
            return f"_MAPPA riga {row}: tipo_chiave deve essere 'numero' o 'testo', trovato: '{tipo_chiave}'"

        return None

    # ========================================================================
    # LETTURA FOGLIO DATI
    # ========================================================================
    def _read_data_sheet(self, wb, entry: Dict) -> Optional[Tuple[List[str], List[Dict]]]:
        """
        Legge un foglio dati secondo le indicazioni della mappa.
        Ritorna (headers, rows) dove rows è lista di dict.
        Separa colonne ART: dalle colonne tecniche.
        """
        foglio_nome = entry["foglio"]
        if foglio_nome not in wb.sheetnames:
            self.errors.append(f"Foglio '{foglio_nome}' non esiste nel file Excel")
            return None

        ws = wb[foglio_nome]
        header_row = entry.get("riga_intestazioni", 1)

        # Leggi intestazioni
        headers_raw = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=col).value
            if v is not None:
                headers_raw.append((col, str(v).strip()))
            else:
                headers_raw.append((col, None))

        # Filtra colonne con intestazione
        headers = [(col, name) for col, name in headers_raw if name]

        if not headers:
            self.errors.append(f"Foglio '{foglio_nome}': nessuna intestazione trovata alla riga {header_row}")
            return None

        # Verifica colonna chiave
        colonna_chiave = entry["colonna_chiave"]
        header_names = [name for _, name in headers]
        if colonna_chiave not in header_names:
            # Prova match case-insensitive
            found = False
            for _, name in headers:
                if name.lower().strip() == colonna_chiave.lower().strip():
                    entry["colonna_chiave"] = name  # Usa il nome esatto
                    found = True
                    break
            if not found:
                self.errors.append(
                    f"Foglio '{foglio_nome}': colonna chiave '{colonna_chiave}' non trovata. "
                    f"Colonne disponibili: {header_names}"
                )
                return None

        # Leggi righe dati
        rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            has_data = False

            for col, name in headers:
                v = ws.cell(row=row_idx, column=col).value
                if v is not None and not (isinstance(v, str) and v.strip() == ""):
                    row_data[name] = v
                    has_data = True
                else:
                    row_data[name] = None

            # Salta righe completamente vuote
            if has_data:
                # Salta righe dove la colonna chiave è vuota
                if row_data.get(entry["colonna_chiave"]) is not None:
                    rows.append(row_data)

        if not rows:
            self.warnings.append(f"Foglio '{foglio_nome}': nessun dato trovato")

        return header_names, rows

    # ========================================================================
    # SEPARAZIONE COLONNE TECNICHE / ARTICOLI
    # ========================================================================
    def _separate_art_columns(self, headers: List[str]) -> Tuple[List[str], List[str]]:
        """Separa intestazioni in colonne tecniche e colonne ART:"""
        tech_cols = []
        art_cols = []
        for h in headers:
            if h.upper().startswith("ART:"):
                art_cols.append(h)
            else:
                tech_cols.append(h)
        return tech_cols, art_cols

    # ========================================================================
    # PREVIEW DI UNA TABELLA
    # ========================================================================
    def _preview_table(self, wb, entry: Dict) -> Optional[Dict]:
        """Genera preview per una singola tabella."""
        result = self._read_data_sheet(wb, entry)
        if result is None:
            return None

        headers, rows = result
        tech_cols, art_cols = self._separate_art_columns(headers)

        preview = {
            "foglio": entry["foglio"],
            "tipo": entry["tipo"],
            "nome_tabella": entry.get("nome_tabella", ""),
            "colonna_chiave": entry["colonna_chiave"],
            "tipo_chiave": entry.get("tipo_chiave", "testo"),
            "partizionato_per": entry.get("partizionato_per"),
            "valore_partizione": entry.get("valore_partizione"),
            "righe": len(rows),
            "colonne_tecniche": tech_cols,
            "colonne_articoli": art_cols,
            "anteprima": []
        }

        # Anteprima prime 5 righe
        for row in rows[:5]:
            preview_row = {}
            for h in tech_cols:
                v = row.get(h)
                if v is not None:
                    preview_row[h] = v
            if art_cols:
                preview_row["_articoli"] = {
                    h: row.get(h) for h in art_cols if row.get(h) is not None
                }
            preview["anteprima"].append(preview_row)

        # Per lookup_range, mostra anche le fasce calcolate
        if entry["tipo"] == "lookup_range":
            chiave = entry["colonna_chiave"]
            kw_values = []
            for row in rows:
                v = row.get(chiave)
                if v is not None:
                    try:
                        kw_values.append(float(v))
                    except (ValueError, TypeError):
                        pass

            if kw_values:
                ranges = self._calcola_ranges(kw_values)
                preview["fasce_calcolate"] = [
                    {"valore": kw, "da": r[0], "a": r[1]}
                    for kw, r in zip(kw_values, ranges)
                ]

        return preview

    # ========================================================================
    # COSTRUZIONE DATA TABLE
    # ========================================================================
    def _group_by_table(self, entries: List[Dict]) -> Dict[str, List[Dict]]:
        """Raggruppa entry per nome_tabella."""
        grouped = {}
        for entry in entries:
            nome = entry["nome_tabella"]
            if nome not in grouped:
                grouped[nome] = []
            grouped[nome].append(entry)
        return grouped

    def _build_table(self, wb, entries: List[Dict]) -> Optional[Dict]:
        """Costruisce una data table JSON da una o più entry della mappa."""
        tipo = entries[0]["tipo"]

        if tipo == "lookup_range":
            return self._build_lookup_range(wb, entries)
        elif tipo == "catalogo":
            return self._build_catalogo(wb, entries)
        elif tipo == "costanti":
            return self._build_costanti(wb, entries)
        else:
            self.errors.append(f"Tipo tabella '{tipo}' non implementato")
            return None

    # --------------------------------------------------------------------
    # BUILDER: lookup_range
    # --------------------------------------------------------------------
    def _build_lookup_range(self, wb, entries: List[Dict]) -> Optional[Dict]:
        """Costruisce tabella lookup_range, con partizioni se più fogli."""
        nome_tabella = entries[0]["nome_tabella"]
        colonna_chiave = entries[0]["colonna_chiave"]
        partizionato_per = entries[0].get("partizionato_per")
        is_partitioned = len(entries) > 1 or partizionato_per

        table = {
            "tipo": "lookup_range",
            "parametro_lookup": self._normalize_name(colonna_chiave),
            "_meta": {
                "nome": nome_tabella,
                "tipo": "lookup_range",
                "generato_il": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "file_origine": "import Excel con _MAPPA",
            }
        }

        if is_partitioned and partizionato_per:
            table["partizionato_per"] = self._normalize_name(partizionato_per)
            table["partizioni"] = {}

            for entry in entries:
                result = self._read_data_sheet(wb, entry)
                if result is None:
                    continue
                headers, rows = result
                partition_key = str(entry.get("valore_partizione", entry["foglio"]))
                ranges = self._rows_to_ranges(rows, entry, headers)
                table["partizioni"][partition_key] = ranges
                table["_meta"][f"righe_{partition_key}"] = len(rows)
        else:
            # Singolo foglio, nessuna partizione
            entry = entries[0]
            result = self._read_data_sheet(wb, entry)
            if result is None:
                return None
            headers, rows = result
            table["ranges"] = self._rows_to_ranges(rows, entry, headers)
            table["_meta"]["righe"] = len(rows)

        return table

    def _rows_to_ranges(self, rows: List[Dict], entry: Dict, headers: List[str]) -> List[Dict]:
        """Converte righe Excel in range con output per lookup_range."""
        colonna_chiave = entry["colonna_chiave"]
        tech_cols, art_cols = self._separate_art_columns(headers)

        # Estrai valori chiave numerici
        kw_values = []
        valid_rows = []
        for row in rows:
            v = row.get(colonna_chiave)
            if v is not None:
                try:
                    kw_values.append(float(v))
                    valid_rows.append(row)
                except (ValueError, TypeError):
                    self.warnings.append(
                        f"Foglio '{entry['foglio']}': valore '{v}' non numerico "
                        f"nella colonna '{colonna_chiave}', riga ignorata"
                    )

        if not kw_values:
            self.errors.append(
                f"Foglio '{entry['foglio']}': nessun valore numerico "
                f"nella colonna chiave '{colonna_chiave}'"
            )
            return []

        fasce = self._calcola_ranges(kw_values)
        result = []

        for row, (da, a) in zip(valid_rows, fasce):
            output = {}
            for col in tech_cols:
                if col == colonna_chiave:
                    continue  # La chiave non va nell'output
                v = row.get(col)
                if v is not None:
                    output[self._normalize_name(col)] = self._clean_value(v)

            # Colonne ART: separate
            art_output = {}
            for col in art_cols:
                v = row.get(col)
                if v is not None:
                    # Rimuovi prefisso "ART: " dal nome
                    art_name = self._normalize_name(col[4:].strip())
                    art_output[art_name] = str(v).strip()

            range_entry = {"da": da, "a": a, "output": output}
            if art_output:
                range_entry["articoli"] = art_output

            result.append(range_entry)

        return result

    # --------------------------------------------------------------------
    # BUILDER: catalogo
    # --------------------------------------------------------------------
    def _build_catalogo(self, wb, entries: List[Dict]) -> Optional[Dict]:
        """Costruisce tabella catalogo."""
        from collections import Counter
        entry = entries[0]
        nome_tabella = entry["nome_tabella"]
        colonna_chiave = self._normalize_name(entry["colonna_chiave"])
        col_suffisso_raw = entry.get("colonna_suffisso") or ""
        col_suffisso = self._normalize_name(col_suffisso_raw) if col_suffisso_raw else ""
        print(f"[BUILD_CATALOGO] col_suffisso_raw='{col_suffisso_raw}', col_suffisso='{col_suffisso}'")

        result = self._read_data_sheet(wb, entry)
        if result is None:
            return None
        headers, rows = result
        tech_cols, art_cols = self._separate_art_columns(headers)

        # Conta occorrenze per chiave — se >1 e col_suffisso definita, espanderemo con suffisso
        conteggio = Counter()
        for row in rows:
            v = row.get(entry["colonna_chiave"])
            if v is not None:
                conteggio[str(v).strip()] += 1

        records = []
        for row in rows:
            record = {}
            for col in tech_cols:
                v = row.get(col)
                if v is not None:
                    record[self._normalize_name(col)] = self._clean_value(v)
            if art_cols:
                art = {}
                for col in art_cols:
                    v = row.get(col)
                    if v is not None:
                        art[self._normalize_name(col[4:].strip())] = str(v).strip()
                if art:
                    record["_articoli"] = art

            # Espandi chiave con suffisso se componente ha righe multiple
            chiave_val = str(row.get(entry["colonna_chiave"], "")).strip()
            if col_suffisso and conteggio.get(chiave_val, 1) > 1:
                suf_raw = record.get(col_suffisso)
                if suf_raw is not None:
                    try:
                        suf = str(int(float(str(suf_raw))))
                    except (ValueError, TypeError):
                        suf = str(suf_raw).replace(" ", "_").lower()
                    record[colonna_chiave] = f"{self._normalize_name(chiave_val)}_{suf}v"

            records.append(record)

        return {
            "tipo": "catalog",
            "colonna_id": colonna_chiave,
            "colonne": [self._normalize_name(c) for c in tech_cols],
            "records": records,
            "_meta": {
                "nome": nome_tabella,
                "tipo": "catalogo",
                "tipo_engine": "catalog",
                "colonna_suffisso": col_suffisso or None,
                "generato_il": datetime.now().isoformat(),
                "file_origine": "import Excel con _MAPPA",
                "righe_totali": len(records),
                "colonna_chiave": colonna_chiave,
            }
        }

    # --------------------------------------------------------------------
    # BUILDER: costanti
    # --------------------------------------------------------------------
    def _build_costanti(self, wb, entries: List[Dict]) -> Optional[Dict]:
        """Costruisce tabella costanti (chiave → valori)."""
        entry = entries[0]
        nome_tabella = entry["nome_tabella"]
        colonna_chiave = entry["colonna_chiave"]

        result = self._read_data_sheet(wb, entry)
        if result is None:
            return None
        headers, rows = result
        tech_cols, art_cols = self._separate_art_columns(headers)

        valori = {}
        for row in rows:
            key_raw = row.get(colonna_chiave)
            if key_raw is None:
                continue
            key = str(key_raw).strip()

            output = {}
            for col in tech_cols:
                if col == colonna_chiave:
                    continue
                v = row.get(col)
                if v is not None:
                    output[self._normalize_name(col)] = self._clean_value(v)

            if art_cols:
                art = {}
                for col in art_cols:
                    v = row.get(col)
                    if v is not None:
                        art[self._normalize_name(col[4:].strip())] = str(v).strip()
                if art:
                    output["_articoli"] = art

            valori[key] = output

        return {
            "tipo": "lookup_mapping",
            "parametro_lookup": self._normalize_name(colonna_chiave),
            "valori": valori,
            "_meta": {
                "nome": nome_tabella,
                "tipo": "costanti",
                "generato_il": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "file_origine": "import Excel con _MAPPA",
                "righe": len(valori),
            }
        }

    # ========================================================================
    # GENERAZIONE REGOLA LOOKUP
    # ========================================================================
    def _build_lookup_rule(self, nome_tabella: str, entries: List[Dict],
                           table_data: Dict) -> Optional[Dict]:
        """Genera la regola JSON che collega il configuratore alla data table."""
        tipo = entries[0]["tipo"]
        colonna_chiave = entries[0]["colonna_chiave"]
        partizionato_per = entries[0].get("partizionato_per")

        rule_id = f"LOOKUP_{nome_tabella.upper()}"
        nome_leggibile = nome_tabella.replace("_", " ").title()

        rule = {
            "id": rule_id,
            "name": f"Lookup {nome_leggibile}",
            "description": f"Tabella lookup generata da Excel per {nome_leggibile}",
            "version": "1.0",
            "enabled": True,
            "priority": 10,
            "conditions": [],
            "actions": []
        }

        if tipo == "lookup_range":
            action = {
                "action": "lookup_table",
                "tabella": nome_tabella,
                "input_field": f"TODO_SEZIONE.{self._normalize_name(colonna_chiave)}",
                "output_prefix": f"_calc.{nome_tabella}."
            }
            if partizionato_per:
                action["partition_field"] = f"TODO_SEZIONE.{self._normalize_name(partizionato_per)}"

            rule["actions"].append(action)
            rule["_todo"] = (
                "COMPLETARE: sostituire 'TODO_SEZIONE' con il nome della sezione "
                "del configuratore (es. 'argano.potenza_motore_kw'). "
                "Aggiungere conditions appropriate."
            )

        elif tipo == "catalogo":
            action = {
                "action": "catalog_match",
                "tabella": nome_tabella,
                "criteri": [],
                "output_prefix": f"_calc.{nome_tabella}."
            }
            rule["actions"].append(action)
            rule["_todo"] = (
                "COMPLETARE: aggiungere criteri di matching e conditions."
            )

        elif tipo == "costanti":
            action = {
                "action": "lookup_table",
                "tabella": nome_tabella,
                "input_field": f"TODO_SEZIONE.{self._normalize_name(colonna_chiave)}",
                "output_prefix": f"_calc.{nome_tabella}."
            }
            rule["actions"].append(action)
            rule["_todo"] = (
                "COMPLETARE: sostituire 'TODO_SEZIONE' con il campo del configuratore."
            )

        return rule

    # ========================================================================
    # UTILITY
    # ========================================================================
    def _open_workbook(self, filepath: str):
        """Apre il file Excel."""
        try:
            import openpyxl
        except ImportError:
            self.errors.append("Libreria openpyxl non installata")
            return None

        if not os.path.exists(filepath):
            self.errors.append(f"File non trovato: {filepath}")
            return None

        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            self.errors.append(f"Errore apertura file Excel: {e}")
            return None

    def _calcola_ranges(self, values: List[float]) -> List[Tuple[float, float]]:
        """Calcola fasce con punto medio tra valori consecutivi."""
        ranges = []
        for i, v in enumerate(values):
            da = 0 if i == 0 else round((values[i - 1] + v) / 2, 2)
            a = 999 if i == len(values) - 1 else round((v + values[i + 1]) / 2, 2)
            ranges.append((da, a))
        return ranges

    def _normalize_name(self, name) -> str:
        """Normalizza un nome per uso come chiave JSON."""
        if name is None:
            return ""
        s = str(name).strip().lower()
        # Rimuovi prefissi comuni tipo "DIR: " o "ST: "
        s = s.replace(":", "_").replace(".", "_")
        s = s.replace(" ", "_").replace("-", "_")
        s = s.replace("(", "").replace(")", "")
        # Rimuovi doppi underscore
        while "__" in s:
            s = s.replace("__", "_")
        return s.strip("_")

    def _clean_value(self, v):
        """Pulisce un valore dalla cella Excel."""
        if isinstance(v, str):
            v = v.strip()
            # Prova conversione numerica
            try:
                if "." in v:
                    return float(v)
                return int(v)
            except ValueError:
                return v
        return v

    def _save_json(self, data: Dict, filepath: str):
        """Salva un dict come JSON."""
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.errors.append(f"Errore scrittura {filepath}: {e}")

    def _result(self, success: bool) -> Dict[str, Any]:
        """Template risultato."""
        return {
            "success": success,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
        }
