"""
analytics_api.py — Pannello analisi self-service
=================================================
Motore di query dinamico con whitelist di fonti e campi.
Nessun SQL esposto all'utente — tutto tramite parametri JSON strutturati.

Endpoint:
  GET  /analytics/fonti               → elenco fonti disponibili + campi
  POST /analytics/query               → esegui query con filtri + colonne
  POST /analytics/export-excel        → scarica risultato come .xlsx
  POST /analytics/export-pdf          → scarica risultato come .pdf
  POST /analytics/viste               → salva una vista
  GET  /analytics/viste               → elenco viste salvate
  DELETE /analytics/viste/{id}        → elimina vista
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

try:
    from database import get_db
except ImportError:
    from main import get_db

router = APIRouter(prefix="/analytics", tags=["analytics"])

# ══════════════════════════════════════════════════════════════════════════════
# CATALOGO FONTI E CAMPI (whitelist)
# ══════════════════════════════════════════════════════════════════════════════

FONTI: dict[str, dict] = {
    "commesse": {
        "label": "Commesse / Ordini",
        "query": """
            SELECT
                p.id                        AS id,
                p.numero_preventivo         AS numero_preventivo,
                p.tipo_preventivo           AS tipo,
                p.status                    AS stato,
                c.ragione_sociale           AS cliente,
                dc.data_offerta             AS data_offerta,
                dc.consegna_richiesta       AS consegna_richiesta,
                dc.prezzo_unitario          AS prezzo_unitario,
                dc.quantita                 AS quantita,
                p.created_at                AS creato_il
            FROM preventivi p
            LEFT JOIN clienti c ON c.id = p.cliente_id
            LEFT JOIN dati_commessa dc ON dc.preventivo_id = p.id
        """,
        "campi": {
            "id":                   {"label": "ID",                "tipo": "numero",   "filtrabile": True},
            "numero_preventivo":    {"label": "Numero",            "tipo": "testo",    "filtrabile": True},
            "tipo":                 {"label": "Tipo",              "tipo": "testo",    "filtrabile": True},
            "stato":                {"label": "Stato",             "tipo": "testo",    "filtrabile": True},
            "cliente":              {"label": "Cliente",           "tipo": "testo",    "filtrabile": True},
            "data_offerta":         {"label": "Data offerta",      "tipo": "data",     "filtrabile": True},
            "consegna_richiesta":   {"label": "Consegna richiesta","tipo": "data",     "filtrabile": True},
            "prezzo_unitario":      {"label": "Prezzo unitario",   "tipo": "numero",   "filtrabile": True},
            "quantita":             {"label": "Quantità",          "tipo": "numero",   "filtrabile": True},
            "creato_il":            {"label": "Creato il",         "tipo": "data",     "filtrabile": True},
        },
    },
    "oda": {
        "label": "Ordini di Acquisto",
        "query": """
            SELECT
                o.id                            AS id,
                o.numero_oda                    AS numero_oda,
                o.stato                         AS stato,
                o.fornitore_denominazione       AS fornitore,
                o.data_emissione                AS data_ordine,
                o.data_consegna_richiesta       AS data_consegna,
                o.imponibile_totale             AS importo,
                o.note                          AS note,
                o.created_at                    AS creato_il
            FROM ordini_acquisto o
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "numero_oda":       {"label": "Numero ODA",      "tipo": "testo",  "filtrabile": True},
            "stato":            {"label": "Stato",           "tipo": "testo",  "filtrabile": True},
            "fornitore":        {"label": "Fornitore",       "tipo": "testo",  "filtrabile": True},
            "data_ordine":      {"label": "Data ordine",     "tipo": "data",   "filtrabile": True},
            "data_consegna":    {"label": "Data consegna",   "tipo": "data",   "filtrabile": True},
            "importo":          {"label": "Importo (€)",     "tipo": "numero", "filtrabile": True},
            "note":             {"label": "Note",            "tipo": "testo",  "filtrabile": False},
            "creato_il":        {"label": "Creato il",       "tipo": "data",   "filtrabile": True},
        },
    },
    "materiali": {
        "label": "Materiali / Distinta",
        "query": """
            SELECT
                m.id                        AS id,
                m.preventivo_id             AS commessa_id,
                p.numero_preventivo         AS commessa,
                m.codice                    AS codice_articolo,
                m.descrizione               AS descrizione,
                m.quantita                  AS quantita,
                m.prezzo_unitario           AS prezzo_unitario,
                m.categoria                 AS categoria,
                m.created_at                AS creato_il
            FROM materiali m
            LEFT JOIN preventivi p ON p.id = m.preventivo_id
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "commessa_id":      {"label": "ID Commessa",     "tipo": "numero", "filtrabile": True},
            "commessa":         {"label": "Commessa",        "tipo": "testo",  "filtrabile": True},
            "codice_articolo":  {"label": "Codice articolo", "tipo": "testo",  "filtrabile": True},
            "descrizione":      {"label": "Descrizione",     "tipo": "testo",  "filtrabile": True},
            "quantita":         {"label": "Quantità",        "tipo": "numero", "filtrabile": True},
            "prezzo_unitario":  {"label": "Prezzo unitario", "tipo": "numero", "filtrabile": True},
            "categoria":        {"label": "Categoria",       "tipo": "testo",  "filtrabile": True},
            "creato_il":        {"label": "Creato il",       "tipo": "data",   "filtrabile": True},
        },
    },
    "produzione": {
        "label": "Produzione — Fasi",
        "query": """
            SELECT
                fp.id                       AS id,
                fp.ordine_id                AS commessa_id,
                p.numero_preventivo         AS commessa,
                fp.nome                     AS fase,
                fp.stato                    AS stato,
                fp.data_inizio_prevista     AS inizio_previsto,
                fp.data_fine_prevista       AS fine_prevista,
                fp.data_inizio_reale        AS inizio_reale,
                fp.data_fine_reale          AS fine_reale,
                fp.durata_stimata_ore       AS ore_stimate,
                fp.durata_reale_ore         AS ore_reali,
                cl.nome                     AS centro_lavoro
            FROM fasi_produzione fp
            LEFT JOIN ordini o  ON o.id = fp.ordine_id
            LEFT JOIN preventivi p ON p.id = o.preventivo_id
            LEFT JOIN centri_lavoro cl ON cl.id = fp.centro_lavoro_id
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "commessa_id":      {"label": "ID Commessa",     "tipo": "numero", "filtrabile": True},
            "commessa":         {"label": "Commessa",        "tipo": "testo",  "filtrabile": True},
            "fase":             {"label": "Fase",            "tipo": "testo",  "filtrabile": True},
            "stato":            {"label": "Stato",           "tipo": "testo",  "filtrabile": True},
            "inizio_previsto":  {"label": "Inizio previsto", "tipo": "data",   "filtrabile": True},
            "fine_prevista":    {"label": "Fine prevista",   "tipo": "data",   "filtrabile": True},
            "inizio_reale":     {"label": "Inizio reale",    "tipo": "data",   "filtrabile": True},
            "fine_reale":       {"label": "Fine reale",      "tipo": "data",   "filtrabile": True},
            "ore_stimate":      {"label": "Ore stimate",     "tipo": "numero", "filtrabile": True},
            "ore_reali":        {"label": "Ore reali",       "tipo": "numero", "filtrabile": True},
            "centro_lavoro":    {"label": "Centro lavoro",   "tipo": "testo",  "filtrabile": True},
        },
    },
    "tempi": {
        "label": "Tempi lavoro (ticket)",
        "query": """
            SELECT
                ts.id                       AS id,
                ts.ticket_id                AS ticket_id,
                t.titolo                    AS ticket,
                u.username                  AS operatore,
                ts.inizio                   AS inizio,
                ts.fine                     AS fine,
                ts.durata_minuti            AS minuti,
                t.numero_ticket             AS numero_ticket,
                t.priorita                  AS priorita
            FROM ticket_sessioni_lavoro ts
            LEFT JOIN tickets t ON t.id = ts.ticket_id
            LEFT JOIN utenti u ON u.id = ts.utente_id
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "ticket_id":        {"label": "Ticket ID",       "tipo": "numero", "filtrabile": True},
            "ticket":           {"label": "Ticket",          "tipo": "testo",  "filtrabile": True},
            "operatore":        {"label": "Operatore",       "tipo": "testo",  "filtrabile": True},
            "inizio":           {"label": "Inizio",          "tipo": "data",   "filtrabile": True},
            "fine":             {"label": "Fine",            "tipo": "data",   "filtrabile": True},
            "minuti":           {"label": "Minuti",          "tipo": "numero", "filtrabile": True},
            "numero_ticket":    {"label": "N. Ticket",       "tipo": "testo",  "filtrabile": True},
            "priorita":         {"label": "Priorità",        "tipo": "testo",  "filtrabile": True},
        },
    },
    "ticket": {
        "label": "Ticket assistenza",
        "query": """
            SELECT
                t.id                        AS id,
                t.numero_ticket             AS numero_ticket,
                t.titolo                    AS titolo,
                t.stato                     AS stato,
                t.priorita                  AS priorita,
                u.username                  AS assegnato_a,
                t.created_at                AS creato_il,
                t.updated_at                AS aggiornato_il
            FROM tickets t
            LEFT JOIN utenti u ON u.id = t.assegnato_a
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "numero_ticket":    {"label": "N. Ticket",       "tipo": "testo",  "filtrabile": True},
            "titolo":           {"label": "Titolo",          "tipo": "testo",  "filtrabile": True},
            "stato":            {"label": "Stato",           "tipo": "testo",  "filtrabile": True},
            "priorita":         {"label": "Priorità",        "tipo": "testo",  "filtrabile": True},
            "assegnato_a":      {"label": "Assegnato a",     "tipo": "testo",  "filtrabile": True},
            "creato_il":        {"label": "Creato il",       "tipo": "data",   "filtrabile": True},
            "aggiornato_il":    {"label": "Aggiornato il",   "tipo": "data",   "filtrabile": True},
        },
    },
    "magazzino": {
        "label": "Magazzino — Movimenti",
        "query": """
            SELECT
                mm.id                       AS id,
                mm.articolo_id              AS articolo_id,
                mm.codice_articolo          AS codice_articolo,
                a.descrizione               AS descrizione,
                mm.tipo                     AS tipo,
                mm.quantita                 AS quantita,
                mm.note                     AS note,
                mm.utente                   AS utente,
                mm.data_movimento           AS data_movimento
            FROM magazzino_movimenti mm
            LEFT JOIN articoli a ON a.id = mm.articolo_id
        """,
        "campi": {
            "id":               {"label": "ID",              "tipo": "numero", "filtrabile": True},
            "articolo_id":      {"label": "ID Articolo",     "tipo": "numero", "filtrabile": True},
            "codice_articolo":  {"label": "Codice articolo", "tipo": "testo",  "filtrabile": True},
            "descrizione":      {"label": "Descrizione",     "tipo": "testo",  "filtrabile": True},
            "tipo":             {"label": "Tipo movimento",  "tipo": "testo",  "filtrabile": True},
            "quantita":         {"label": "Quantità",        "tipo": "numero", "filtrabile": True},
            "note":             {"label": "Note",            "tipo": "testo",  "filtrabile": False},
            "utente":           {"label": "Utente",          "tipo": "testo",  "filtrabile": True},
            "data_movimento":   {"label": "Data",            "tipo": "data",   "filtrabile": True},
        },
    },
}

# ══════════════════════════════════════════════════════════════════════════════
# OPERATORI FILTRO CONSENTITI
# ══════════════════════════════════════════════════════════════════════════════

OPERATORI = {
    "uguale":           ("=",        False),
    "diverso":          ("!=",       False),
    "contiene":         ("LIKE",     True),   # True = wrap con %
    "non_contiene":     ("NOT LIKE", True),
    "inizia_con":       ("LIKE",     "start"),
    "finisce_con":      ("LIKE",     "end"),
    "maggiore":         (">",        False),
    "maggiore_uguale":  (">=",       False),
    "minore":           ("<",        False),
    "minore_uguale":    ("<=",       False),
    "vuoto":            ("IS NULL",  None),
    "non_vuoto":        ("IS NOT NULL", None),
}

# ══════════════════════════════════════════════════════════════════════════════
# MODELLI
# ══════════════════════════════════════════════════════════════════════════════

class Filtro(BaseModel):
    campo: str
    operatore: str
    valore: Optional[Any] = None

class QueryRequest(BaseModel):
    fonte: str
    colonne: List[str]
    filtri: List[Filtro] = []
    ordina_per: Optional[str] = None
    ordine_desc: bool = False
    limite: int = 500

class VistaCreate(BaseModel):
    nome: str
    fonte: str
    colonne: List[str]
    filtri: List[dict] = []
    ordina_per: Optional[str] = None
    ordine_desc: bool = False

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _valida_fonte(fonte: str) -> dict:
    if fonte not in FONTI:
        raise HTTPException(400, f"Fonte '{fonte}' non disponibile")
    return FONTI[fonte]

def _valida_colonne(fonte_cfg: dict, colonne: list[str]) -> list[str]:
    campi_ok = set(fonte_cfg["campi"].keys())
    bad = [c for c in colonne if c not in campi_ok]
    if bad:
        raise HTTPException(400, f"Campi non disponibili: {bad}")
    return colonne if colonne else list(campi_ok)

def _build_query(fonte_cfg: dict, colonne: list[str], filtri: list[Filtro],
                 ordina_per: Optional[str], ordine_desc: bool) -> tuple[str, dict]:
    base = fonte_cfg["query"].strip().rstrip(";")
    campi_ok = set(fonte_cfg["campi"].keys())

    # SELECT con solo le colonne richieste
    sel = ", ".join(f"t.{c}" for c in colonne) if colonne else "t.*"
    sql = f"SELECT {sel} FROM ({base}) t"

    params: dict = {}
    where_parts: list[str] = []

    for i, f in enumerate(filtri):
        if f.campo not in campi_ok:
            raise HTTPException(400, f"Campo filtro '{f.campo}' non disponibile")
        if f.operatore not in OPERATORI:
            raise HTTPException(400, f"Operatore '{f.operatore}' non riconosciuto")
        if not fonte_cfg["campi"][f.campo].get("filtrabile", False):
            raise HTTPException(400, f"Campo '{f.campo}' non filtrabile")

        op, wrap = OPERATORI[f.operatore]
        col = f"t.{f.campo}"

        if wrap is None:  # IS NULL / IS NOT NULL
            where_parts.append(f"({col} {op})")
        elif wrap is False:
            key = f"p{i}"
            where_parts.append(f"({col} {op} :{key})")
            params[key] = f.valore
        elif wrap is True:  # LIKE con % su entrambi i lati
            key = f"p{i}"
            where_parts.append(f"({col} {op} :{key})")
            params[key] = f"%{f.valore}%"
        elif wrap == "start":
            key = f"p{i}"
            where_parts.append(f"({col} {op} :{key})")
            params[key] = f"{f.valore}%"
        elif wrap == "end":
            key = f"p{i}"
            where_parts.append(f"({col} {op} :{key})")
            params[key] = f"%{f.valore}"

    if where_parts:
        sql += " WHERE " + " AND ".join(where_parts)

    if ordina_per and ordina_per in campi_ok:
        direction = "DESC" if ordine_desc else "ASC"
        sql += f" ORDER BY t.{ordina_per} {direction}"

    return sql, params


def _esegui_query(db: Session, fonte_cfg: dict, colonne: list[str],
                  filtri: list[Filtro], ordina_per: Optional[str],
                  ordine_desc: bool, limite: int) -> tuple[list[str], list[list]]:
    if not colonne:
        colonne = list(fonte_cfg["campi"].keys())

    sql, params = _build_query(fonte_cfg, colonne, filtri, ordina_per, ordine_desc)
    sql_limited = sql + f" LIMIT {min(limite, 5000)}"

    try:
        rows = db.execute(text(sql_limited), params).fetchall()
    except Exception as e:
        raise HTTPException(500, f"Errore query: {e}")

    data = [list(r) for r in rows]
    return colonne, data


def _tabella_a_excel(colonne: list[str], righe: list[list],
                     labels: dict[str, str]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl non installato")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisi"

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)

    for ci, col in enumerate(colonne, 1):
        cell = ws.cell(row=1, column=ci, value=labels.get(col, col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(15, len(labels.get(col, col)) + 4)

    for ri, row in enumerate(righe, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tabella_a_pdf(colonne: list[str], righe: list[list],
                   labels: dict[str, str], titolo: str) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
    except ImportError:
        raise HTTPException(500, "reportlab non installato")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(titolo, styles["Title"]))
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph(
        f"Esportato il {datetime.now().strftime('%d/%m/%Y %H:%M')} — {len(righe)} righe",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 0.4*cm))

    headers = [labels.get(c, c) for c in colonne]
    table_data = [headers] + [[str(v) if v is not None else "" for v in row] for row in righe]

    col_w = (landscape(A4)[0] - 2*cm) / max(len(colonne), 1)
    col_widths = [col_w] * len(colonne)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/fonti")
def get_fonti():
    result = []
    for codice, cfg in FONTI.items():
        result.append({
            "codice": codice,
            "label": cfg["label"],
            "campi": [
                {
                    "codice": k,
                    "label": v["label"],
                    "tipo": v["tipo"],
                    "filtrabile": v.get("filtrabile", False),
                }
                for k, v in cfg["campi"].items()
            ],
        })
    return result


@router.post("/query")
def esegui_query(body: QueryRequest, db: Session = Depends(get_db)):
    import traceback
    try:
        fonte_cfg = _valida_fonte(body.fonte)
        colonne = _valida_colonne(fonte_cfg, body.colonne)
        colonne, righe = _esegui_query(
            db, fonte_cfg, colonne, body.filtri,
            body.ordina_per, body.ordine_desc, body.limite,
        )
        labels = {k: v["label"] for k, v in fonte_cfg["campi"].items()}
        return {
            "colonne": colonne,
            "labels": {c: labels.get(c, c) for c in colonne},
            "righe": righe,
            "totale": len(righe),
        }
    except Exception as e:
        traceback.print_exc()
        raise


@router.post("/export-excel")
def export_excel(body: QueryRequest, db: Session = Depends(get_db)):
    fonte_cfg = _valida_fonte(body.fonte)
    colonne = _valida_colonne(fonte_cfg, body.colonne)
    colonne, righe = _esegui_query(
        db, fonte_cfg, colonne, body.filtri,
        body.ordina_per, body.ordine_desc, body.limite,
    )
    labels = {k: v["label"] for k, v in fonte_cfg["campi"].items()}
    data = _tabella_a_excel(colonne, righe, labels)
    nome = f"analisi_{body.fonte}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@router.post("/export-pdf")
def export_pdf(body: QueryRequest, db: Session = Depends(get_db)):
    fonte_cfg = _valida_fonte(body.fonte)
    colonne = _valida_colonne(fonte_cfg, body.colonne)
    colonne, righe = _esegui_query(
        db, fonte_cfg, colonne, body.filtri,
        body.ordina_per, body.ordine_desc, body.limite,
    )
    labels = {k: v["label"] for k, v in fonte_cfg["campi"].items()}
    titolo = f"Analisi — {fonte_cfg['label']}"
    data = _tabella_a_pdf(colonne, righe, labels, titolo)
    nome = f"analisi_{body.fonte}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@router.get("/viste")
def get_viste(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text(
            "SELECT id, nome, fonte, colonne, filtri, ordina_per, ordine_desc, created_at "
            "FROM analytics_viste ORDER BY created_at DESC"
        )).fetchall()
    except Exception:
        return []
    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "nome": r[1],
            "fonte": r[2],
            "colonne": json.loads(r[3]) if r[3] else [],
            "filtri": json.loads(r[4]) if r[4] else [],
            "ordina_per": r[5],
            "ordine_desc": bool(r[6]),
            "created_at": r[7],
        })
    return result


@router.post("/viste")
def salva_vista(body: VistaCreate, db: Session = Depends(get_db)):
    _valida_fonte(body.fonte)
    fonte_cfg = FONTI[body.fonte]
    _valida_colonne(fonte_cfg, body.colonne)

    # Crea tabella se non esiste
    try:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS analytics_viste (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                fonte TEXT NOT NULL,
                colonne TEXT,
                filtri TEXT,
                ordina_per TEXT,
                ordine_desc INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now'))
            )
        """))
        db.commit()
    except Exception:
        pass

    now = datetime.now().isoformat()
    try:
        result = db.execute(text("""
            INSERT INTO analytics_viste (nome, fonte, colonne, filtri, ordina_per, ordine_desc, created_at)
            VALUES (:nome, :fonte, :colonne, :filtri, :ordina_per, :ordine_desc, :now)
        """), {
            "nome": body.nome,
            "fonte": body.fonte,
            "colonne": json.dumps(body.colonne),
            "filtri": json.dumps(body.filtri),
            "ordina_per": body.ordina_per,
            "ordine_desc": 1 if body.ordine_desc else 0,
            "now": now,
        })
        db.commit()
        return {"id": result.lastrowid, "nome": body.nome}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))


@router.delete("/viste/{vista_id}")
def elimina_vista(vista_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM analytics_viste WHERE id = :id"), {"id": vista_id})
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))
    return {"deleted": vista_id}
