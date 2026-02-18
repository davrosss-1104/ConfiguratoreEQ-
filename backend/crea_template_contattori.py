"""
crea_template_contattori.py
===========================
Crea il template Excel che Elettroquadri deve compilare.
Il file generato ha 2 fogli (50Hz e 60Hz) con le colonne corrette
e i dati attuali già pre-compilati come esempio.

ESECUZIONE:
  python crea_template_contattori.py
  → genera: template_contattori_oleo.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, size=9, name="Arial")
DATA_FONT = Font(size=9, name="Arial")
NONE_FONT = Font(size=9, name="Arial", color="999999")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("A", "kW",              8,  "Potenza nominale motore (kW)"),
    ("B", "IN (A)",          8,  "Corrente nominale (A)"),
    # Diretto
    ("C", "DIR: Cont.",      10, "Taglia contattore KM diretto (es. D18). Vuoto = non disponibile"),
    ("D", "DIR: Mors.",      10, "Sezione morsetti mm² diretto"),
    ("E", "DIR: Filo",       10, "Sezione filo mm² diretto"),
    # Stella-Triangolo
    ("F", "ST: KS",          10, "Taglia contattore KS stella"),
    ("G", "ST: KM",          10, "Taglia contattore KM triangolo. Vuoto = non previsto"),
    ("H", "ST: Mors RST",    12, "Sezione morsetti mm² RST stella-triangolo"),
    ("I", "ST: Mors UVW",    12, "Sezione morsetti mm² UVW stella-triangolo"),
    ("J", "ST: Filo RST",    12, "Sezione filo mm² RST stella-triangolo"),
    ("K", "ST: Filo UVW",    12, "Sezione filo mm² UVW stella-triangolo"),
    # Soft Starter
    ("L", "SS: Cont.",       10, "Taglia contattore bypass soft starter. Vuoto = solo SS"),
    ("M", "SS: Mors.",       10, "Sezione morsetti mm² soft starter"),
    ("N", "SS: Filo",        10, "Sezione filo mm² soft starter"),
    ("O", "SS: Modello",     12, "Modello soft starter (es. V40, V70, V105, V150)"),
]

TABELLA_50HZ = [
    (4.4,  10.4,  "D18",10,2.5,   "D18","D12",10,10,2.5,2.5,   "D18",10,2.5,"V40"),
    (5.9,  14.6,  "D18",10,4,     "D18","D12",10,10,2.5,2.5,   "D18",10,2.5,"V40"),
    (7.7,  18.5,  "D25",10,4,     "D18","D12",10,10,2.5,2.5,   "D18",10,2.5,"V40"),
    (9.6,  23.4,  "D25",10,4,     "D18","D12",10,10,2.5,2.5,   "D18",10,2.5,"V40"),
    (11.8, 27.8,  "D32",10,6,     "D25","D18",10,10,4,4,       "D25",10,4,"V40"),
    (14.7, 32,    "D50",16,10,    "D32","D25",10,10,6,6,       "D25",10,4,"V40"),
    (18.4, 40,    "D80",35,16,    "D50","D32",16,16,10,10,     "D32",10,6,"V40"),
    (22.1, 47,    "D80",35,16,    "D50","D32",16,16,10,10,     "D50",16,10,"V70"),
    (29.4, 63,    None,None,None, "D80","D50",35,35,16,16,     "D50",16,10,"V70"),
    (36.8, 77.3,  None,None,None, "D80","D50",35,35,16,16,     "D80",35,16,"V70"),
    (44.1, 91.4,  None,None,None, None,None,None,None,None,None, None,None,None,"V105"),
    (58.8, 119.1, None,None,None, None,None,None,None,None,None, None,None,None,"V105"),
    (73.5, 146.3, None,None,None, None,None,None,None,None,None, None,None,None,"V150"),
]

TABELLA_60HZ = [
    (3.6,  7.9,   "D18",10,2.5,   "D18",None,10,10,10,10,     "D18",10,4,"V40"),
    (5.3,  11.4,  "D18",10,2.5,   "D18",None,10,10,10,10,     "D18",10,4,"V40"),
    (7.3,  15,    "D25",16,4,     "D18",None,10,10,10,10,     "D18",10,4,"V40"),
    (9.2,  18.9,  "D25",16,4,     "D18",None,10,10,10,10,     "D18",10,4,"V40"),
    (11,   23.1,  "D32",16,6,     "D25",None,16,16,16,16,     "D25",10,6,"V40"),
    (15,   31,    "D50",16,10,    "D32",None,16,16,16,16,     "D25",10,6,"V40"),
    (18.5, 36,    "D50",35,10,    "D50",None,35,35,35,35,     "D32",16,6,"V40"),
    (24,   46,    "D80",35,16,    "D80",None,35,35,35,35,     "D50",16,10,"V40"),
    (29,   54,    "D80",35,16,    "D80",None,35,35,35,35,     "D50",16,10,"V70"),
    (37,   70,    None,None,None, "D80",None,35,35,35,35,     "D80",35,16,"V70"),
    (48,   90,    None,None,None, None,None,None,None,None,None, None,None,None,"V105"),
    (57,   105,   None,None,None, None,None,None,None,None,None, None,None,None,"V105"),
    (72,   131,   None,None,None, None,None,None,None,None,None, None,None,None,"V150"),
]


def popola_foglio(ws, titolo, tabella):
    # Riga 1: titolo
    ws.merge_cells("A1:O1")
    ws["A1"] = titolo
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Riga 2: sottogruppi
    for col, start, end, label in [
        ("A", "A2", "B2", "Motore"),
        ("C", "C2", "E2", "Avviamento Diretto"),
        ("F", "F2", "K2", "Avviamento Stella-Triangolo"),
        ("L", "L2", "O2", "Avviamento Soft Starter"),
    ]:
        ws.merge_cells(f"{start}:{end}")
        ws[start] = label
        ws[start].font = SUBHEADER_FONT
        ws[start].fill = SUBHEADER_FILL
        ws[start].alignment = Alignment(horizontal="center")

    # Riga 3: intestazioni colonna
    for col_letter, header, width, comment_text in COLUMNS:
        cell = ws[f"{col_letter}3"]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width
        if comment_text:
            from openpyxl.comments import Comment
            cell.comment = Comment(comment_text, "Configuratore")

    # Riga 4+: dati
    for row_idx, row_data in enumerate(tabella, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if value is None:
                cell.font = NONE_FONT
            else:
                cell.value = value
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center")

    # Riga istruzioni in fondo
    last_row = len(tabella) + 5
    ws.merge_cells(f"A{last_row}:O{last_row}")
    ws[f"A{last_row}"] = ("ISTRUZIONI: Celle vuote = combinazione non disponibile. "
                           "Aggiungere righe per nuove potenze. Non modificare le intestazioni.")
    ws[f"A{last_row}"].font = Font(italic=True, size=9, color="666666", name="Arial")


def main():
    wb = Workbook()
    ws50 = wb.active
    ws50.title = "50Hz_400V"
    popola_foglio(ws50, "Tabella Contattori Oleodinamici - 50Hz / 400V", TABELLA_50HZ)

    ws60 = wb.create_sheet("60Hz_440V")
    popola_foglio(ws60, "Tabella Contattori Oleodinamici - 60Hz / 440V", TABELLA_60HZ)

    output = "template_contattori_oleo.xlsx"
    wb.save(output)
    print(f"✅ Template creato: {output}")
    print(f"   Foglio 1: 50Hz/400V ({len(TABELLA_50HZ)} righe)")
    print(f"   Foglio 2: 60Hz/440V ({len(TABELLA_60HZ)} righe)")


if __name__ == "__main__":
    main()
