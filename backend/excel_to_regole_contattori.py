"""
excel_to_regole_contattori.py
=============================
Legge il file Excel delle tabelle contattori oleodinamici e genera
automaticamente le regole JSON (1 lookup_table + 5 materiali).

ESECUZIONE:
  python excel_to_regole_contattori.py [percorso_excel]

  Se non viene passato un percorso, cerca: template_contattori_oleo.xlsx

FORMATO EXCEL ATTESO:
  - 2 fogli: "50Hz_400V" e "60Hz_440V" (o i primi 2 fogli)
  - Riga 3: intestazioni (kW, IN, DIR:Cont, DIR:Mors, DIR:Filo,
    ST:KS, ST:KM, ST:Mors RST, ST:Mors UVW, ST:Filo RST, ST:Filo UVW,
    SS:Cont, SS:Mors, SS:Filo, SS:Modello)
  - Riga 4+: dati (celle vuote = non disponibile)

OUTPUT:
  ./rules/rule_CALC_OLEO_CONTATTORI.json  (1 regola lookup)
  ./rules/rule_MAT_OLEO_*.json            (5 regole materiali)
"""

import json
import os
import sys
import glob

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl non installato. Esegui: pip install openpyxl")
    sys.exit(1)

RULES_DIR = "./rules"

# Mapping colonne Excel → indici (1-based come openpyxl)
# A=1  B=2  C=3  D=4  E=5  F=6  G=7  H=8  I=9  J=10  K=11  L=12  M=13  N=14  O=15
COL = {
    "kw": 1, "in_a": 2,
    "dir_cont": 3, "dir_mors": 4, "dir_filo": 5,
    "st_ks": 6, "st_km": 7, "st_mors_rst": 8, "st_mors_uvw": 9,
    "st_filo_rst": 10, "st_filo_uvw": 11,
    "ss_cont": 12, "ss_mors": 13, "ss_filo": 14, "ss_model": 15,
}

HEADER_ROW = 3   # Riga con le intestazioni
DATA_START = 4   # Prima riga dati


def cell_val(ws, row, col):
    """Legge una cella, ritorna None se vuota."""
    v = ws.cell(row=row, column=col).value
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    return v


def read_sheet(ws):
    """Legge un foglio Excel e ritorna lista di dict per ogni riga."""
    rows = []
    row_idx = DATA_START
    
    while True:
        kw = cell_val(ws, row_idx, COL["kw"])
        if kw is None:
            break  # Fine dati
        
        try:
            kw = float(kw)
        except (ValueError, TypeError):
            print(f"  ⚠️  Riga {row_idx}: kW '{kw}' non numerico, skip")
            row_idx += 1
            continue
        
        row = {"kw": kw, "in_a": cell_val(ws, row_idx, COL["in_a"])}
        
        # Diretto
        dir_cont = cell_val(ws, row_idx, COL["dir_cont"])
        if dir_cont is not None:
            row["dir"] = {
                "cont": str(dir_cont),
                "mors": cell_val(ws, row_idx, COL["dir_mors"]),
                "filo": cell_val(ws, row_idx, COL["dir_filo"]),
            }
        else:
            row["dir"] = None
        
        # Stella-Triangolo
        st_ks = cell_val(ws, row_idx, COL["st_ks"])
        if st_ks is not None:
            row["st"] = {
                "ks": str(st_ks),
                "km": str(cell_val(ws, row_idx, COL["st_km"])) if cell_val(ws, row_idx, COL["st_km"]) else None,
                "mors_rst": cell_val(ws, row_idx, COL["st_mors_rst"]),
                "mors_uvw": cell_val(ws, row_idx, COL["st_mors_uvw"]),
                "filo_rst": cell_val(ws, row_idx, COL["st_filo_rst"]),
                "filo_uvw": cell_val(ws, row_idx, COL["st_filo_uvw"]),
            }
        else:
            row["st"] = None
        
        # Soft Starter
        ss_model = cell_val(ws, row_idx, COL["ss_model"])
        if ss_model is not None:
            row["ss"] = {
                "cont": str(cell_val(ws, row_idx, COL["ss_cont"])) if cell_val(ws, row_idx, COL["ss_cont"]) else None,
                "mors": cell_val(ws, row_idx, COL["ss_mors"]),
                "filo": cell_val(ws, row_idx, COL["ss_filo"]),
                "model": str(ss_model),
            }
        else:
            row["ss"] = None
        
        rows.append(row)
        row_idx += 1
    
    return rows


def calcola_ranges(rows):
    """Calcola range kW con punto medio tra valori consecutivi."""
    kw_list = [r["kw"] for r in rows]
    ranges = []
    for i, kw in enumerate(kw_list):
        kw_min = 0 if i == 0 else round((kw_list[i - 1] + kw) / 2, 2)
        kw_max = 999 if i == len(kw_list) - 1 else round((kw + kw_list[i + 1]) / 2, 2)
        ranges.append((kw_min, kw_max))
    return ranges


def row_to_set_vars(row):
    """Converte una riga parsed nel dict 'set' per lookup_table."""
    s = {"_calc.potenza_kw": str(row["kw"])}
    
    # Diretto
    if row["dir"]:
        s["_calc.dir_disponibile"] = "si"
        s["_calc.cont_dir"] = row["dir"]["cont"]
        s["_calc.mors_dir"] = str(row["dir"]["mors"])
        s["_calc.filo_dir"] = str(row["dir"]["filo"])
    else:
        s["_calc.dir_disponibile"] = "no"
    
    # Stella-Triangolo
    if row["st"]:
        s["_calc.st_disponibile"] = "si"
        s["_calc.cont_ks"] = row["st"]["ks"]
        s["_calc.cont_km_st"] = row["st"]["km"] or "NONE"
        s["_calc.mors_rst_st"] = str(row["st"]["mors_rst"])
        s["_calc.mors_uvw_st"] = str(row["st"]["mors_uvw"])
        s["_calc.filo_rst_st"] = str(row["st"]["filo_rst"])
        s["_calc.filo_uvw_st"] = str(row["st"]["filo_uvw"])
    else:
        s["_calc.st_disponibile"] = "no"
    
    # Soft Starter
    if row["ss"]:
        s["_calc.ss_model"] = row["ss"]["model"]
        s["_calc.cont_ss"] = row["ss"]["cont"] or "NONE"
        s["_calc.mors_ss"] = str(row["ss"]["mors"]) if row["ss"].get("mors") else "NONE"
        s["_calc.filo_ss"] = str(row["ss"]["filo"]) if row["ss"].get("filo") else "NONE"
    
    return s


def build_lookup_rows(rows):
    """Costruisce le righe lookup da una lista di righe parsed."""
    ranges = calcola_ranges(rows)
    return [
        {"min": rng[0], "max": rng[1], "set": row_to_set_vars(row)}
        for row, rng in zip(rows, ranges)
    ]


def save_rule(rule, filename):
    filepath = os.path.join(RULES_DIR, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(rule, f, indent=2, ensure_ascii=False)


def cleanup_old():
    patterns = ["rule_CONT_OLEO_*.json", "rule_CALC_OLEO_*.json", "rule_MAT_OLEO_*.json"]
    removed = 0
    for pat in patterns:
        for f in glob.glob(os.path.join(RULES_DIR, pat)):
            os.remove(f)
            removed += 1
    return removed


def genera_regola_calcolo(rows_50, rows_60):
    rule = {
        "id": "CALC_OLEO_CONTATTORI",
        "name": "Calcolo componenti contattori oleodinamici",
        "description": "Lookup table: potenza motore + frequenza → taglie contattori, morsetti, fili",
        "version": "3.0",
        "phase": 1,
        "priority": 10,
        "enabled": True,
        "conditions": [
            {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
            {"field": "argano.potenza_motore_kw", "operator": "is_not_empty"},
        ],
        "actions": [{
            "action": "lookup_table",
            "lookup_field": "argano.potenza_motore_kw",
            "partition_field": "tensioni.frequenza_rete",
            "rows": {
                "50": build_lookup_rows(rows_50),
                "60": build_lookup_rows(rows_60),
            }
        }],
        "materials": []
    }
    save_rule(rule, "rule_CALC_OLEO_CONTATTORI.json")
    return 1


def genera_regole_materiali():
    """Le 5 regole Phase 2 sono fisse (usano placeholder _calc.*)."""
    rules = [
        {
            "id": "MAT_OLEO_DIRETTO",
            "name": "Materiali contattori oleo - Avviamento diretto",
            "version": "3.0", "phase": 2, "priority": 50, "enabled": True,
            "conditions": [
                {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
                {"field": "argano.tipo_avviamento_motore", "operator": "equals", "value": "diretto"},
                {"field": "_calc.dir_disponibile", "operator": "equals", "value": "si"},
            ],
            "materials": [
                {"codice": "CONT-{{_calc.cont_dir}}-KM", "descrizione": "Contattore {{_calc.cont_dir}} marcia (KM) - motore {{_calc.potenza_kw}}kW", "quantita": 1, "prezzo_unitario": 0, "categoria": "Contattori"},
                {"codice": "MORS-{{_calc.mors_dir}}MM2-RST", "descrizione": "Morsetti {{_calc.mors_dir}}mm² linea R-S-T", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "MORS-{{_calc.mors_dir}}MM2-UVW", "descrizione": "Morsetti {{_calc.mors_dir}}mm² motore U-V-W", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "FILO-{{_calc.filo_dir}}MM2-RST", "descrizione": "Filo {{_calc.filo_dir}}mm² linea R-S-T", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
                {"codice": "FILO-{{_calc.filo_dir}}MM2-UVW", "descrizione": "Filo {{_calc.filo_dir}}mm² motore U-V-W", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
            ]
        },
        {
            "id": "MAT_OLEO_ST_BASE",
            "name": "Materiali contattori oleo - Stella-triangolo (base)",
            "version": "3.0", "phase": 2, "priority": 50, "enabled": True,
            "conditions": [
                {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
                {"field": "argano.tipo_avviamento_motore", "operator": "equals", "value": "stella_triangolo"},
                {"field": "_calc.st_disponibile", "operator": "equals", "value": "si"},
            ],
            "materials": [
                {"codice": "CONT-{{_calc.cont_ks}}-KS", "descrizione": "Contattore {{_calc.cont_ks}} stella (KS) - motore {{_calc.potenza_kw}}kW", "quantita": 1, "prezzo_unitario": 0, "categoria": "Contattori"},
                {"codice": "MORS-{{_calc.mors_rst_st}}MM2-RST", "descrizione": "Morsetti {{_calc.mors_rst_st}}mm² linea R-S-T", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "MORS-{{_calc.mors_uvw_st}}MM2-UVW", "descrizione": "Morsetti {{_calc.mors_uvw_st}}mm² motore U-V-W", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "FILO-{{_calc.filo_rst_st}}MM2-RST", "descrizione": "Filo {{_calc.filo_rst_st}}mm² linea R-S-T", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
                {"codice": "FILO-{{_calc.filo_uvw_st}}MM2-UVW", "descrizione": "Filo {{_calc.filo_uvw_st}}mm² motore U-V-W", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
            ]
        },
        {
            "id": "MAT_OLEO_ST_KM",
            "name": "Materiali contattori oleo - Stella-triangolo (KM triangolo)",
            "version": "3.0", "phase": 2, "priority": 51, "enabled": True,
            "conditions": [
                {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
                {"field": "argano.tipo_avviamento_motore", "operator": "equals", "value": "stella_triangolo"},
                {"field": "_calc.st_disponibile", "operator": "equals", "value": "si"},
                {"field": "_calc.cont_km_st", "operator": "not_equals", "value": "NONE"},
            ],
            "materials": [
                {"codice": "CONT-{{_calc.cont_km_st}}-KM", "descrizione": "Contattore {{_calc.cont_km_st}} triangolo (KM) - motore {{_calc.potenza_kw}}kW", "quantita": 1, "prezzo_unitario": 0, "categoria": "Contattori"},
            ]
        },
        {
            "id": "MAT_OLEO_SS_BASE",
            "name": "Materiali soft starter oleo (base)",
            "version": "3.0", "phase": 2, "priority": 50, "enabled": True,
            "conditions": [
                {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
                {"field": "argano.tipo_avviamento_motore", "operator": "equals", "value": "soft_starter"},
                {"field": "_calc.ss_model", "operator": "is_not_empty"},
            ],
            "materials": [
                {"codice": "SS-{{_calc.ss_model}}", "descrizione": "Soft Starter {{_calc.ss_model}} - motore {{_calc.potenza_kw}}kW", "quantita": 1, "prezzo_unitario": 0, "categoria": "Soft Starter"},
            ]
        },
        {
            "id": "MAT_OLEO_SS_CONT",
            "name": "Materiali soft starter oleo (contattore + cablaggio)",
            "version": "3.0", "phase": 2, "priority": 51, "enabled": True,
            "conditions": [
                {"field": "argano.trazione", "operator": "equals", "value": "oleodinamica"},
                {"field": "argano.tipo_avviamento_motore", "operator": "equals", "value": "soft_starter"},
                {"field": "_calc.cont_ss", "operator": "not_equals", "value": "NONE"},
            ],
            "materials": [
                {"codice": "CONT-{{_calc.cont_ss}}-KM", "descrizione": "Contattore {{_calc.cont_ss}} by-pass (KM) - motore {{_calc.potenza_kw}}kW", "quantita": 1, "prezzo_unitario": 0, "categoria": "Contattori"},
                {"codice": "MORS-{{_calc.mors_ss}}MM2-RST", "descrizione": "Morsetti {{_calc.mors_ss}}mm² linea R-S-T", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "MORS-{{_calc.mors_ss}}MM2-UVW", "descrizione": "Morsetti {{_calc.mors_ss}}mm² motore U-V-W", "quantita": 3, "prezzo_unitario": 0, "categoria": "Morsetteria"},
                {"codice": "FILO-{{_calc.filo_ss}}MM2-RST", "descrizione": "Filo {{_calc.filo_ss}}mm² linea R-S-T", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
                {"codice": "FILO-{{_calc.filo_ss}}MM2-UVW", "descrizione": "Filo {{_calc.filo_ss}}mm² motore U-V-W", "quantita": 1, "prezzo_unitario": 0, "unita_misura": "mt", "categoria": "Cablaggio"},
            ]
        },
    ]
    for rule in rules:
        save_rule(rule, f"rule_{rule['id']}.json")
    return len(rules)


# ============================================================================
# MAIN
# ============================================================================

def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "template_contattori_oleo.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ File non trovato: {excel_path}")
        print(f"   Usa: python {sys.argv[0]} percorso/file.xlsx")
        sys.exit(1)
    
    print(f"📖 Lettura: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Trova i 2 fogli (per nome o per posizione)
    sheet_names = wb.sheetnames
    ws_50 = None
    ws_60 = None
    
    for name in sheet_names:
        if "50" in name:
            ws_50 = wb[name]
        elif "60" in name:
            ws_60 = wb[name]
    
    if ws_50 is None and len(sheet_names) >= 1:
        ws_50 = wb[sheet_names[0]]
        print(f"   ⚠️  Foglio 50Hz non trovato per nome, uso primo foglio: '{sheet_names[0]}'")
    if ws_60 is None and len(sheet_names) >= 2:
        ws_60 = wb[sheet_names[1]]
        print(f"   ⚠️  Foglio 60Hz non trovato per nome, uso secondo foglio: '{sheet_names[1]}'")
    
    if ws_50 is None or ws_60 is None:
        print("❌ Servono almeno 2 fogli (50Hz e 60Hz)")
        sys.exit(1)
    
    rows_50 = read_sheet(ws_50)
    rows_60 = read_sheet(ws_60)
    print(f"   50Hz: {len(rows_50)} righe lette")
    print(f"   60Hz: {len(rows_60)} righe lette")
    
    if not rows_50 or not rows_60:
        print("❌ Nessun dato trovato. Controlla che i dati partano dalla riga 4.")
        sys.exit(1)
    
    # Genera regole
    os.makedirs(RULES_DIR, exist_ok=True)
    removed = cleanup_old()
    if removed:
        print(f"🗑️  Rimossi {removed} file regole precedenti")
    
    n_calc = genera_regola_calcolo(rows_50, rows_60)
    n_mat = genera_regole_materiali()
    totale = n_calc + n_mat
    
    print(f"\n✅ Generate {totale} regole in {RULES_DIR}/:")
    print(f"   Phase 1: {n_calc} regola  (CALC_OLEO_CONTATTORI)")
    print(f"   Phase 2: {n_mat} regole  (MAT_OLEO_*)")
    print(f"   Lookup:  {len(rows_50)} righe 50Hz + {len(rows_60)} righe 60Hz")
    
    # Validazione rapida
    lt_file = os.path.join(RULES_DIR, "rule_CALC_OLEO_CONTATTORI.json")
    with open(lt_file) as f:
        lt = json.load(f)
    r50 = lt["actions"][0]["rows"]["50"]
    r60 = lt["actions"][0]["rows"]["60"]
    print(f"\n🔍 Validazione:")
    print(f"   Prima riga 50Hz: {r50[0]['set']['_calc.potenza_kw']}kW "
          f"[{r50[0]['min']}, {r50[0]['max']}) → cont_dir={r50[0]['set'].get('_calc.cont_dir', 'N/A')}")
    print(f"   Ultima riga 50Hz: {r50[-1]['set']['_calc.potenza_kw']}kW "
          f"[{r50[-1]['min']}, {r50[-1]['max']}) → ss={r50[-1]['set'].get('_calc.ss_model', 'N/A')}")
    print(f"   Prima riga 60Hz: {r60[0]['set']['_calc.potenza_kw']}kW "
          f"[{r60[0]['min']}, {r60[0]['max']}) → cont_dir={r60[0]['set'].get('_calc.cont_dir', 'N/A')}")


if __name__ == "__main__":
    main()
